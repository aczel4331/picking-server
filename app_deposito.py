import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import fitz
from PIL import Image, ImageTk
import os
import tempfile
import re
import json
import subprocess
from datetime import datetime
try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
RAILWAY_URL = "https://picking-server-production.up.railway.app"

def obtener_impresoras_windows():
    try:
        resultado = subprocess.run(
            ["powershell", "-Command", "Get-Printer | Select-Object -ExpandProperty Name"],
            capture_output=True, text=True, timeout=10
        )
        impresoras = [l.strip() for l in resultado.stdout.splitlines() if l.strip()]
        return impresoras if impresoras else []
    except Exception:
        return []

def cargar_config():
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"impresora": "", "excel_path": "", "codigo_supervisor": "1234", "servidor_nube": "https://picking-server-production.up.railway.app", "clave_nube": "everest2024"}


def _limpiar_nombre_pasillo(sheet_name):
    """Extrae un nombre corto de pasillo desde el nombre de la hoja."""
    s = sheet_name.strip()
    # Normalizar: quitar espacios extra
    s = re.sub(r'\s+', ' ', s)
    # Si dice PASILLO N algo → "Pasillo N · Algo"
    m = re.match(r'PASILLO\s+(\d+)\s*[-–]?\s*(.*)', s, re.IGNORECASE)
    if m:
        num   = m.group(1)
        resto = m.group(2).strip().title() if m.group(2).strip() else ""
        return f"Pasillo {num}" + (f" · {resto}" if resto else "")
    # Otros nombres especiales
    upper = s.upper()
    if "MOPA" in upper:
        return "Stock Mopa"
    if "MASCOTA" in upper:
        return "Pasillo 7 · Mascotas"
    if "PISO" in upper or "CLIMA" in upper:
        return "Piso · Climatización"
    if "BOMBA" in upper:
        return "Stock Bomba"
    if "CHIQUITAJE" in upper:
        return "Stock Chiquitaje"
    if "ODOO" in upper:
        return ""   # hoja interna, sin pasillo visible
    if "PALLET" in upper:
        return "Pallets"
    return s.title()


def leer_excel_skus(ruta):
    """
    Lee todas las hojas del Excel y devuelve un dict:
        { SKU_UPPER: {"nombre": str, "pasillo": str, "estanteria": str} }
    """
    if not _OPENPYXL_OK or not ruta or not os.path.exists(ruta):
        return {}
    db = {}
    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws      = wb[sheet_name]
            rows    = list(ws.iter_rows(values_only=True))
            pasillo = _limpiar_nombre_pasillo(sheet_name)
            if not rows:
                continue

            # ── Hoja ODOO: nombre(A) | ref_interna(B) — sin estantería ────────
            if "ODOO" in sheet_name.upper():
                for fila in rows:
                    if len(fila) < 2:
                        continue
                    nombre_raw, ref_raw = fila[0], fila[1]
                    if not nombre_raw or not ref_raw:
                        continue
                    nombre = str(nombre_raw).strip()
                    sku    = str(ref_raw).strip().upper()
                    if sku and nombre and sku not in ("REFERENCIA INTERNA", "SKU", "COD"):
                        db.setdefault(sku, {"nombre": nombre, "pasillo": "", "estanteria": ""})
                continue

            # ── Detectar layout de la hoja ────────────────────────────────────
            # Pasillo 1: col0=estanteria, col1=producto, col2=SKU  (3 cols)
            # Pasillo 2+: col0=producto, col1=SKU               (2 cols)
            ncols = max((len(r) for r in rows[:10] if r), default=0)

            col_est  = None
            col_prod = None
            col_sku  = None

            for fila in rows[:15]:
                vals = [str(v).strip().upper() if v else "" for v in fila]
                if "PRODUCTO" in vals and "SKU" in vals:
                    col_prod = vals.index("PRODUCTO")
                    col_sku  = vals.index("SKU")
                    # Si hay una columna antes de PRODUCTO, es estantería
                    if col_prod > 0:
                        col_est = col_prod - 1
                    break
                for i, v in enumerate(vals):
                    if "COD Y DESCRIPCION" in v:
                        col_prod = i
                        col_sku  = None
                        break
                if col_prod is not None:
                    break

            if col_prod is None:
                if ncols >= 3:
                    col_est, col_prod, col_sku = 0, 1, 2
                elif ncols == 2:
                    col_prod, col_sku = 0, 1
                else:
                    continue

            # ── Iterar filas de datos ─────────────────────────────────────────
            estanteria_actual = ""
            for fila in rows:
                if not fila or all(v is None for v in fila):
                    continue

                # Detectar fila de sección de estantería (col0 = "Estantería X")
                if col_est is not None and col_est < len(fila) and fila[col_est]:
                    val_est = str(fila[col_est]).strip().replace("\n", " ")
                    if val_est and "PASILLO" not in val_est.upper():
                        estanteria_actual = val_est

                nombre_raw = fila[col_prod] if col_prod < len(fila) else None
                sku_raw    = fila[col_sku]  if col_sku is not None and col_sku < len(fila) else None

                if not nombre_raw or not sku_raw:
                    continue
                nombre = str(nombre_raw).strip().replace("\n", " ")
                sku    = str(sku_raw).strip().upper()

                # Saltar encabezados
                if sku  in ("SKU", "PRODUCTO", "COD", "CÓDIGO", "CODIGO", ""):
                    continue
                if nombre.upper() in ("PRODUCTO", "SKU", ""):
                    continue
                if len(nombre) < 2:
                    continue

                if sku not in db:
                    db[sku] = {
                        "nombre":     nombre,
                        "pasillo":    pasillo,
                        "estanteria": estanteria_actual,
                    }

        wb.close()
    except Exception:
        pass
    return db

def guardar_config(config: dict):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

# ─────────────────────────────────────────────────────────────────────────────
C = {
    "bg_dark":    "#0F172A",
    "panel":      "#1E293B",
    "card":       "#0F172A",
    "border":     "#334155",
    "accent":     "#3B82F6",
    "accent2":    "#6366F1",
    "success":    "#10B981",
    "warning":    "#F59E0B",
    "danger":     "#EF4444",
    "text_hi":    "#F1F5F9",
    "text_mid":   "#94A3B8",
    "text_lo":    "#475569",
    "preview_bg": "#0B1120",
    "bar_bg":     "#1E3A5F",
    "bar_fg":     "#3B82F6",
    "bar_ok":     "#10B981",
}

FONT_TITLE  = ("Segoe UI Semibold", 11)
FONT_BODY   = ("Segoe UI",          10)
FONT_SMALL  = ("Segoe UI",           9)
FONT_MONO   = ("Consolas",          13, "bold")
FONT_GIANT  = ("Segoe UI Black",    52, "bold")
FONT_BTN    = ("Segoe UI Semibold", 10)


class VentanaConfiguracion(tk.Toplevel):
    def __init__(self, parent, config_actual, callback_guardar):
        super().__init__(parent)
        self.title("Configuracion")
        self.geometry("560x700")
        self.minsize(540, 580)
        self.resizable(True, True)
        self.config(bg=C["bg_dark"])
        self.grab_set()
        self.callback_guardar = callback_guardar
        self._config = dict(config_actual)

        # Header
        hdr = tk.Frame(self, bg=C["accent"], pady=12)
        hdr.pack(fill="x", side="top")
        tk.Label(hdr, text="CONFIGURACION DEL SISTEMA",
                 font=("Segoe UI Semibold", 13), bg=C["accent"], fg="white").pack()

        # Footer (siempre visible)
        footer = tk.Frame(self, bg=C["bg_dark"], padx=20, pady=10)
        footer.pack(fill="x", side="bottom")
        tk.Frame(footer, bg=C["border"], height=1).pack(fill="x", pady=(0,10))
        btn_row = tk.Frame(footer, bg=C["bg_dark"])
        btn_row.pack(fill="x")
        tk.Button(btn_row, text="GUARDAR", font=("Segoe UI Semibold", 10),
                  bg=C["accent"], fg="white", activebackground=C["accent2"],
                  activeforeground="white", relief="flat", cursor="hand2",
                  padx=24, pady=7, bd=0, command=self._guardar).pack(side="right")
        tk.Button(btn_row, text="Cancelar", font=FONT_BODY,
                  bg=C["panel"], fg=C["text_mid"], relief="flat", cursor="hand2",
                  padx=16, pady=7, bd=0, command=self.destroy).pack(side="right", padx=(0,8))

        # Canvas scrollable
        canvas = tk.Canvas(self, bg=C["bg_dark"], highlightthickness=0)
        sb     = tk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        body = tk.Frame(canvas, bg=C["bg_dark"], padx=26, pady=14)
        _win = canvas.create_window((0,0), window=body, anchor="nw")
        body.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(_win, width=e.width))
        canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        def sep():
            tk.Frame(body, bg=C["border"], height=1).pack(fill="x", pady=(14,12))

        def lbl_sec(txt):
            tk.Label(body, text=txt, font=("Segoe UI", 9, "bold"),
                     bg=C["bg_dark"], fg=C["text_mid"]).pack(anchor="w")

        def lbl_sub(txt):
            tk.Label(body, text=txt, font=("Segoe UI", 8),
                     bg=C["bg_dark"], fg=C["text_lo"],
                     justify="left", wraplength=480).pack(anchor="w", pady=(2,0))

        def entry_wrap(color=None):
            w = tk.Frame(body, bg=color or C["border"], padx=2, pady=2)
            w.pack(fill="x", pady=(4,0))
            i = tk.Frame(w, bg=C["card"])
            i.pack(fill="x")
            return i

        # IMPRESORA
        lbl_sec("IMPRESORA DE ETIQUETAS")
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Dark.TCombobox",
                         fieldbackground=C["panel"], background=C["panel"],
                         foreground=C["text_hi"], arrowcolor=C["accent"],
                         selectbackground=C["accent"], selectforeground="white",
                         bordercolor=C["border"], lightcolor=C["border"],
                         darkcolor=C["border"])
        self.combo_var = tk.StringVar()
        self.combo = ttk.Combobox(body, textvariable=self.combo_var,
                                  state="readonly", font=("Segoe UI", 11),
                                  style="Dark.TCombobox")
        self.combo.pack(fill="x", pady=(6,0), ipady=5)
        self.lbl_est = tk.Label(body, text="Buscando impresoras...",
                                bg=C["bg_dark"], fg=C["text_mid"], font=FONT_SMALL)
        self.lbl_est.pack(anchor="w", pady=(4,0))
        self.after(120, lambda: self._cargar(config_actual.get("impresora", "")))

        # EXCEL
        sep()
        lbl_sec("BASE DE DATOS EXCEL  (SKU a Nombre de producto)")
        lbl_sub("Este archivo se carga automaticamente al iniciar la app.")
        excel_row = tk.Frame(body, bg=C["bg_dark"])
        excel_row.pack(fill="x", pady=(6,0))
        self.lbl_excel_ruta = tk.Label(excel_row,
            text=self._resumir_ruta(config_actual.get("excel_path", "")),
            font=("Segoe UI", 9), bg=C["panel"], fg=C["text_hi"],
            relief="flat", anchor="w", padx=8, pady=5, wraplength=340, justify="left")
        self.lbl_excel_ruta.pack(side="left", fill="x", expand=True, padx=(0,8))
        tk.Button(excel_row, text="Buscar", font=FONT_SMALL,
                  bg=C["accent"], fg="white", activebackground=C["accent2"],
                  activeforeground="white", relief="flat", cursor="hand2",
                  padx=10, pady=5, bd=0,
                  command=self._seleccionar_excel).pack(side="right")
        self.lbl_excel_est = tk.Label(body,
            text=self._estado_excel_inicial(config_actual.get("excel_path", "")),
            font=FONT_SMALL, bg=C["bg_dark"],
            fg=C["success"] if config_actual.get("excel_path") and
               os.path.exists(config_actual.get("excel_path","")) else C["text_lo"])
        self.lbl_excel_est.pack(anchor="w", pady=(4,0))

        # SUPERVISOR
        sep()
        lbl_sec("CODIGO DE SUPERVISOR  (para saltear Fase 1)")
        lbl_sub("Para cambiar el codigo, ingresa primero el codigo actual.")
        tk.Label(body, text="Codigo actual", font=("Segoe UI", 8),
                 bg=C["bg_dark"], fg=C["text_lo"]).pack(anchor="w", pady=(6,0))
        act_row = tk.Frame(body, bg=C["bg_dark"])
        act_row.pack(fill="x", pady=(3,6))
        act_wrap = tk.Frame(act_row, bg=C["border"], padx=2, pady=2)
        act_wrap.pack(side="left", fill="x", expand=True, padx=(0,8))
        act_in = tk.Frame(act_wrap, bg=C["card"]); act_in.pack(fill="x")
        self.entry_sup_actual = tk.Entry(act_in, font=("Consolas",12,"bold"),
            justify="center", bg=C["card"], fg=C["text_hi"],
            insertbackground=C["accent"], relief="flat", bd=0, show="*")
        self.entry_sup_actual.pack(fill="x", ipady=5, padx=8)
        self.btn_toggle_actual = tk.Button(act_row, text="ver", font=("Segoe UI",9),
            bg=C["panel"], fg=C["text_mid"], relief="flat", cursor="hand2",
            padx=8, pady=3, bd=0,
            command=lambda: self._toggle_entry(self.entry_sup_actual, self.btn_toggle_actual))
        self.btn_toggle_actual.pack(side="left")

        tk.Label(body, text="Codigo nuevo", font=("Segoe UI", 8),
                 bg=C["bg_dark"], fg=C["text_lo"]).pack(anchor="w")
        new_row = tk.Frame(body, bg=C["bg_dark"])
        new_row.pack(fill="x", pady=(3,0))
        new_wrap = tk.Frame(new_row, bg=C["danger"], padx=2, pady=2)
        new_wrap.pack(side="left", fill="x", expand=True, padx=(0,8))
        new_in = tk.Frame(new_wrap, bg=C["card"]); new_in.pack(fill="x")
        self.entry_sup_nuevo = tk.Entry(new_in, font=("Consolas",12,"bold"),
            justify="center", bg=C["card"], fg=C["text_hi"],
            insertbackground=C["accent"], relief="flat", bd=0, show="*")
        self.entry_sup_nuevo.pack(fill="x", ipady=5, padx=8)
        self.btn_toggle_nuevo = tk.Button(new_row, text="ver", font=("Segoe UI",9),
            bg=C["panel"], fg=C["text_mid"], relief="flat", cursor="hand2",
            padx=8, pady=3, bd=0,
            command=lambda: self._toggle_entry(self.entry_sup_nuevo, self.btn_toggle_nuevo))
        self.btn_toggle_nuevo.pack(side="left")
        self.lbl_sup_error = tk.Label(body, text="", font=("Segoe UI",8),
                                      bg=C["bg_dark"], fg=C["danger"])
        self.lbl_sup_error.pack(anchor="w", pady=(4,0))

        # RAILWAY
        sep()
        lbl_sec("SERVIDOR RAILWAY")
        lbl_sub("Pega la URL de tu deploy en Railway. Los celulares podran acceder desde cualquier red.")
        tk.Label(body, text="URL del servidor (ej: https://picking-xxx.up.railway.app)",
                 font=("Segoe UI", 8, "bold"), bg=C["bg_dark"],
                 fg=C["text_lo"]).pack(anchor="w", pady=(8,0))
        nube_wrap = tk.Frame(body, bg=C["accent"], padx=2, pady=2)
        nube_wrap.pack(fill="x", pady=(4,0))
        nube_in = tk.Frame(nube_wrap, bg=C["card"]); nube_in.pack(fill="x")
        self.entry_nube = tk.Entry(nube_in, font=("Segoe UI", 11),
            bg=C["card"], fg=C["text_hi"],
            insertbackground=C["accent"], relief="flat", bd=0)
        self.entry_nube.insert(0, config_actual.get("servidor_nube", ""))
        self.entry_nube.pack(fill="x", ipady=7, padx=8)
        self.lbl_nube_estado = tk.Label(body, text="", font=("Segoe UI", 8),
                                        bg=C["bg_dark"], fg=C["text_lo"])
        self.lbl_nube_estado.pack(anchor="w", pady=(4,0))
        tk.Button(body, text="Probar conexion",
                  font=FONT_SMALL, bg=C["panel"], fg=C["text_mid"],
                  activebackground=C["border"], activeforeground=C["text_hi"],
                  relief="flat", cursor="hand2", padx=10, pady=4, bd=0,
                  command=self._probar_nube).pack(anchor="w", pady=(4,0))

        tk.Label(body, text="Clave de API  (PICKING_API_KEY en Railway Variables)",
                 font=("Segoe UI", 8, "bold"), bg=C["bg_dark"],
                 fg=C["text_lo"]).pack(anchor="w", pady=(10,0))
        clave_row = tk.Frame(body, bg=C["bg_dark"])
        clave_row.pack(fill="x", pady=(4,0))
        clave_wrap = tk.Frame(clave_row, bg=C["border"], padx=2, pady=2)
        clave_wrap.pack(side="left", fill="x", expand=True, padx=(0,8))
        clave_in = tk.Frame(clave_wrap, bg=C["card"]); clave_in.pack(fill="x")
        self.entry_clave_nube = tk.Entry(clave_in, font=("Segoe UI",10), show="*",
            bg=C["card"], fg=C["text_hi"],
            insertbackground=C["accent"], relief="flat", bd=0)
        self.entry_clave_nube.insert(0, config_actual.get("clave_nube", "everest2024"))
        self.entry_clave_nube.pack(fill="x", ipady=6, padx=8)
        self.btn_toggle_clave = tk.Button(clave_row, text="ver", font=("Segoe UI",9),
            bg=C["panel"], fg=C["text_mid"], relief="flat", cursor="hand2",
            padx=8, pady=3, bd=0,
            command=lambda: self._toggle_entry(self.entry_clave_nube, self.btn_toggle_clave))
        self.btn_toggle_clave.pack(side="left")

        # ── MERCADOLIBRE ──────────────────────────────────────────────────────
        sep()
        lbl_sec("MERCADOLIBRE")
        lbl_sub("Conecta tu cuenta para descargar pedidos y etiquetas automaticamente.")

        ml_frame = tk.Frame(body, bg=C["bg_dark"])
        ml_frame.pack(fill="x", pady=(8,0))

        # Estado de conexión (se actualiza en _cfg_verificar_ml)
        self.lbl_ml_estado_cfg = tk.Label(
            body, text="Verificando conexion...",
            font=("Segoe UI", 9), bg=C["bg_dark"], fg=C["text_lo"])
        self.lbl_ml_estado_cfg.pack(anchor="w", pady=(0,8))

        # Panel de cuentas conectadas (se llena dinámicamente)
        self.frame_cuentas_cfg = tk.Frame(body, bg=C["bg_dark"])
        self.frame_cuentas_cfg.pack(fill="x", pady=(0,6))

        # Botones de acción
        btn_ml_row = tk.Frame(body, bg=C["bg_dark"])
        btn_ml_row.pack(fill="x", pady=(0,4))

        self.btn_cfg_agregar_ml = tk.Button(
            btn_ml_row, text="Agregar cuenta MercadoLibre",
            font=("Segoe UI Semibold", 10),
            bg="#CA8A04", fg="black",
            activebackground="#EAD700", activeforeground="black",
            relief="flat", cursor="hand2", padx=16, pady=7, bd=0,
            command=self._cfg_conectar_ml)
        self.btn_cfg_agregar_ml.pack(side="left")

        tk.Button(btn_ml_row, text="Verificar conexion",
                  font=FONT_SMALL, bg=C["panel"], fg=C["text_mid"],
                  activebackground=C["border"], activeforeground=C["text_hi"],
                  relief="flat", cursor="hand2", padx=10, pady=5, bd=0,
                  command=self._cfg_verificar_ml).pack(side="left", padx=(8,0))

        tk.Button(btn_ml_row, text="💾 Guardar tokens",
                  font=FONT_SMALL, bg="#0891B2", fg="white",
                  activebackground="#0E7490", activeforeground="white",
                  relief="flat", cursor="hand2", padx=10, pady=5, bd=0,
                  command=self._cfg_exportar_tokens).pack(side="left", padx=(8,0))

        # Verificar estado al abrir
        self.after(300, self._cfg_verificar_ml)

        tk.Frame(body, bg=C["bg_dark"], height=16).pack()

    def _cfg_conectar_ml(self):
        """Abre el browser para conectar/agregar una cuenta ML."""
        import webbrowser
        # Determinar el próximo cuenta_id
        n = len(getattr(self, "_cfg_cuentas_actuales", []))
        cuenta_id = f"cuenta_{n}"
        url = RAILWAY_URL.rstrip("/") + f"/auth/login?cuenta={cuenta_id}"
        webbrowser.open(url)
        self.lbl_ml_estado_cfg.config(
            text=f"Se abrio el navegador. Autorizá la cuenta en MercadoLibre y volvé aqui.",
            fg=C["warning"])
        self.after(8000, self._cfg_verificar_ml)

    def _cfg_verificar_ml(self):
        """Verifica el estado de ML y actualiza la UI de configuracion."""
        import threading, urllib.request, json as _j
        def _worker():
            try:
                url = RAILWAY_URL.rstrip("/") + "/auth/status"
                with urllib.request.urlopen(url, timeout=5) as r:
                    d = _j.loads(r.read())
                def _safe_update():
                    try:
                        if self.winfo_exists():
                            self._cfg_actualizar_ml_ui(d)
                    except Exception:
                        pass
                self.after(0, _safe_update)
            except Exception:
                def _safe_err():
                    try:
                        if self.winfo_exists() and \
                           hasattr(self,'lbl_ml_estado_cfg') and \
                           self.lbl_ml_estado_cfg.winfo_exists():
                            self.lbl_ml_estado_cfg.config(
                                text="No se pudo verificar (Railway inactivo?)",
                                fg=C["danger"])
                    except Exception:
                        pass
                self.after(0, _safe_err)
        threading.Thread(target=_worker, daemon=True).start()

    def _cfg_actualizar_ml_ui(self, d):
        """Actualiza el panel de cuentas ML en la ventana de configuración."""
        # Verificar que la ventana sigue abierta
        try:
            if not self.winfo_exists():
                return
            if not hasattr(self, 'frame_cuentas_cfg') or \
               not self.frame_cuentas_cfg.winfo_exists():
                return
        except Exception:
            return

        cuentas = d.get("cuentas", [])
        self._cfg_cuentas_actuales = cuentas

        # Limpiar panel de cuentas
        for w in self.frame_cuentas_cfg.winfo_children():
            w.destroy()

        COLORES = ["#7C3AED", "#0891B2", "#D97706", "#DC2626", "#059669"]

        if cuentas:
            # Mostrar cada cuenta como chip con botón de desconectar
            for idx, cuenta in enumerate(cuentas):
                color = COLORES[idx % len(COLORES)]
                nick  = cuenta["nickname"]
                cid   = cuenta["cuenta_id"]

                chip = tk.Frame(self.frame_cuentas_cfg, bg=color,
                                padx=2, pady=2)
                chip.pack(side="left", padx=(0,6), pady=2)

                inner = tk.Frame(chip, bg=color)
                inner.pack()

                tk.Label(inner, text=f"✅  {nick}",
                         font=("Segoe UI Semibold", 9),
                         bg=color, fg="white",
                         padx=8, pady=4).pack(side="left")

                tk.Button(inner, text="×",
                          font=("Segoe UI", 10, "bold"),
                          bg=color, fg="white",
                          activebackground=C["danger"], activeforeground="white",
                          relief="flat", cursor="hand2", padx=6, pady=4, bd=0,
                          command=lambda c=cid, n=nick: self._cfg_desconectar_cuenta(c, n)
                          ).pack(side="left")

            total_ped = d.get("pedidos", 0)
            self.lbl_ml_estado_cfg.config(
                text=f"✅  {len(cuentas)} cuenta(s) conectada(s)  ·  {total_ped} pedidos en Railway",
                fg=C["success"])
            self.btn_cfg_agregar_ml.config(
                text="+ Agregar otra cuenta ML",
                bg=C["success"], fg="white",
                activebackground="#059669")
        else:
            self.lbl_ml_estado_cfg.config(
                text="Sin cuentas conectadas — hacé clic en 'Conectar MercadoLibre'",
                fg=C["text_mid"])
            self.btn_cfg_agregar_ml.config(
                text="Conectar MercadoLibre",
                bg="#CA8A04", fg="black",
                activebackground="#EAD700")

    def _cfg_desvincular_todo(self):
        """Desconecta TODAS las cuentas ML de Railway."""
        cuentas = getattr(self, "_cfg_cuentas_actuales", [])
        if not cuentas:
            messagebox.showinfo("Sin cuentas",
                                "No hay cuentas conectadas para desconectar.",
                                parent=self)
            return
        nicks = ", ".join(c["nickname"] for c in cuentas)
        if not messagebox.askyesno(
                "Desvincular todas las cuentas",
                f"¿Desvincular TODAS las cuentas?\n\n"
                f"Cuentas: {nicks}\n\n"
                f"Tendrás que volver a conectar MercadoLibre para traer pedidos.",
                parent=self):
            return

        import threading, urllib.request
        self.lbl_ml_estado_cfg.config(
            text="Desvinculando cuentas...", fg=C["warning"])
        self.btn_cfg_agregar_ml.config(state="disabled")

        def _w():
            errores = []
            for c in cuentas:
                try:
                    url = RAILWAY_URL.rstrip("/")
                    req = urllib.request.Request(
                        f"{url}/api/cuentas/{c['cuenta_id']}/logout",
                        method="POST",
                        headers={"Content-Type": "application/json",
                                 "X-API-Key": "everest2024"})
                    urllib.request.urlopen(req, timeout=8)
                except Exception as e:
                    errores.append(str(e))

            def _actualizar_ui():
                try:
                    if not self.winfo_exists():
                        return
                    # Limpiar chips de cuentas
                    if hasattr(self, "frame_cuentas_cfg") and \
                       self.frame_cuentas_cfg.winfo_exists():
                        for w in self.frame_cuentas_cfg.winfo_children():
                            w.destroy()
                    # Actualizar estado
                    self._cfg_cuentas_actuales = []
                    self.lbl_ml_estado_cfg.config(
                        text="✅ Todas las cuentas desvinculadas correctamente",
                        fg=C["success"])
                    self.btn_cfg_agregar_ml.config(
                        text="Conectar MercadoLibre",
                        bg="#CA8A04", fg="black",
                        state="normal")
                    # Actualizar también el botón principal del panel ML
                    if hasattr(self.master, 'btn_ml_login'):
                        self.master.btn_ml_login.config(
                            text="Conectar MercadoLibre",
                            bg="#CA8A04", fg="black")
                except Exception:
                    pass

            self.after(0, _actualizar_ui)

        threading.Thread(target=_w, daemon=True).start()

    def _cfg_exportar_tokens(self):
        """
        Obtiene el JSON de tokens de Railway y abre ventana para copiarlo.
        Si Railway no tiene el endpoint (version vieja), genera el JSON localmente
        consultando /auth/status.
        """
        import threading, urllib.request, json as _j

        def _worker():
            url_base = RAILWAY_URL.rstrip("/")

            # Intentar el endpoint nuevo primero
            try:
                req = urllib.request.Request(
                    url_base + "/api/tokens_export",
                    headers={"X-API-Key": "everest2024"})
                with urllib.request.urlopen(req, timeout=8) as r:
                    d = _j.loads(r.read())
                self.after(0, lambda: self._cfg_mostrar_tokens(d))
                return
            except urllib.error.HTTPError as e:
                if e.code != 404:
                    body = e.read().decode("utf-8","ignore") if e.fp else str(e)
                    self.after(0, lambda: messagebox.showerror(
                        "Error", f"HTTP {e.code}:\n{body[:200]}", parent=self))
                    return
                # 404 = Railway tiene version vieja, continuar con fallback

            except Exception as e:
                self.after(0, lambda: messagebox.showerror(
                    "Error de conexion", str(e), parent=self))
                return

            # Fallback: usar /auth/status para mostrar instrucciones manuales
            try:
                with urllib.request.urlopen(
                        url_base + "/auth/status", timeout=6) as r:
                    st = _j.loads(r.read())
                cuentas = st.get("cuentas", [])
                if not cuentas:
                    self.after(0, lambda: messagebox.showwarning(
                        "Sin cuentas",
                        "No hay cuentas conectadas a MercadoLibre.\n"
                        "Conectá primero desde la pestaña Pedidos ML.",
                        parent=self))
                    return

                # Mostrar instruccion para subir server.py nuevo
                msg = (
                    "Railway tiene una versión antigua del servidor.\n\n"
                    "Para guardar los tokens necesitás:\n\n"
                    "1. Subir el server.py nuevo a GitHub\n"
                    "2. Esperar que Railway redeplye\n"
                    "3. Volver a presionar 'Guardar tokens'\n\n"
                    f"Cuentas conectadas ahora: "
                    f"{', '.join(c.get('nickname','?') for c in cuentas)}\n\n"
                    "¿Querés abrir Railway para actualizar?"
                )
                resp = self.after(0, lambda: None)
                def _preguntar():
                    if messagebox.askyesno(
                            "Servidor desactualizado", msg, parent=self):
                        __import__("webbrowser").open("https://railway.app")
                self.after(0, _preguntar)

            except Exception as e:
                self.after(0, lambda: messagebox.showerror(
                    "Error", str(e), parent=self))

        threading.Thread(target=_worker, daemon=True).start()

    def _cfg_mostrar_tokens(self, d):
        """Muestra ventana con el JSON de tokens para copiar a Railway."""
        if not d.get("ok"):
            messagebox.showwarning(
                "Sin tokens",
                d.get("msg", "No hay cuentas conectadas."),
                parent=self)
            return

        tokens_json = d.get("ML_TOKENS_JSON", "")
        cuentas     = d.get("cuentas", [])

        win = tk.Toplevel(self)
        win.title("Guardar tokens en Railway")
        win.geometry("620x480")
        win.resizable(False, False)
        win.config(bg=C["bg_dark"])
        win.grab_set()

        # Header
        hdr = tk.Frame(win, bg="#0891B2", pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="💾  Guardar tokens ML en Railway",
                 font=("Segoe UI Black", 12), bg="#0891B2",
                 fg="white").pack()

        body = tk.Frame(win, bg=C["bg_dark"], padx=20, pady=16)
        body.pack(fill="both", expand=True)

        # Cuentas conectadas
        cuentas_txt = "  ·  ".join(
            f"{c['nickname']} ({c['expira']})" for c in cuentas)
        tk.Label(body, text=f"Cuentas: {cuentas_txt}",
                 font=("Segoe UI Semibold", 9), bg=C["bg_dark"],
                 fg=C["success"]).pack(anchor="w", pady=(0,10))

        # Instrucciones
        pasos = (
            "1. Clic en 'Copiar JSON'\n"
            "2. Ir a railway.app → tu proyecto → picking-server → Variables\n"
            "3. Buscar o crear la variable  ML_TOKENS_JSON\n"
            "4. Pegar el valor copiado\n"
            "5. Guardar — listo, los tokens sobreviven cualquier redeploy"
        )
        tk.Label(body, text=pasos,
                 font=("Segoe UI", 9), bg=C["bg_dark"],
                 fg=C["text_mid"], justify="left").pack(anchor="w", pady=(0,10))

        # JSON box
        tk.Label(body, text="Valor para ML_TOKENS_JSON:",
                 font=("Segoe UI", 8, "bold"), bg=C["bg_dark"],
                 fg=C["text_lo"]).pack(anchor="w")

        txt_frame = tk.Frame(body, bg=C["border"], padx=1, pady=1)
        txt_frame.pack(fill="x", pady=(4,0))
        txt = tk.Text(txt_frame, height=6, font=("Consolas", 8),
                      bg=C["card"], fg=C["text_hi"], relief="flat",
                      wrap="word", state="normal")
        txt.insert("1.0", tokens_json)
        txt.config(state="disabled")
        txt.pack(fill="x")

        # Botones
        sep = tk.Frame(body, bg=C["border"], height=1)
        sep.pack(fill="x", pady=(14,10))

        btn_row = tk.Frame(body, bg=C["bg_dark"])
        btn_row.pack(fill="x")

        def _copiar():
            win.clipboard_clear()
            win.clipboard_append(tokens_json)
            btn_copy.config(text="✅ Copiado!", bg=C["success"])
            win.after(2000, lambda: btn_copy.config(
                text="📋 Copiar JSON", bg="#0891B2"))

        btn_copy = tk.Button(btn_row, text="📋 Copiar JSON",
                             font=("Segoe UI Semibold", 10),
                             bg="#0891B2", fg="white",
                             activebackground="#0E7490",
                             relief="flat", cursor="hand2",
                             padx=16, pady=7, bd=0,
                             command=_copiar)
        btn_copy.pack(side="left")

        tk.Button(btn_row, text="Abrir Railway Variables",
                  font=FONT_SMALL, bg=C["panel"], fg=C["text_mid"],
                  activebackground=C["border"],
                  relief="flat", cursor="hand2", padx=12, pady=5, bd=0,
                  command=lambda: __import__("webbrowser").open(
                      "https://railway.app")).pack(side="left", padx=(8,0))

        tk.Button(btn_row, text="Cerrar",
                  font=FONT_SMALL, bg=C["panel"], fg=C["text_mid"],
                  relief="flat", cursor="hand2", padx=12, pady=5, bd=0,
                  command=win.destroy).pack(side="right")
        """Desconecta una cuenta ML específica desde la ventana de configuración."""
        if not messagebox.askyesno(
                "Desconectar",
                f"¿Desconectar la cuenta '{nick}'?\n\n"
                f"Podés volver a conectarla cuando quieras.",
                parent=self):
            return

        import threading, urllib.request
        self.lbl_ml_estado_cfg.config(
            text=f"Desconectando {nick}...", fg=C["warning"])

        def _w():
            try:
                url = RAILWAY_URL.rstrip("/")
                req = urllib.request.Request(
                    f"{url}/api/cuentas/{cuenta_id}/logout",
                    method="POST",
                    headers={"Content-Type": "application/json",
                             "X-API-Key": "everest2024"})
                urllib.request.urlopen(req, timeout=8)
            except Exception:
                pass

            def _actualizar():
                try:
                    if self.winfo_exists():
                        self._cfg_verificar_ml()
                except Exception:
                    pass
            self.after(500, _actualizar)

        threading.Thread(target=_w, daemon=True).start()

    def _resumir_ruta(self, ruta):
        if not ruta:
            return "Sin archivo configurado"
        return os.path.basename(ruta)

    def _estado_excel_inicial(self, ruta):
        if not ruta:
            return "No configurado"
        if not os.path.exists(ruta):
            return "Archivo no encontrado en la ruta guardada."
        return "Archivo configurado y disponible."

    def _seleccionar_excel(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar base de datos Excel",
            filetypes=[("Archivos Excel", "*.xlsx *.xls *.xlsm")],
            parent=self)
        if not ruta:
            return
        self._config["excel_path"] = ruta
        self.lbl_excel_ruta.config(text=self._resumir_ruta(ruta))
        db = leer_excel_skus(ruta)
        if db:
            self.lbl_excel_est.config(
                text=f"OK  {len(db)} productos encontrados.",
                fg=C["success"])
        else:
            self.lbl_excel_est.config(
                text="No se encontraron SKUs en el archivo.",
                fg=C["warning"])

    def _probar_nube(self):
        import urllib.request, json as _j, threading
        url = self.entry_nube.get().strip().rstrip("/")
        if not url:
            self.lbl_nube_estado.config(text="Ingresa la URL primero.", fg=C["warning"])
            return
        # Auto-agregar https:// si falta el protocolo
        if url and not url.startswith("http"):
            url = "https://" + url
            self.entry_nube.delete(0, tk.END)
            self.entry_nube.insert(0, url)
        self.lbl_nube_estado.config(text="Probando conexion...", fg=C["text_mid"])
        self.update_idletasks()

        def _worker():
            try:
                with urllib.request.urlopen(url + "/api/ping", timeout=8) as r:
                    d = _j.loads(r.read())
                if d.get("ok"):
                    self.after(0, lambda: self.lbl_nube_estado.config(
                        text=f"Conectado correctamente  {d.get('ts','')}",
                        fg=C["success"]))
                else:
                    self.after(0, lambda: self.lbl_nube_estado.config(
                        text="El servidor respondio con error.", fg=C["warning"]))
            except Exception as e:
                self.after(0, lambda: self.lbl_nube_estado.config(
                    text=f"No se pudo conectar: {e}", fg=C["danger"]))

        threading.Thread(target=_worker, daemon=True).start()

    def _cargar(self, sel):
        imp = obtener_impresoras_windows()
        if imp:
            self.combo["values"] = imp
            self.combo_var.set(sel if sel in imp else imp[0])
            self.lbl_est.config(text=f"{len(imp)} impresora(s) encontradas.", fg=C["success"])
        else:
            self.combo["values"] = ["(Sin impresoras detectadas)"]
            self.combo.current(0)
            self.lbl_est.config(text="No se detectaron impresoras.", fg=C["warning"])

    def _toggle_entry(self, entry, btn):
        show = entry.cget("show")
        entry.config(show="" if show else "*")
        if hasattr(btn, "config"):
            btn.config(fg=C["accent"] if show else C["text_mid"])

    def _guardar(self):
        imp = self.combo_var.get()
        if not imp or imp.startswith("("):
            messagebox.showwarning("Atencion", "Selecciona una impresora valida.", parent=self)
            return

        codigo_actual_guardado = self._config.get("codigo_supervisor", "1234")
        actual_ingresado       = self.entry_sup_actual.get().strip()
        nuevo_ingresado        = self.entry_sup_nuevo.get().strip()

        if actual_ingresado or nuevo_ingresado:
            if not actual_ingresado:
                self.lbl_sup_error.config(text="Ingresa el codigo actual primero.")
                self.entry_sup_actual.focus(); return
            if actual_ingresado != codigo_actual_guardado:
                self.lbl_sup_error.config(text="El codigo actual es incorrecto.")
                self.entry_sup_actual.delete(0, tk.END)
                self.entry_sup_actual.config(bg=C["danger"])
                self.after(400, lambda: self.entry_sup_actual.config(bg=C["card"]))
                self.entry_sup_actual.focus(); return
            if not nuevo_ingresado:
                self.lbl_sup_error.config(text="Ingresa el nuevo codigo.")
                self.entry_sup_nuevo.focus(); return
            if len(nuevo_ingresado) < 4:
                self.lbl_sup_error.config(text="Minimo 4 caracteres.")
                self.entry_sup_nuevo.focus(); return
            self._config["codigo_supervisor"] = nuevo_ingresado

        nube_url = self.entry_nube.get().strip().rstrip("/")
        if nube_url and not nube_url.startswith("http"):
            nube_url = "https://" + nube_url
        self._config["impresora"]     = imp
        self._config["servidor_nube"] = nube_url
        self._config["clave_nube"]    = self.entry_clave_nube.get().strip()
        self.callback_guardar(self._config)
        messagebox.showinfo("Guardado", "Configuracion guardada.", parent=self)
        self.destroy()


class VentanaBaseSKUs(tk.Toplevel):
    """
    Permite ver, agregar, editar y eliminar entradas de la base interna de SKUs.
    Los datos se guardan en skus_db.json al lado del script.
    """
    @staticmethod
    def get_db_path():
        import sys
        base = os.path.dirname(sys.executable) if getattr(sys,'frozen',False) else os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base, "skus_db.json")

    def __init__(self, parent, callback_actualizar):
        super().__init__(parent)
        self.title("Base de datos de SKUs")
        self.geometry("820x560")
        self.minsize(700, 420)
        self.config(bg=C["bg_dark"])
        self.grab_set()
        self.callback_actualizar = callback_actualizar
        self.db = self._cargar_db()
        self._build()

    # ── Persistencia ──────────────────────────────────────────────────────────
    def _cargar_db(self):
        if os.path.exists(self.get_db_path()):
            try:
                with open(self.get_db_path(), "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return {}

    def _guardar_db(self):
        with open(self.get_db_path(), "w", encoding="utf-8") as f:
            json.dump(self.db, f, ensure_ascii=False, indent=2)

    # ── UI ─────────────────────────────────────────────────────────────────────
    def _build(self):
        # Header
        hdr = tk.Frame(self, bg=C["accent2"], pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="BASE DE DATOS INTERNA DE SKUs",
                 font=("Segoe UI Black", 12), bg=C["accent2"], fg="white").pack()
        tk.Label(hdr, text="Estos datos tienen prioridad sobre el Excel externo",
                 font=("Segoe UI", 8), bg=C["accent2"], fg="#C7D2FE").pack()

        # Toolbar
        toolbar = tk.Frame(self, bg=C["panel"], padx=14, pady=10)
        toolbar.pack(fill="x")

        self.var_buscar = tk.StringVar()
        self.var_buscar.trace_add("write", lambda *_: self._filtrar_debounced())
        buscar_wrap = tk.Frame(toolbar, bg=C["border"], padx=1, pady=1)
        buscar_wrap.pack(side="left", fill="x", expand=True, padx=(0, 10))
        tk.Entry(buscar_wrap, textvariable=self.var_buscar,
                 font=FONT_BODY, bg=C["card"], fg=C["text_hi"],
                 insertbackground=C["accent"], relief="flat", bd=0).pack(
                 fill="x", ipady=5, padx=6)

        tk.Button(toolbar, text="+ Agregar",
                  font=FONT_BTN, bg=C["success"], fg="white",
                  activebackground="#059669", activeforeground="white",
                  relief="flat", cursor="hand2", padx=14, pady=5, bd=0,
                  command=self._form_nuevo).pack(side="right")
        tk.Button(toolbar, text="Importar Excel",
                  font=FONT_BTN, bg=C["accent"], fg="white",
                  activebackground=C["accent2"], activeforeground="white",
                  relief="flat", cursor="hand2", padx=14, pady=5, bd=0,
                  command=self._importar_excel).pack(side="right", padx=(0, 8))

        # Cabecera de tabla fija
        tabla_frame = tk.Frame(self, bg=C["bg_dark"])
        tabla_frame.pack(fill="both", expand=True, padx=10, pady=(0, 0))

        hdr_t = tk.Frame(tabla_frame, bg=C["bar_bg"])
        hdr_t.pack(fill="x")
        for texto, ancho in [("SKU", 14), ("Nombre del producto", 36),
                              ("Pasillo", 18), ("Estanteria", 14), ("", 6)]:
            tk.Label(hdr_t, text=texto, font=("Segoe UI", 9, "bold"),
                     bg=C["bar_bg"], fg=C["accent"],
                     width=ancho, anchor="w").pack(side="left", padx=4, pady=5)

        # Canvas scrollable
        canvas = tk.Canvas(tabla_frame, bg=C["bg_dark"], highlightthickness=0)
        sb = tk.Scrollbar(tabla_frame, orient="vertical", command=canvas.yview)
        self.frame_filas = tk.Frame(canvas, bg=C["bg_dark"])
        canvas.create_window((0, 0), window=self.frame_filas, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        self.frame_filas.bind("<Configure>",
                              lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        self._canvas = canvas

        # Paginación
        self._PAGE_SIZE = 80
        self._page      = 0
        self._items_filtrados = []

        pag_frame = tk.Frame(self, bg=C["panel"], padx=14, pady=6)
        pag_frame.pack(fill="x")
        self.lbl_total = tk.Label(pag_frame, text="", font=FONT_SMALL,
                                   bg=C["panel"], fg=C["text_mid"])
        self.lbl_total.pack(side="left")
        self.btn_prev = tk.Button(pag_frame, text="< Anterior",
                  font=FONT_SMALL, bg=C["card"], fg=C["text_mid"],
                  relief="flat", cursor="hand2", padx=10, pady=3, bd=0,
                  command=self._pag_prev)
        self.btn_prev.pack(side="left", padx=(10, 4))
        self.lbl_pag = tk.Label(pag_frame, text="", font=FONT_SMALL,
                                 bg=C["panel"], fg=C["text_hi"])
        self.lbl_pag.pack(side="left")
        self.btn_next = tk.Button(pag_frame, text="Siguiente >",
                  font=FONT_SMALL, bg=C["card"], fg=C["text_mid"],
                  relief="flat", cursor="hand2", padx=10, pady=3, bd=0,
                  command=self._pag_next)
        self.btn_next.pack(side="left", padx=(4, 0))

        # Footer
        footer = tk.Frame(self, bg=C["panel"], padx=14, pady=10)
        footer.pack(fill="x")
        tk.Button(footer, text="Cerrar y aplicar",
                  font=("Segoe UI Semibold", 10), bg=C["success"], fg="white",
                  activebackground="#059669", activeforeground="white",
                  relief="flat", cursor="hand2", padx=20, pady=7, bd=0,
                  command=self._cerrar_aplicar).pack(side="right")

        # Carga inicial diferida — no bloquea la apertura de la ventana
        self.after(50, self._filtrar)

    def _filtrar_debounced(self):
        """Espera 300ms después del último keystroke antes de filtrar."""
        if hasattr(self, '_filtrar_timer'):
            self.after_cancel(self._filtrar_timer)
        self._filtrar_timer = self.after(300, self._filtrar)

    def _filtrar(self):
        """Filtra los items y muestra la primera página. No bloquea la UI."""
        q = self.var_buscar.get().strip().upper()
        items = sorted(self.db.items())
        if q:
            items = [(k, v) for k, v in items
                     if q in k.upper()
                     or q in v.get("nombre", "").upper()
                     or q in v.get("pasillo", "").upper()
                     or q in v.get("estanteria", "").upper()]
        self._items_filtrados = items
        self._page = 0
        self._render_pagina()

    def _render_pagina(self):
        """Renderiza solo la página actual (PAGE_SIZE filas)."""
        # Destruir filas anteriores
        for w in self.frame_filas.winfo_children():
            w.destroy()

        items  = self._items_filtrados
        total  = len(items)
        start  = self._page * self._PAGE_SIZE
        end    = min(start + self._PAGE_SIZE, total)
        pagina = items[start:end]

        bg_alt = [C["card"], C["bg_dark"]]
        for i, (sku, datos) in enumerate(pagina):
            bg   = bg_alt[i % 2]
            fila = tk.Frame(self.frame_filas, bg=bg)
            fila.pack(fill="x")

            def _lbl(text, width, mono=False):
                f = ("Consolas", 9) if mono else FONT_SMALL
                tk.Label(fila, text=text, font=f, bg=bg,
                         fg=C["text_hi"], width=width,
                         anchor="w", wraplength=200).pack(side="left", padx=4, pady=4)

            _lbl(sku, 14, mono=True)
            _lbl(datos.get("nombre", ""), 36)
            _lbl(datos.get("pasillo", ""), 18)
            _lbl(datos.get("estanteria", ""), 14)

            bf = tk.Frame(fila, bg=bg)
            bf.pack(side="right", padx=4)
            tk.Button(bf, text="editar", font=("Segoe UI", 8),
                      bg=bg, fg=C["accent"], relief="flat", cursor="hand2", bd=0,
                      command=lambda s=sku: self._form_editar(s)).pack(side="left")
            tk.Button(bf, text="borrar", font=("Segoe UI", 8),
                      bg=bg, fg=C["danger"], relief="flat", cursor="hand2", bd=0,
                      command=lambda s=sku: self._eliminar(s)).pack(side="left", padx=(4,0))

        # Actualizar paginación
        total_pag = max(1, (total + self._PAGE_SIZE - 1) // self._PAGE_SIZE)
        self.lbl_total.config(
            text=f"{total} registros  (mostrando {start+1}-{end})" if total else "Sin resultados")
        self.lbl_pag.config(text=f"Página {self._page+1} / {total_pag}")
        self.btn_prev.config(state="normal" if self._page > 0 else "disabled")
        self.btn_next.config(state="normal" if end < total else "disabled")
        self._canvas.yview_moveto(0)

    def _pag_prev(self):
        if self._page > 0:
            self._page -= 1
            self._render_pagina()

    def _pag_next(self):
        total = len(self._items_filtrados)
        if (self._page + 1) * self._PAGE_SIZE < total:
            self._page += 1
            self._render_pagina()

    def _form_nuevo(self):
        self._form_editar(None)

    def _form_editar(self, sku_orig):
        win = tk.Toplevel(self)
        win.title("Nuevo SKU" if sku_orig is None else f"Editar: {sku_orig}")
        win.geometry("440x380")
        win.minsize(440, 380)
        win.resizable(False, False)
        win.config(bg=C["bg_dark"])
        win.grab_set()

        datos = self.db.get(sku_orig, {}) if sku_orig else {}

        fields = [("SKU *", "sku", sku_orig or ""),
                  ("Nombre del producto *", "nombre", datos.get("nombre", "")),
                  ("Pasillo", "pasillo", datos.get("pasillo", "")),
                  ("Estantería / Ubicación", "estanteria", datos.get("estanteria", ""))]

        vars_ = {}

        # Footer fijo abajo — se empaqueta PRIMERO para que nunca quede oculto
        footer = tk.Frame(win, bg=C["bg_dark"], padx=20, pady=12)
        footer.pack(side="bottom", fill="x")

        lbl_err = tk.Label(footer, text="", font=FONT_SMALL,
                           bg=C["bg_dark"], fg=C["danger"])
        lbl_err.pack(anchor="w", pady=(0, 6))

        sep = tk.Frame(footer, bg=C["border"], height=1)
        sep.pack(fill="x", pady=(0, 10))

        btn_row = tk.Frame(footer, bg=C["bg_dark"])
        btn_row.pack(fill="x")

        def _guardar():
            sku_nuevo = vars_["sku"].get().strip().upper()
            nombre    = vars_["nombre"].get().strip()
            if not sku_nuevo:
                lbl_err.config(text="⚠  El SKU es obligatorio."); return
            if not nombre:
                lbl_err.config(text="⚠  El nombre es obligatorio."); return
            if sku_orig and sku_orig != sku_nuevo and sku_orig in self.db:
                del self.db[sku_orig]
            self.db[sku_nuevo] = {
                "nombre":     nombre,
                "pasillo":    vars_["pasillo"].get().strip(),
                "estanteria": vars_["estanteria"].get().strip(),
            }
            self._guardar_db()
            win.destroy()
            self._filtrar()

        tk.Button(btn_row, text="✅  GUARDAR",
                  font=("Segoe UI Semibold", 10),
                  bg=C["success"], fg="white",
                  activebackground="#059669", activeforeground="white",
                  relief="flat", cursor="hand2",
                  padx=22, pady=8, bd=0,
                  command=_guardar).pack(side="right")
        tk.Button(btn_row, text="Cancelar", font=FONT_BODY,
                  bg=C["panel"], fg=C["text_mid"],
                  relief="flat", cursor="hand2",
                  padx=14, pady=8, bd=0,
                  command=win.destroy).pack(side="right", padx=(0, 8))

        # Campos — en el espacio restante
        body = tk.Frame(win, bg=C["bg_dark"], padx=20, pady=14)
        body.pack(side="top", fill="both", expand=True)

        for label, key, val in fields:
            tk.Label(body, text=label, font=("Segoe UI", 8, "bold"),
                     bg=C["bg_dark"], fg=C["text_mid"]).pack(anchor="w", pady=(6, 2))
            var = tk.StringVar(value=val)
            vars_[key] = var
            wrap = tk.Frame(body, bg=C["border"], padx=1, pady=1)
            wrap.pack(fill="x")
            e = tk.Entry(wrap, textvariable=var, font=FONT_BODY,
                         bg=C["card"], fg=C["text_hi"],
                         insertbackground=C["accent"], relief="flat", bd=0)
            e.pack(fill="x", ipady=6, padx=6)
            if key == "sku" and sku_orig:
                e.config(state="disabled")
            # Enter en cualquier campo guarda
            e.bind("<Return>", lambda ev: _guardar())

    def _eliminar(self, sku):
        if messagebox.askyesno("Confirmar", f"¿Eliminar '{sku}'?", parent=self):
            del self.db[sku]
            self._guardar_db()
            self._filtrar()

    def _importar_excel(self):
        if not _OPENPYXL_OK:
            messagebox.showerror("Módulo faltante",
                                 "Instalá openpyxl:\npip install openpyxl", parent=self)
            return
        ruta = filedialog.askopenfilename(
            title="Seleccionar Excel",
            filetypes=[("Excel", "*.xlsx *.xls *.xlsm")], parent=self)
        if not ruta:
            return
        db_excel = leer_excel_skus(ruta)
        antes = len(self.db)
        for sku, datos in db_excel.items():
            if isinstance(datos, dict):
                self.db.setdefault(sku, datos)
            else:
                self.db.setdefault(sku, {"nombre": datos, "pasillo": "", "estanteria": ""})
        self._guardar_db()
        nuevos = len(self.db) - antes
        messagebox.showinfo("Importado",
                            f"✅  {nuevos} SKUs nuevos importados desde Excel.\n"
                            f"Los SKUs ya existentes no fueron modificados.",
                            parent=self)
        self._filtrar()

    def _cerrar_aplicar(self):
        self.callback_actualizar(self.db)
        self.destroy()


# =============================================================================
# VENTANA DE AUTORIZACIÓN DE SUPERVISOR
# =============================================================================
class VentanaCodigoSupervisor(tk.Toplevel):
    """
    Pide el código de supervisor para autorizar el pasaje a Fase 2
    cuando la Fase 1 (colecta) no fue completada.
    Llama a callback_ok() si el código es correcto.
    """
    def __init__(self, parent, codigo_correcto, incompletos, callback_ok):
        super().__init__(parent)
        self.title("Autorización requerida")
        self.geometry("480x420")
        self.resizable(False, False)
        self.config(bg=C["bg_dark"])
        self.grab_set()
        self.codigo_correcto = str(codigo_correcto).strip()
        self.callback_ok     = callback_ok
        self._intentos       = 0
        self._MAX_INTENTOS   = 3

        # ── Header rojo de advertencia ────────────────────────────────────────
        hdr = tk.Frame(self, bg=C["danger"], pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🔒  AUTORIZACIÓN DE SUPERVISOR",
                 font=("Segoe UI Black", 12), bg=C["danger"], fg="white").pack()
        tk.Label(hdr, text="La Fase 1 no fue completada",
                 font=("Segoe UI", 9), bg=C["danger"], fg="#FECACA").pack(pady=(2, 0))

        body = tk.Frame(self, bg=C["bg_dark"], padx=28, pady=16)
        body.pack(fill="both", expand=True)

        # ── Lista de artículos incompletos ────────────────────────────────────
        tk.Label(body, text="Artículos pendientes de colecta:",
                 font=("Segoe UI", 9, "bold"), bg=C["bg_dark"],
                 fg=C["text_mid"]).pack(anchor="w")

        lista_frame = tk.Frame(body, bg=C["panel"], padx=10, pady=8)
        lista_frame.pack(fill="x", pady=(4, 12))

        mostrar = incompletos[:6]
        for sku, colectado, requerido in mostrar:
            fila = tk.Frame(lista_frame, bg=C["panel"])
            fila.pack(fill="x", pady=1)
            tk.Label(fila, text=f"• {sku}",
                     font=("Consolas", 9), bg=C["panel"],
                     fg=C["text_hi"], anchor="w").pack(side="left")
            tk.Label(fila, text=f"{colectado}/{requerido}",
                     font=("Segoe UI Semibold", 9), bg=C["panel"],
                     fg=C["danger"], anchor="e").pack(side="right")

        if len(incompletos) > 6:
            tk.Label(lista_frame,
                     text=f"  … y {len(incompletos) - 6} artículo(s) más sin colectar",
                     font=("Segoe UI", 8), bg=C["panel"],
                     fg=C["text_lo"]).pack(anchor="w", pady=(4, 0))

        # ── Campo de código ───────────────────────────────────────────────────
        tk.Label(body, text="CÓDIGO DE SUPERVISOR",
                 font=("Segoe UI", 9, "bold"), bg=C["bg_dark"],
                 fg=C["text_mid"]).pack(anchor="w")

        entry_wrap = tk.Frame(body, bg=C["danger"], padx=2, pady=2)
        entry_wrap.pack(fill="x", pady=(6, 0))
        inner = tk.Frame(entry_wrap, bg=C["card"])
        inner.pack(fill="x")

        self.entry_codigo = tk.Entry(
            inner, font=("Consolas", 18, "bold"), justify="center",
            bg=C["card"], fg=C["text_hi"], insertbackground=C["accent"],
            relief="flat", bd=0, show="●")
        self.entry_codigo.pack(fill="x", ipady=10, padx=10)
        self.entry_codigo.bind("<Return>", lambda e: self._verificar())
        self.entry_codigo.focus()

        self.lbl_error = tk.Label(
            body, text="",
            font=("Segoe UI Semibold", 9), bg=C["bg_dark"], fg=C["danger"])
        self.lbl_error.pack(pady=(6, 0))

        # ── Botones ───────────────────────────────────────────────────────────
        sep = tk.Frame(body, bg=C["border"], height=1)
        sep.pack(fill="x", pady=(12, 10))

        row = tk.Frame(body, bg=C["bg_dark"])
        row.pack(fill="x")

        tk.Button(row, text="✔  AUTORIZAR", font=("Segoe UI Semibold", 10),
                  bg=C["danger"], fg="white",
                  activebackground="#B91C1C", activeforeground="white",
                  relief="flat", cursor="hand2", padx=24, pady=7, bd=0,
                  command=self._verificar).pack(side="right")

        tk.Button(row, text="Cancelar", font=FONT_BODY,
                  bg=C["panel"], fg=C["text_mid"],
                  activebackground=C["border"], activeforeground=C["text_hi"],
                  relief="flat", cursor="hand2", padx=16, pady=7, bd=0,
                  command=self.destroy).pack(side="right", padx=(0, 8))

    def _verificar(self):
        ingresado = self.entry_codigo.get().strip()
        if ingresado == self.codigo_correcto:
            self.destroy()
            self.callback_ok()
        else:
            self._intentos += 1
            restantes = self._MAX_INTENTOS - self._intentos
            if restantes <= 0:
                self.lbl_error.config(
                    text="❌  Demasiados intentos fallidos. Acceso bloqueado.")
                self.entry_codigo.config(state="disabled")
                self.after(2000, self.destroy)
            else:
                self.lbl_error.config(
                    text=f"❌  Código incorrecto. {restantes} intento(s) restante(s).")
                self.entry_codigo.delete(0, tk.END)
                # Parpadeo rojo en el entry
                self.entry_codigo.config(bg=C["danger"])
                self.after(300, lambda: self.entry_codigo.config(bg=C["card"]))
                self.entry_codigo.focus()


# =============================================================================
# APLICACIÓN PRINCIPAL
# =============================================================================
class AsistenteDepositoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Picking — Pro Edition")
        self.root.geometry("1400x800")
        self.root.minsize(1000, 600)
        self.root.config(bg=C["bg_dark"])

        self.ruta_pdf          = ""
        self.pedidos           = {}
        self.pedido_en_proceso = None
        self.imagen_actual     = None
        self.config            = cargar_config()
        self.fase_actual       = 1
        self.colecta_global    = {}  # {sku: qty} acumulado en Fase 1
        self.fase1_items       = {}  # widgets de la lista consolidada Fase 1
        self.sku_descripciones = {}  # {sku: descripcion} extraído del PDF
        self.db_nombres        = {}  # {sku: nombre} cargado desde Excel
        self._servidor_thread  = None
        self._servidor_activo  = False
        self._export_pending   = False   # debounce flag para exportar estado
        self._redraw_pending   = False   # debounce flag para redibujar fase1

        self._build_ui()
        self.root.bind('<Escape>', lambda e: self.cancelar_paquete())

        # Carga automática del Excel guardado en config
        self.root.after(200, self._autocargar_excel)
        # Cargar BD interna de SKUs
        self.root.after(100, self._cargar_db_interna)
        # Primer arranque
        self.root.after(600, self._verificar_primera_vez)

    def _build_ui(self):
        # TOPBAR
        topbar = tk.Frame(self.root, bg=C["panel"], height=56)
        topbar.pack(fill="x", side="top")
        topbar.pack_propagate(False)

        left = tk.Frame(topbar, bg=C["panel"])
        left.pack(side="left", padx=20, fill="y")
        tk.Label(left, text="⬡", font=("Segoe UI", 20), bg=C["panel"],
                 fg=C["accent"]).pack(side="left", padx=(0, 10))
        col = tk.Frame(left, bg=C["panel"])
        col.pack(side="left", fill="y", pady=8)
        tk.Label(col, text="SISTEMA DE PICKING", font=("Segoe UI Black", 11),
                 bg=C["panel"], fg=C["text_hi"]).pack(anchor="w")
        tk.Label(col, text="Mercado Libre · Fase 1 (Colecta) + Fase 2 (Preparación)",
                 font=FONT_SMALL, bg=C["panel"], fg=C["text_mid"]).pack(anchor="w")

        center = tk.Frame(topbar, bg=C["panel"])
        center.place(relx=0.5, rely=0.5, anchor="center")
        self.lbl_contador = tk.Label(center, text="— etiquetas pendientes",
                                     font=("Segoe UI Semibold", 12),
                                     bg=C["panel"], fg=C["text_hi"])
        self.lbl_contador.pack()

        right = tk.Frame(topbar, bg=C["panel"])
        right.pack(side="right", padx=16, fill="y")

        # Fila superior: label impresora
        self.lbl_printer = tk.Label(right, text=self._texto_impresora(),
                                    font=FONT_SMALL, bg=C["panel"], fg=C["text_mid"])
        self.lbl_printer.pack(anchor="e", pady=(4, 2))

        # Fila inferior: dos botones uno al lado del otro
        btns_row = tk.Frame(right, bg=C["panel"])
        btns_row.pack(anchor="e")

        self.btn_servidor = tk.Button(
                  btns_row, text="📱  App Móvil",
                  font=FONT_SMALL, bg=C["success"], fg="white",
                  activebackground="#059669", activeforeground="white",
                  relief="flat", cursor="hand2", padx=10, pady=4, bd=0,
                  command=self._mostrar_info_movil)
        self.btn_servidor.pack(side="left", padx=(0, 6))

        tk.Button(btns_row, text="⚙  Impresora",
                  font=FONT_SMALL, bg=C["accent"], fg="white",
                  activebackground=C["accent2"], activeforeground="white",
                  relief="flat", cursor="hand2", padx=10, pady=4, bd=0,
                  command=self.abrir_configuracion).pack(side="left")

        self.lbl_servidor_ip = tk.Label(right, text="",
                                        font=("Segoe UI", 8), bg=C["panel"], fg=C["success"])
        self.lbl_servidor_ip.pack(anchor="e")

        # TABS bajo el topbar
        tabs_bar = tk.Frame(self.root, bg=C["panel"], height=36)
        tabs_bar.pack(fill="x", side="top")
        tabs_bar.pack_propagate(False)

        self._tab_actual = tk.StringVar(value="dashboard")
        def _tab(nombre, texto, activo=False):
            btn = tk.Button(tabs_bar, text=texto,
                            font=("Segoe UI Semibold", 9),
                            bg=C["accent"] if activo else C["panel"],
                            fg="white" if activo else C["text_mid"],
                            activebackground=C["accent2"], activeforeground="white",
                            relief="flat", cursor="hand2", padx=18, pady=0, bd=0,
                            command=lambda n=nombre: self._switch_tab(n))
            btn.pack(side="left", fill="y")
            return btn

        self._btn_tab_dash    = _tab("dashboard", "⬡  Inicio",          activo=True)
        self._btn_tab_ml      = _tab("ml",        "📦  Pedidos ML",      activo=False)
        self._btn_tab_picking = _tab("picking",   "🔍  Picking / Packing", activo=False)

        # Indicador de pedidos pendientes
        self.lbl_ml_badge = tk.Label(tabs_bar, text="",
                                     font=("Segoe UI", 8), bg=C["accent"],
                                     fg="white", padx=6, pady=2)
        self.lbl_ml_badge.pack(side="left")

        # BODY: contenedor que alterna entre paneles
        self.body_container = tk.Frame(self.root, bg=C["bg_dark"])
        self.body_container.pack(fill="both", expand=True)

        # ── Panel Dashboard ───────────────────────────────────────────────
        self.panel_dashboard = tk.Frame(self.body_container, bg=C["bg_dark"])
        self.panel_dashboard.pack(fill="both", expand=True)  # visible por defecto
        self._build_dashboard()

        # ── Panel ML ─────────────────────────────────────────────────────
        self.panel_ml = tk.Frame(self.body_container, bg=C["bg_dark"])
        self._build_ml_panel()

        # ── Panel Picking (3 columnas original) ──────────────────────────
        self.panel_picking = tk.Frame(self.body_container, bg=C["bg_dark"])

        body = self.panel_picking

        # Columna 1: FASE 1 (Colecta)
        self.col_fase1 = tk.Frame(body, bg=C["panel"], width=360)
        self.col_fase1.pack(side="left", fill="both", expand=False)
        self.col_fase1.pack_propagate(False)

        # Columna 2: Controles de escaneo
        self.col_control = tk.Frame(body, bg=C["panel"], width=400)
        self.col_control.pack(side="left", fill="both", expand=False)
        self.col_control.pack_propagate(False)

        # Columna 3: Vista previa etiqueta
        self.col_preview = tk.Frame(body, bg=C["preview_bg"])
        self.col_preview.pack(side="left", fill="both", expand=True)

        self._build_fase1()
        self._build_control_panel()
        self._build_preview_panel()

    # =========================================================================
    # PANEL DE PEDIDOS ML
    # =========================================================================
    def _build_ml_panel(self):
        p = self.panel_ml

        # ── Colores por tipo de logística ────────────────────────────────────
        self._logistica_cfg = {
            "flex":        {"label": "⚡ FLEX",      "color": "#7C3AED", "emoji": "⚡"},
            "me2":         {"label": "🚚 ME2",        "color": "#2563EB", "emoji": "🚚"},
            "me1":         {"label": "📦 ME1",        "color": "#0891B2", "emoji": "📦"},
            "desconocido": {"label": "— Sin envío",   "color": "#475569", "emoji": "—"},
        }
        # Etiquetas por sub-tipo de logistica (según doc oficial ML)
        self._subtipo_label = {
            "self_service":  "⚡ Flex",
            "cross_docking": "🚚 Colecta",
            "xd_drop_off":   "📍 Places",
            "drop_off":      "🏪 Drop Off",
            "fulfillment":   "🏭 Full",
            "turbo":         "⚡ Turbo",
            "default":       "📦 ME1",
        }

        # ── Toolbar ───────────────────────────────────────────────────────────
        tb = tk.Frame(p, bg=C["panel"], pady=8)
        tb.pack(fill="x")

        self.btn_ml_login = tk.Button(
            tb, text="Conectar MercadoLibre",
            font=("Segoe UI Semibold", 10),
            bg="#CA8A04", fg="black",
            activebackground="#EAD700", activeforeground="black",
            relief="flat", cursor="hand2", padx=14, pady=5, bd=0,
            command=self._ml_abrir_login)
        self.btn_ml_login.pack(side="left", padx=(14, 6))

        self.btn_ml_refresh = tk.Button(
            tb, text="Actualizar",
            font=FONT_SMALL, bg=C["panel"], fg=C["text_mid"],
            activebackground=C["border"], activeforeground=C["text_hi"],
            relief="flat", cursor="hand2", padx=10, pady=5, bd=0,
            command=self._ml_refresh_pedidos)
        self.btn_ml_refresh.pack(side="left")

        # Filtros de fecha rápidos
        sep_frame = tk.Frame(tb, bg=C["border"], width=1)
        sep_frame.pack(side="left", fill="y", padx=8, pady=4)

        for label, dias in [("Hoy", 0), ("Ayer", 1), ("7d", 7), ("14d", 14), ("30d", 30)]:
            tk.Button(tb, text=label,
                      font=("Segoe UI", 8), bg=C["panel"], fg=C["text_mid"],
                      activebackground=C["accent"], activeforeground="white",
                      relief="flat", cursor="hand2", padx=8, pady=3, bd=0,
                      command=lambda d=dias: self._ml_refresh_rapido(d)
                      ).pack(side="left", padx=1)

        self.lbl_ml_estado = tk.Label(
            tb, text="Sin conexion a MercadoLibre",
            font=FONT_SMALL, bg=C["panel"], fg=C["text_lo"])
        self.lbl_ml_estado.pack(side="left", padx=(10, 0))

        # Buscar
        bw = tk.Frame(tb, bg=C["border"], padx=1, pady=1)
        bw.pack(side="right", padx=14)
        self.var_buscar_ml = tk.StringVar()
        self.var_buscar_ml.trace_add("write", lambda *_: self._ml_filtrar())
        tk.Entry(bw, textvariable=self.var_buscar_ml,
                 font=FONT_BODY, bg=C["card"], fg=C["text_hi"],
                 insertbackground=C["accent"], relief="flat", bd=0,
                 width=20).pack(fill="x", ipady=4, padx=6)

        # Generar lote
        self.btn_ml_lote = tk.Button(
            tb, text="▶  Generar Lote de Picking",
            font=("Segoe UI Semibold", 10),
            bg=C["success"], fg="white",
            activebackground="#059669", activeforeground="white",
            relief="flat", cursor="hand2", padx=14, pady=5, bd=0,
            state="disabled", command=self._ml_generar_lote)
        self.btn_ml_lote.pack(side="right", padx=(0, 8))

        # ── Pestañas de cuentas ML ────────────────────────────────────────────
        self._cuenta_tabs_frame = tk.Frame(p, bg=C["panel"], pady=4)
        self._cuenta_tabs_frame.pack(fill="x", padx=6)
        # Se llena dinámicamente en _rebuild_cuenta_tabs()

        # ── Sub-pestañas: Todos / Flex / Mercado Envíos ───────────────────────
        sub_tabs = tk.Frame(p, bg=C["bg_dark"])
        sub_tabs.pack(fill="x")

        self._ml_filtro_tipo = tk.StringVar(value="todos")
        self._ml_sub_btns = {}

        for key, txt, color in [
            ("todos",  "Todos los pedidos", C["accent"]),
            ("flex",   "⚡ Mercado Flex",    "#7C3AED"),
            ("me2",    "🚚 Mercado Envios",  "#2563EB"),
            ("me1",    "📦 ME1 / Propio",    "#0891B2"),
        ]:
            btn = tk.Button(sub_tabs, text=txt,
                            font=("Segoe UI Semibold", 9),
                            bg=color if key == "todos" else C["panel"],
                            fg="white" if key == "todos" else C["text_mid"],
                            activebackground=color, activeforeground="white",
                            relief="flat", cursor="hand2",
                            padx=14, pady=6, bd=0,
                            command=lambda k=key, c=color: self._ml_set_filtro(k, c))
            btn.pack(side="left")
            self._ml_sub_btns[key] = (btn, color)

        # Separador
        tk.Frame(sub_tabs, bg=C["border"], width=1).pack(
            side="left", fill="y", padx=10, pady=4)

        # Toggle "Solo pendientes" — ACTIVO POR DEFECTO
        self.var_solo_pendientes = tk.BooleanVar(value=True)
        chk = tk.Checkbutton(
            sub_tabs, text="⏳ Solo pendientes",
            variable=self.var_solo_pendientes,
            font=("Segoe UI Semibold", 9),
            bg=C["bg_dark"], fg=C["warning"],
            activebackground=C["bg_dark"],
            selectcolor=C["card"],
            relief="flat", cursor="hand2", bd=0,
            command=self._ml_filtrar)
        chk.pack(side="left", padx=4)

        # Badges de conteo
        self.lbl_flex_badge  = tk.Label(sub_tabs, text="",
                                         font=("Segoe UI", 8), bg="#7C3AED",
                                         fg="white", padx=6, pady=2)
        self.lbl_me2_badge   = tk.Label(sub_tabs, text="",
                                         font=("Segoe UI", 8), bg="#2563EB",
                                         fg="white", padx=6, pady=2)

        # ── Cabecera tabla ────────────────────────────────────────────────────
        hdr = tk.Frame(p, bg=C["bar_bg"])
        hdr.pack(fill="x")
        for txt, w in [("Tipo", 8), ("Pedido", 10), ("Fecha", 8),
                       ("Comprador", 16), ("SKU · Nombre del Producto", 42),
                       ("Estado Envio", 14), ("Etiqueta", 8)]:
            tk.Label(hdr, text=txt, font=("Segoe UI", 8, "bold"),
                     bg=C["bar_bg"], fg=C["accent"],
                     width=w, anchor="w").pack(side="left", padx=4, pady=5)

        # ── Lista scrollable ──────────────────────────────────────────────────
        lista = tk.Frame(p, bg=C["bg_dark"])
        lista.pack(fill="both", expand=True)
        self.canvas_ml = tk.Canvas(lista, bg=C["bg_dark"], highlightthickness=0)
        sb = tk.Scrollbar(lista, orient="vertical", command=self.canvas_ml.yview)
        self.frame_ml  = tk.Frame(self.canvas_ml, bg=C["bg_dark"])
        self.canvas_ml.create_window((0, 0), window=self.frame_ml, anchor="nw")
        self.canvas_ml.configure(yscrollcommand=sb.set)
        self.frame_ml.bind("<Configure>",
                           lambda e: self.canvas_ml.configure(
                               scrollregion=self.canvas_ml.bbox("all")))
        self.canvas_ml.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.canvas_ml.bind("<MouseWheel>",
                            lambda e: self.canvas_ml.yview_scroll(
                                int(-1*(e.delta/120)), "units"))

        # ── Estado inicial ────────────────────────────────────────────────────
        self._ml_pedidos       = {}
        self._ml_autenticado   = False
        self._ml_filtro_actual = "todos"
        self._ml_cuentas       = {}       # { cuenta_id: nickname }
        self._ml_cuenta_filtro = "todas"  # "todas" o cuenta_id específico
        self._cuenta_tab_widgets = []

        self._ml_mostrar_placeholder("Conéctate a MercadoLibre para ver los pedidos")
        self.root.after(1200, self._ml_verificar_conexion)

    def _tipo_logistica(self, ped):
        """
        Clasifica un pedido en: flex, me2, me1, desconocido.
        SIEMPRE recalcula desde logistica. No asume me2 sin confirmar.
        """
        log = (ped.get("logistica") or "").lower().strip()

        # Clasificacion definitiva por logistic_type
        if log == "self_service":
            return "flex"
        if log in ("cross_docking", "drop_off", "xd_drop_off",
                   "fulfillment", "turbo", "xd_same_day"):
            return "me2"
        if log in ("default", "custom", "not_specified"):
            return "me1"

        # Fallback por tags
        tags_str = " ".join(t.lower() for t in ped.get("tags", []))
        if "self_service" in tags_str or "flex" in tags_str:
            return "flex"

        # Fallback por tipo pre-calculado del servidor
        tipo = (ped.get("tipo") or "").lower()
        if tipo in ("flex", "me1", "me2"):
            return tipo

        # Sin info confirmada → desconocido (NO asumir me2)
        return "desconocido"

    def _ml_set_filtro(self, key, color):
        """Cambia la sub-pestaña activa."""
        self._ml_filtro_actual = key
        self._ml_filtro_tipo.set(key)  # SINCRONIZAR el StringVar
        for k, (btn, c) in self._ml_sub_btns.items():
            if k == key:
                btn.config(bg=c, fg="white")
            else:
                btn.config(bg=C["panel"], fg=C["text_mid"])
        self._ml_filtrar()

    def _ml_filtrar(self):
        q    = self.var_buscar_ml.get().strip().lower()
        tipo = self._ml_filtro_tipo.get()
        cta  = getattr(self, "_ml_cuenta_filtro", "todas")
        solo_pendientes = getattr(self, "var_solo_pendientes",
                                  tk.BooleanVar(value=True)).get()
        peds = list(self._ml_pedidos.values())

        # Filtrar por cuenta ML
        if cta != "todas":
            peds = [p for p in peds if p.get("_cuenta") == cta]

        # ── SOLO PENDIENTES (por defecto activo) ──────────────────────────────
        # Solo muestra pedidos NO impresos con estado ready_to_ship o sin estado
        if solo_pendientes:
            peds = [p for p in peds
                    if not p.get("impreso", False)
                    and p.get("estado_envio","") not in
                    ("shipped", "delivered", "cancelled", "not_delivered")]

        # Filtrar por tipo de logística
        if tipo != "todos":
            peds = [p for p in peds if self._tipo_logistica(p) == tipo]

        # Filtrar por búsqueda de texto
        if q:
            peds = [p for p in peds if
                    q in p.get("comprador","").lower() or
                    q in str(p.get("order_id","")).lower() or
                    any(q in it.get("titulo","").lower() or
                        q in it.get("sku","").lower()
                        for it in p.get("items",[]))]

        # Actualizar badges — contar solo pendientes por tipo
        base = [p for p in self._ml_pedidos.values()
                if not p.get("impreso", False)
                and p.get("estado_envio","") not in
                ("shipped","delivered","cancelled","not_delivered")]
        n_flex = sum(1 for p in base if self._tipo_logistica(p) == "flex")
        n_me2  = sum(1 for p in base if self._tipo_logistica(p) == "me2")
        n_me1  = sum(1 for p in base if self._tipo_logistica(p) == "me1")

        for key, badge_attr, count in [
            ("flex", "lbl_flex_badge", n_flex),
            ("me2",  "lbl_me2_badge",  n_me2),
        ]:
            lbl = getattr(self, badge_attr, None)
            if lbl:
                lbl.config(text=f" {count} " if count else "")
                if count:
                    btn = self._ml_sub_btns[key][0]
                    lbl.pack(side="left", after=btn)

        self._ml_render_pedidos(peds)

    def _ml_render_pedidos(self, pedidos):
        for w in self.frame_ml.winfo_children():
            w.destroy()

        if not pedidos:
            tk.Label(self.frame_ml, text="Sin pedidos en esta categoría",
                     font=("Segoe UI", 11), bg=C["bg_dark"],
                     fg=C["text_lo"]).pack(pady=50)
            return

        tipo_filtro = self._ml_filtro_tipo.get()

        if tipo_filtro == "todos":
            # Agrupar con encabezados de sección
            grupos = {}
            for ped in pedidos:
                t = self._tipo_logistica(ped)
                grupos.setdefault(t, []).append(ped)

            SECCIONES = [
                ("flex",        "⚡  MERCADO FLEX",     "#7C3AED"),
                ("me2",         "🚚  MERCADO ENVÍOS",   "#2563EB"),
                ("me1",         "📦  ME1 / PROPIO",     "#0891B2"),
                ("desconocido", "❓  SIN CLASIFICAR",   "#475569"),
            ]
            primer = True
            for key, titulo, color in SECCIONES:
                lista = grupos.get(key, [])
                if not lista:
                    continue
                # Encabezado de sección con color
                hdr = tk.Frame(self.frame_ml, bg=color)
                hdr.pack(fill="x", pady=(0 if primer else 8, 0))
                tk.Label(hdr,
                         text=f"   {titulo}   ·   {len(lista)} pedido(s)",
                         font=("Segoe UI Black", 10), bg=color,
                         fg="white", anchor="w", pady=7).pack(fill="x")
                self._ml_render_filas(lista, color)
                primer = False
        else:
            # Filtro específico — sin encabezado, solo las filas
            COLORES = {"flex":"#7C3AED","me2":"#2563EB","me1":"#0891B2","desconocido":"#475569"}
            color = COLORES.get(tipo_filtro, C["accent"])
            self._ml_render_filas(pedidos, color)

    def _ml_render_filas(self, pedidos, color_tipo=None):
        """Renderiza filas de pedidos con columnas alineadas."""

        # Cabecera de columnas
        hdr = tk.Frame(self.frame_ml, bg=C["panel"])
        hdr.pack(fill="x")
        tk.Frame(hdr, bg=C["panel"], width=4).pack(side="left", fill="y")  # borde
        for txt, w in [("Tipo",6), ("Pedido",10), ("Fecha",9),
                       ("Comprador",16), ("SKU · Producto",40), ("Estado",12), ("",8)]:
            tk.Label(hdr, text=txt, font=("Segoe UI", 8, "bold"),
                     bg=C["panel"], fg=C["text_lo"],
                     width=w, anchor="w").pack(side="left", padx=4)

        tk.Frame(self.frame_ml, bg=C["border"], height=1).pack(fill="x")

        bg_alt = [C["card"], C["bg_dark"]]
        for i, ped in enumerate(pedidos):
            bg    = bg_alt[i % 2]
            tipo  = self._tipo_logistica(ped)
            log   = (ped.get("logistica") or "").lower()
            impreso = ped.get("impreso", False)

            # Color del borde y chip según tipo
            TIPO_COLOR = {
                "flex": "#7C3AED", "me2": "#2563EB",
                "me1": "#0891B2",  "desconocido": "#475569"
            }
            SUBTIPO_LABEL = {
                "self_service": "⚡ Flex",
                "xd_drop_off":  "📍 Places",
                "cross_docking":"🚚 Colecta",
                "drop_off":     "🏪 Drop Off",
                "fulfillment":  "🏭 Full",
                "turbo":        "⚡ Turbo",
                "default":      "📦 ME1",
            }
            borde_color = TIPO_COLOR.get(tipo, "#475569")
            chip_txt    = SUBTIPO_LABEL.get(log, tipo.upper() if tipo != "desconocido" else "—")

            # Fila
            fila = tk.Frame(self.frame_ml, bg=bg)
            fila.pack(fill="x")

            # Borde de color izquierdo
            tk.Frame(fila, bg=borde_color, width=4).pack(side="left", fill="y")

            inner = tk.Frame(fila, bg=bg, pady=4)
            inner.pack(side="left", fill="x", expand=True)

            # Función helper para labels
            def _col(txt, w, fg=None, bold=False, mono=False, bg_=bg):
                font = ("Consolas", 9) if mono else \
                       ("Segoe UI Semibold", 9) if bold else FONT_SMALL
                tk.Label(inner, text=str(txt), font=font, bg=bg_,
                         fg=fg or (C["text_mid"] if impreso else C["text_hi"]),
                         width=w, anchor="w", wraplength=240).pack(side="left", padx=4)

            # Col 1: Chip de tipo
            tk.Label(inner, text=chip_txt,
                     font=("Segoe UI Semibold", 7),
                     bg=borde_color, fg="white",
                     padx=5, pady=2, width=8, anchor="center").pack(side="left", padx=4)

            # Col 2: Pedido #
            _col(f"#{ped.get('order_id','')}", 10,
                 fg=C["text_mid"] if impreso else C["accent"], bold=True)

            # Col 3: Fecha
            _col(ped.get("fecha","")[:10], 9, fg=C["text_mid"])

            # Col 4: Comprador
            _col((ped.get("comprador","")[:16] or "—"), 16, bold=True)

            # Col 5: SKUs y productos
            items = ped.get("items", [])
            lineas = []
            for it in items[:3]:
                sku = it.get("sku","")
                nom = it.get("titulo","")
                if sku and sku in self.db_nombres:
                    info = self.db_nombres[sku]
                    nom  = info.get("nombre","") if isinstance(info,dict) else str(info)
                qty = it.get("cantidad",1)
                if sku and nom:
                    lineas.append(f"{sku}  {nom[:20]}  ×{qty}")
                elif sku:
                    lineas.append(f"{sku}  ×{qty}")
                elif nom:
                    lineas.append(f"{nom[:28]}  ×{qty}")
            if len(items) > 3:
                lineas.append(f"+{len(items)-3} más")
            _col("\n".join(lineas) if lineas else "—", 40, mono=True)

            # Col 6: Estado envío
            est = ped.get("estado_envio","") or "—"
            col_est = C["success"]  if "ready" in est else \
                      C["warning"]  if est not in ("—","delivered","cancelled") else \
                      C["text_lo"]
            _col(est[:14], 12, fg=col_est)

            # Col 7: Botón etiqueta / impreso
            if impreso:
                tk.Label(inner, text="✅ Impreso",
                         font=("Segoe UI", 8), bg=bg,
                         fg=C["success"]).pack(side="left", padx=4)
            elif ped.get("shipping_id"):
                tk.Button(inner, text="🏷 Etiqueta",
                          font=("Segoe UI Semibold", 8),
                          bg=borde_color, fg="white",
                          activebackground=C["accent2"],
                          activeforeground="white",
                          relief="flat", cursor="hand2",
                          padx=8, pady=2, bd=0,
                          command=lambda oid=ped["order_id"]: self._ml_ver_etiqueta(oid)
                          ).pack(side="left", padx=4)

        self.canvas_ml.yview_moveto(0)

    def _switch_tab(self, nombre):
        """Alterna entre Dashboard, Pedidos ML y Picking/Packing."""
        # Ocultar todos
        for panel in [self.panel_dashboard, self.panel_ml, self.panel_picking]:
            panel.pack_forget()

        # Resetear todos los botones
        for btn, tab in [
            (self._btn_tab_dash,    "dashboard"),
            (self._btn_tab_ml,      "ml"),
            (self._btn_tab_picking, "picking"),
        ]:
            activo = (tab == nombre)
            btn.config(
                bg=C["accent"] if activo else C["panel"],
                fg="white"     if activo else C["text_mid"])

        # Mostrar el panel correcto
        if nombre == "dashboard":
            self.panel_dashboard.pack(fill="both", expand=True)
            self._refresh_dashboard()
        elif nombre == "ml":
            self.panel_ml.pack(fill="both", expand=True)
        else:
            self.panel_picking.pack(fill="both", expand=True)
            self.entrada_sku.focus()

    # ── Conexión ML ──────────────────────────────────────────────────────────

    def _ml_verificar_conexion(self):
        """Consulta el servidor Railway para ver todas las cuentas activas."""
        import threading
        def _worker():
            try:
                import urllib.request, json as _j
                url = RAILWAY_URL.rstrip("/")
                with urllib.request.urlopen(url + "/auth/status", timeout=6) as r:
                    d = _j.loads(r.read())
                self.root.after(0, lambda: self._ml_on_status(d))
            except Exception:
                self.root.after(0, lambda: self.lbl_ml_estado.config(
                    text="Sin conexion al servidor Railway", fg=C["danger"]))
        threading.Thread(target=_worker, daemon=True).start()

    def _ml_on_status(self, d):
        cuentas = d.get("cuentas", [])
        if cuentas:
            self._ml_autenticado = True
            n_ped  = d.get("pedidos", 0)

            # Actualizar botón principal — mostrar nicknames conectados
            if len(cuentas) == 1:
                nick = cuentas[0]["nickname"]
                self.btn_ml_login.config(
                    text=f"✅  {nick}",
                    bg=C["success"], fg="white",
                    activebackground="#059669",
                    command=self._ml_agregar_cuenta)   # clic → agregar otra cuenta
            else:
                nicks = "  +  ".join(c["nickname"] for c in cuentas)
                self.btn_ml_login.config(
                    text=f"✅  {nicks}",
                    bg=C["success"], fg="white",
                    activebackground="#059669",
                    command=self._ml_agregar_cuenta)

            self.lbl_ml_estado.config(
                text=f"{'  ·  '.join(c['nickname'] for c in cuentas)}"
                     f"  ·  {n_ped} pedidos  ·  {d.get('ultimo_refresh','')}",
                fg=C["success"])

            # Reconstruir pestañas de cuentas
            self._ml_cuentas = {c["cuenta_id"]: c["nickname"] for c in cuentas}
            self._rebuild_cuenta_tabs()

            if n_ped > 0:
                self._ml_refresh_pedidos()
            else:
                self.root.after(500, self._ml_refresh_pedidos)
        else:
            # No hay cuentas conectadas
            self.btn_ml_login.config(
                text="Conectar MercadoLibre",
                bg="#CA8A04", fg="black",
                activebackground="#EAD700",
                command=self._ml_abrir_login)
            self.lbl_ml_estado.config(
                text="No conectado a MercadoLibre",
                fg=C["warning"])

    def _rebuild_cuenta_tabs(self):
        """Reconstruye las pestañas de cuentas en la toolbar ML."""
        # Eliminar botones de cuenta anteriores
        for w in getattr(self, "_cuenta_tab_widgets", []):
            try: w.destroy()
            except Exception: pass
        self._cuenta_tab_widgets = []

        if not hasattr(self, "_ml_cuentas"):
            return

        # Contenedor de pestañas de cuentas (se inserta en _cuenta_tabs_frame)
        frame = getattr(self, "_cuenta_tabs_frame", None)
        if not frame or not frame.winfo_exists():
            return

        # Limpiar frame
        for w in frame.winfo_children():
            w.destroy()

        COLORES = ["#7C3AED", "#0891B2", "#D97706", "#DC2626", "#059669"]
        for idx, (cid, nick) in enumerate(self._ml_cuentas.items()):
            color = COLORES[idx % len(COLORES)]
            is_active = getattr(self, "_ml_cuenta_filtro", "todas") == cid

            btn = tk.Button(
                frame,
                text=f"  {nick}  ",
                font=("Segoe UI Semibold", 9),
                bg=color if is_active else C["panel"],
                fg="white" if is_active else C["text_mid"],
                activebackground=color, activeforeground="white",
                relief="flat", cursor="hand2", padx=10, pady=5, bd=0,
                command=lambda c=cid, co=color: self._ml_set_cuenta_filtro(c, co))
            btn.pack(side="left")
            self._cuenta_tab_widgets.append(btn)

            # X para desconectar
            x_btn = tk.Button(
                frame,
                text="×",
                font=("Segoe UI", 9, "bold"),
                bg=color if is_active else C["panel"],
                fg="white" if is_active else C["text_lo"],
                activebackground=C["danger"], activeforeground="white",
                relief="flat", cursor="hand2", padx=4, pady=5, bd=0,
                command=lambda c=cid, n=nick: self._ml_desconectar_cuenta(c, n))
            x_btn.pack(side="left", padx=(0, 6))
            self._cuenta_tab_widgets.append(x_btn)

        # Botón agregar nueva cuenta
        add_btn = tk.Button(
            frame,
            text="+ Agregar cuenta",
            font=("Segoe UI", 8),
            bg=C["panel"], fg=C["text_mid"],
            activebackground=C["success"], activeforeground="white",
            relief="flat", cursor="hand2", padx=8, pady=5, bd=0,
            command=self._ml_agregar_cuenta)
        add_btn.pack(side="left", padx=(4, 0))
        self._cuenta_tab_widgets.append(add_btn)

    def _ml_set_cuenta_filtro(self, cuenta_id, color):
        """Filtra la lista por cuenta específica."""
        self._ml_cuenta_filtro = cuenta_id
        self._rebuild_cuenta_tabs()
        self._ml_filtrar()

    def _ml_desconectar_cuenta(self, cuenta_id, nick):
        if not messagebox.askyesno(
                "Desconectar cuenta",
                f"¿Desconectar la cuenta '{nick}'?\n"
                f"Se eliminarán sus pedidos de la lista.",
                parent=self.root):
            return
        import threading, urllib.request, json as _j
        def _w():
            try:
                url = RAILWAY_URL.rstrip("/")
                req = urllib.request.Request(
                    f"{url}/api/cuentas/{cuenta_id}/logout",
                    method="POST",
                    headers={"Content-Type": "application/json",
                             "X-API-Key": self.config.get("clave_nube", "everest2024")})
                urllib.request.urlopen(req, timeout=6)
            except Exception:
                pass
            self.root.after(0, lambda: self._ml_verificar_conexion())
        threading.Thread(target=_w, daemon=True).start()

    def _ml_abrir_login(self):
        """Abre el browser para autenticar la cuenta_0 (primera cuenta)."""
        import webbrowser
        url = RAILWAY_URL.rstrip("/") + "/auth/login?cuenta=cuenta_0"
        webbrowser.open(url)
        messagebox.showinfo(
            "Autenticación MercadoLibre",
            "Se abrió el navegador.\n\n"
            "1. Iniciá sesión con tu cuenta ML\n"
            "2. Autorizá la aplicación\n"
            "3. Volvé acá y hacé clic en 'Actualizar pedidos'",
            parent=self.root)
        self.root.after(8000, self._ml_verificar_conexion)

    def _ml_agregar_cuenta(self):
        """Abre el login de ML para una nueva cuenta."""
        # Calcular próximo ID de cuenta
        cuentas = getattr(self, "_ml_cuentas", {})
        n = len(cuentas)
        cuenta_id = f"cuenta_{n}"

        import webbrowser
        url = RAILWAY_URL.rstrip("/") + f"/auth/login?cuenta={cuenta_id}"
        webbrowser.open(url)

        # Ventana de espera elegante
        win = tk.Toplevel(self.root)
        win.title("Conectar cuenta ML")
        win.geometry("440x260")
        win.resizable(False, False)
        win.config(bg=C["bg_dark"])
        win.grab_set()

        hdr = tk.Frame(win, bg="#CA8A04", pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Conectar nueva cuenta MercadoLibre",
                 font=("Segoe UI Semibold", 12), bg="#CA8A04", fg="black").pack()

        body = tk.Frame(win, bg=C["bg_dark"], padx=28, pady=20)
        body.pack(fill="both", expand=True)

        tk.Label(body,
                 text="Se abrió el navegador.\n\n"
                      "1. Iniciá sesión con la cuenta ML a agregar\n"
                      "2. Autorizá la aplicación\n"
                      "3. Volvé aquí y hacé clic en Verificar",
                 font=("Segoe UI", 10), bg=C["bg_dark"], fg=C["text_hi"],
                 justify="left").pack(anchor="w")

        sep = tk.Frame(body, bg=C["border"], height=1)
        sep.pack(fill="x", pady=(16, 10))

        btn_row = tk.Frame(body, bg=C["bg_dark"])
        btn_row.pack(fill="x")

        lbl_res = tk.Label(body, text="", font=FONT_SMALL,
                           bg=C["bg_dark"], fg=C["text_mid"])
        lbl_res.pack(anchor="w", pady=(6, 0))

        def _verificar():
            lbl_res.config(text="Verificando...", fg=C["accent"])
            win.update_idletasks()
            import urllib.request, json as _j, threading
            def _w():
                try:
                    url2 = RAILWAY_URL.rstrip("/") + "/auth/status"
                    with urllib.request.urlopen(url2, timeout=6) as r:
                        d = _j.loads(r.read())
                    cuentas_nuevas = {c["cuenta_id"]: c["nickname"]
                                      for c in d.get("cuentas", [])}
                    if cuenta_id in cuentas_nuevas:
                        nick = cuentas_nuevas[cuenta_id]
                        self.root.after(0, lambda: [
                            lbl_res.config(text=f"✅ Conectado como {nick}",
                                           fg=C["success"]),
                            win.after(1200, win.destroy),
                            self._ml_on_status(d),
                            self._ml_refresh_pedidos(),
                        ])
                    else:
                        self.root.after(0, lambda: lbl_res.config(
                            text="Aún no conectado. Completá la autorización en el browser.",
                            fg=C["warning"]))
                except Exception as e:
                    self.root.after(0, lambda: lbl_res.config(
                        text=f"Error: {e}", fg=C["danger"]))
            threading.Thread(target=_w, daemon=True).start()

        tk.Button(btn_row, text="Cancelar",
                  font=FONT_BODY, bg=C["panel"], fg=C["text_mid"],
                  relief="flat", cursor="hand2", padx=14, pady=6, bd=0,
                  command=win.destroy).pack(side="left")
        tk.Button(btn_row, text="✅  Ya autoricé — Verificar",
                  font=("Segoe UI Semibold", 10),
                  bg=C["success"], fg="white",
                  activebackground="#059669", activeforeground="white",
                  relief="flat", cursor="hand2", padx=16, pady=6, bd=0,
                  command=_verificar).pack(side="right")

    def _ml_refresh_rapido(self, dias):
        """Refresh rápido con rango de días predefinido."""
        from datetime import datetime, timedelta
        if dias == 0:
            f_desde = datetime.now().strftime("%Y-%m-%d")
            f_hasta = f_desde
        elif dias == 1:
            f_desde = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
            f_hasta = f_desde
        else:
            f_desde = (datetime.now() - timedelta(days=dias)).strftime("%Y-%m-%d")
            f_hasta = datetime.now().strftime("%Y-%m-%d")
        self._ml_refresh_pedidos(fecha_desde=f_desde, fecha_hasta=f_hasta)
        import webbrowser
        url = RAILWAY_URL.rstrip("/") + "/auth/login?cuenta=cuenta_0"
        webbrowser.open(url)
        messagebox.showinfo(
            "Autenticación MercadoLibre",
            "Se abrió el navegador.\n\n"
            "1. Iniciá sesión con tu cuenta ML\n"
            "2. Autorizá la aplicación\n"
            "3. Volvé acá y hacé clic en 'Actualizar pedidos'",
            parent=self.root)
        self.root.after(8000, self._ml_verificar_conexion)

    def _ml_refresh_pedidos(self, fecha_desde=None, fecha_hasta=None):
        """Trae pedidos de TODAS las cuentas con filtro de fecha."""
        import threading
        from datetime import datetime, timedelta
        # Por defecto: últimos 7 días
        if not fecha_desde:
            fecha_desde = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
        if not fecha_hasta:
            fecha_hasta = datetime.now().strftime("%Y-%m-%d")

        self.lbl_ml_estado.config(
            text=f"Actualizando pedidos ({fecha_desde} → {fecha_hasta})...",
            fg=C["accent"])
        self.btn_ml_refresh.config(state="disabled")

        def _worker():
            try:
                import urllib.request, json as _j
                url = RAILWAY_URL.rstrip("/")
                body = json.dumps({
                    "fecha_desde": fecha_desde,
                    "fecha_hasta": fecha_hasta
                }).encode("utf-8")
                try:
                    req = urllib.request.Request(
                        url + "/api/pedidos/refresh", method="POST",
                        data=body,
                        headers={"Content-Type": "application/json"})
                    urllib.request.urlopen(req, timeout=5)
                    import time; time.sleep(3)
                except Exception:
                    pass
                with urllib.request.urlopen(url + "/api/pedidos", timeout=15) as r:
                    d = _j.loads(r.read())
                self.root.after(0, lambda: self._ml_on_pedidos(d))
            except Exception as e:
                self.root.after(0, lambda: self._ml_on_error(str(e)))

        threading.Thread(target=_worker, daemon=True).start()

    def _ml_on_pedidos(self, d):
        self.btn_ml_refresh.config(state="normal")
        if not d.get("ok"):
            if "login" in str(d.get("msg", "")):
                self.lbl_ml_estado.config(
                    text="Sesión expirada — reconectá con ML", fg=C["warning"])
                self._ml_mostrar_placeholder("Sesión expirada. Clic en 'Conectar MercadoLibre'")
            else:
                self._ml_on_error(d.get("msg", "Error desconocido"))
            return

        self._ml_pedidos = {p["order_id"]: p for p in d.get("pedidos", [])}
        total  = len(self._ml_pedidos)
        pend   = sum(1 for p in self._ml_pedidos.values() if not p.get("impreso"))

        # Contar por cuenta
        por_cuenta = {}
        for p in self._ml_pedidos.values():
            cid = p.get("_cuenta", "cuenta_0")
            nick = p.get("_nickname", cid)
            por_cuenta.setdefault(nick, 0)
            por_cuenta[nick] += 1

        resumen = "  ·  ".join(f"{n}: {c}" for n, c in por_cuenta.items()) or f"{total} pedidos"
        self.lbl_ml_estado.config(
            text=f"✅  {resumen}  ·  {pend} pendientes  ·  {d.get('ts','')}",
            fg=C["success"])
        self.btn_ml_lote.config(state="normal" if pend > 0 else "disabled")

        if pend > 0:
            self.lbl_ml_badge.config(text=f" {pend} ", bg=C["warning"], fg="black")
            self.lbl_ml_badge.pack(side="left")
        else:
            self.lbl_ml_badge.config(text="")

        self._ml_filtrar()

    def _ml_on_error(self, msg):
        self.btn_ml_refresh.config(state="normal")
        self.lbl_ml_estado.config(text=f"Error: {msg}", fg=C["danger"])

    # ── Renderizado de pedidos ────────────────────────────────────────────────

    def _ml_filtrar(self):
        q = self.var_buscar_ml.get().strip().lower()
        peds = list(self._ml_pedidos.values())
        if q:
            peds = [p for p in peds if
                    q in p.get("comprador","").lower() or
                    q in str(p.get("order_id","")).lower() or
                    any(q in it.get("titulo","").lower() or q in it.get("sku","").lower()
                        for it in p.get("items",[]))]
        self._ml_render_pedidos(peds)

    def _ml_render_pedidos(self, pedidos):
        for w in self.frame_ml.winfo_children():
            w.destroy()

        if not pedidos:
            self._ml_mostrar_placeholder("Sin pedidos que mostrar")
            return

        bg_alt = [C["card"], C["bg_dark"]]
        for i, ped in enumerate(pedidos):
            bg  = bg_alt[i % 2]
            fila = tk.Frame(self.frame_ml, bg=bg, pady=4)
            fila.pack(fill="x")

            def _lbl(txt, w, fg=None, mono=False, bold=False):
                f = ("Consolas", 9) if mono else (("Segoe UI Semibold", 9) if bold else FONT_SMALL)
                tk.Label(fila, text=txt, font=f, bg=bg,
                         fg=fg or C["text_hi"], width=w,
                         anchor="w", wraplength=250).pack(side="left", padx=4)

            # Pedido
            _lbl(f"#{ped['order_id']}", 10, fg=C["accent"], bold=True)
            # Fecha
            _lbl(ped.get("fecha","")[:10], 8, fg=C["text_mid"])
            # Comprador
            _lbl(ped.get("comprador","")[:20], 18, bold=True)
            # Productos
            items_txt = "\n".join(
                f"{it.get('sku','?')}  {(self.db_nombres.get(it.get('sku',''),{}) or {}).get('nombre', it.get('titulo',''))[:22]}  x{it.get('cantidad',1)}"
                for it in ped.get("items",[])[:4])
            if len(ped.get("items",[])) > 4:
                items_txt += f"\n+{len(ped['items'])-4} más"
            _lbl(items_txt, 42, mono=True)
            # Estado envío (sin columna de precio)
            est = ped.get("estado_envio","") or ped.get("logistica","")
            col_est = C["success"] if "ready" in est else \
                      (C["warning"] if est else C["text_lo"])
            _lbl(est[:14], 14, fg=col_est)
            # Botón etiqueta
            tk.Button(fila, text="Etiqueta",
                      font=("Segoe UI", 8), bg="#CA8A04", fg="black",
                      activebackground="#EAD700", activeforeground="black",
                      relief="flat", cursor="hand2", padx=8, pady=2, bd=0,
                      command=lambda oid=ped["order_id"]: self._ml_ver_etiqueta(oid)
                      ).pack(side="left", padx=4)

        self.canvas_ml.yview_moveto(0)

    def _ml_mostrar_placeholder(self, texto):
        for w in self.frame_ml.winfo_children():
            w.destroy()
        tk.Label(self.frame_ml, text=texto,
                 font=("Segoe UI", 11), bg=C["bg_dark"],
                 fg=C["text_lo"]).pack(pady=60)

    # ── Etiqueta ──────────────────────────────────────────────────────────────

    def _ml_ver_etiqueta(self, order_id):
        """Abre la etiqueta directamente desde el proxy de Railway."""
        import webbrowser
        ped = self._ml_pedidos.get(order_id, {})

        # La URL del proxy sirve el PDF directamente con el token del servidor
        # No necesitamos llamar a Railway primero — abrimos la URL proxy directo
        proxy_url = RAILWAY_URL.rstrip("/") + f"/api/etiqueta/{order_id}"
        self.lbl_ml_estado.config(
            text=f"Abriendo etiqueta #{order_id}...", fg=C["accent"])
        webbrowser.open(proxy_url)
        self.lbl_ml_estado.config(
            text=f"Etiqueta abierta en el navegador — NO marcada en ML",
            fg=C["success"])

    def _ml_on_etiqueta(self, d, order_id):
        self.lbl_ml_estado.config(
            text=f"Etiqueta obtenida para #{order_id} — NO marcada en ML",
            fg=C["success"])
        if d.get("ok") and d.get("url"):
            import webbrowser
            # Construir URL con access_token para autenticar la descarga
            url = d["url"]
            at  = d.get("access_token", "")
            if at and "access_token" not in url:
                url += f"&access_token={at}"
            webbrowser.open(url)
            # NO marcar como impreso en nuestro sistema
            # para no confundir con el estado real de ML
            self._ml_filtrar()
        else:
            messagebox.showwarning(
                "Sin etiqueta",
                d.get("msg", "No se encontró etiqueta para este pedido.\n"
                      "Verificá que el envío esté en estado 'ready_to_ship'."),
                parent=self.root)

    # ── Generar lote de picking ───────────────────────────────────────────────

    def _ml_generar_lote(self):
        """Toma los pedidos ML pendientes y genera el lote de picking."""
        pendientes = [p for p in self._ml_pedidos.values() if not p.get("impreso")]
        if not pendientes:
            messagebox.showwarning("Sin pedidos", "No hay pedidos pendientes.", parent=self.root)
            return

        # Consolidar SKUs de todos los pedidos
        total_req  = {}
        sku_nombre = {}
        for ped in pendientes:
            for it in ped.get("items", []):
                sku = (it.get("sku") or it.get("item_id","")).upper().strip()
                if not sku:
                    continue
                total_req[sku]  = total_req.get(sku, 0) + it.get("cantidad", 1)
                if sku not in sku_nombre:
                    sku_nombre[sku] = it.get("titulo", sku)

        if not total_req:
            messagebox.showwarning(
                "Sin SKUs",
                "Los pedidos no tienen SKUs identificados.\n\n"
                "Asegurate de tener el campo 'seller_custom_field' o 'SKU' cargado en ML.",
                parent=self.root)
            return

        # Enriquecer con BD interna (pasillo, estantería, nombre)
        grupos_dict = {}
        SIN_PASILLO = "Sin ubicacion en BD"
        for sku, qty in sorted(total_req.items()):
            info = self.db_nombres.get(sku, {})
            if isinstance(info, dict):
                nombre     = info.get("nombre", "") or sku_nombre.get(sku, sku)
                pasillo    = info.get("pasillo", "") or SIN_PASILLO
                estanteria = info.get("estanteria", "")
            else:
                nombre     = str(info) or sku_nombre.get(sku, sku)
                pasillo    = SIN_PASILLO
                estanteria = ""
            grupos_dict.setdefault(pasillo, []).append({
                "sku": sku, "nombre": nombre, "req": qty,
                "pasillo": pasillo, "estanteria": estanteria
            })

        def _orden_pas(n):
            import re as _re
            m = _re.search(r'(\d+)', n)
            return (0, int(m.group(1)), n) if m else (1, 0, n)

        grupos = [{"pasillo": k, "items": v}
                  for k, v in sorted(grupos_dict.items(), key=lambda x: _orden_pas(x[0]))]

        total_skus = len(total_req)
        total_uds  = sum(total_req.values())
        sin_bd     = sum(1 for sku in total_req if sku not in self.db_nombres)

        # Confirmar con el usuario
        msg = (f"Se van a preparar {len(pendientes)} pedido(s):\n"
               f"• {total_skus} SKUs distintos\n"
               f"• {total_uds} unidades en total\n"
               f"• {len(grupos)} pasillo(s)")
        if sin_bd > 0:
            msg += f"\n\n⚠  {sin_bd} SKUs sin ubicacion en la BD.\nApareceran en 'Sin ubicacion en BD'."

        if not messagebox.askokcancel("Generar Lote de Picking", msg, parent=self.root):
            return

        # Construir pedidos internos para Fase 1 y Fase 2
        self.pedidos.clear()
        for idx, ped in enumerate(pendientes, start=1):
            skus_req = {}
            for it in ped.get("items", []):
                sku = (it.get("sku") or it.get("item_id","")).upper().strip()
                if sku:
                    skus_req[sku] = skus_req.get(sku, 0) + it.get("cantidad", 1)
            if skus_req:
                self.pedidos[idx] = {
                    "pagina":         idx,
                    "skus_requeridos": skus_req,
                    "skus_escaneados": {},
                    "impreso":         False,
                    "descripcion":     ped.get("comprador",""),
                    "_order_id":       ped["order_id"],
                    "_shipping_id":    ped.get("shipping_id",""),
                }

        # Resetear estado
        self.pedido_en_proceso = None
        self.fase_actual       = 1
        self.colecta_global    = {}
        self.ruta_pdf          = ""  # sin PDF, usamos datos ML
        self.lbl_resultado.config(text="")
        self.lbl_num_caja.config(text="")
        self.lbl_imprimiendo.config(text="")
        self._mostrar_placeholder_visor()

        # Dibujar Fase 1 agrupado por pasillo
        self._dibujar_fase1()
        self.actualizar_contador_global()
        self.entrada_sku.config(state="normal")
        self.btn_fase2.config(state="normal", bg=C["success"], fg="white")
        self.lbl_fase_actual.config(text="● FASE 1: COLECTA EN DEPOSITO", fg=C["accent"])
        self.lbl_col1_header.config(text="FASE 1: COLECTA EN DEPOSITO")
        self.lbl_estado_pdf.config(
            text=f"Lote ML: {len(self.pedidos)} pedidos / {total_uds} unidades — Subiendo a celulares…",
            fg=C["accent"])

        # Siempre subir a Railway para que los celulares reciban el lote
        self._exportar_estado_movil()
        self._subir_a_nube_async()
        self.root.after(5000, self._sincronizar_desde_nube)

        # Cambiar a la pestaña de Picking
        self._switch_tab("picking")
        self.entrada_sku.focus()

        n_sin = sin_bd
        if n_sin > 0:
            messagebox.showinfo(
                "Lote generado",
                f"Lote generado correctamente.\n\n"
                f"Tip: {n_sin} SKUs sin pasillo en la BD.\n"
                f"Agregalos en 'Base de SKUs' para que el colector sepa donde ir.",
                parent=self.root)


    # =========================================================================
    # DASHBOARD
    # =========================================================================
    def _build_dashboard(self):
        p = self.panel_dashboard
        canvas = tk.Canvas(p, bg=C["bg_dark"], highlightthickness=0)
        sb = tk.Scrollbar(p, orient="vertical", command=canvas.yview)
        self._dash_frame = tk.Frame(canvas, bg=C["bg_dark"])
        win = canvas.create_window((0,0), window=self._dash_frame, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        self._dash_frame.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
            lambda e: canvas.itemconfig(win, width=e.width))
        canvas.bind("<MouseWheel>",
            lambda e: canvas.yview_scroll(int(-1*(e.delta/120)),"units"))
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._build_dashboard_content()

    def _build_dashboard_content(self):
        f = self._dash_frame
        for w in f.winfo_children():
            w.destroy()
        pad = 24
        from datetime import datetime as _dt
        hora  = _dt.now().hour
        hola  = "Buenos días" if hora < 12 else ("Buenas tardes" if hora < 19 else "Buenas noches")
        fecha = _dt.now().strftime("%A %d de %B — %Y").capitalize()

        # ── Saludo ─────────────────────────────────────────────────────────
        top = tk.Frame(f, bg=C["bg_dark"])
        top.pack(fill="x", padx=pad, pady=(22,4))
        tk.Label(top, text=f"{hola}  —  Sistema de Picking",
                 font=("Segoe UI Black", 17), bg=C["bg_dark"],
                 fg=C["text_hi"]).pack(anchor="w")
        tk.Label(top, text=fecha, font=("Segoe UI", 10),
                 bg=C["bg_dark"], fg=C["text_mid"]).pack(anchor="w", pady=(3,0))
        tk.Frame(f, bg=C["border"], height=1).pack(fill="x", padx=pad, pady=(14,18))

        # ── Métricas ───────────────────────────────────────────────────────
        total_ped  = len(self.pedidos)
        impresas   = sum(1 for d in self.pedidos.values() if d.get("impreso"))
        pendientes = total_ped - impresas
        pct        = int(impresas / total_ped * 100) if total_ped else 0

        total_ml   = len(self._ml_pedidos) if hasattr(self,"_ml_pedidos") else 0
        ml_cuentas = len(self._ml_cuentas)  if hasattr(self,"_ml_cuentas")  else 0

        total_req  = {}
        for d in self.pedidos.values():
            for s,q in d.get("skus_requeridos",{}).items():
                total_req[s] = total_req.get(s,0) + q
        col_total  = sum(total_req.values())
        col_hecho  = sum(min(self.colecta_global.get(s,0),q) for s,q in total_req.items())
        pct_col    = int(col_hecho/col_total*100) if col_total else 0

        metricas = [
            ("Lote actual",   str(total_ped),         "pedidos cargados",     C["accent"]),
            ("Procesados",    str(impresas),           f"{pct}% completado",   C["success"]),
            ("Pendientes",    str(pendientes),         "por imprimir",         C["warning"] if pendientes else C["success"]),
            ("Colecta",       f"{pct_col}%",           f"{col_hecho}/{col_total} uds", C["accent2"]),
            ("Pedidos ML",    str(total_ml),           f"{ml_cuentas} cuenta(s) ML",   "#0891B2"),
            ("Servidor",      "Online" if RAILWAY_URL else "Sin URL",
                              "Railway",               C["success"] if RAILWAY_URL else C["danger"]),
        ]

        cards_f = tk.Frame(f, bg=C["bg_dark"])
        cards_f.pack(fill="x", padx=pad, pady=(0,18))
        COLS = 3
        for i,(tit, val, sub, color) in enumerate(metricas):
            r, c = divmod(i, COLS)
            card = tk.Frame(cards_f, bg=C["panel"])
            card.grid(row=r, column=c, sticky="nsew", padx=7, pady=7)

            accent_bar = tk.Frame(card, bg=color, width=4)
            accent_bar.pack(side="left", fill="y")

            inn = tk.Frame(card, bg=C["panel"], padx=14, pady=12)
            inn.pack(fill="both", expand=True)

            tk.Label(inn, text=tit, font=("Segoe UI", 9),
                     bg=C["panel"], fg=C["text_mid"]).pack(anchor="w")
            tk.Label(inn, text=val, font=("Segoe UI Black", 28),
                     bg=C["panel"], fg=color).pack(anchor="w", pady=(2,1))
            tk.Label(inn, text=sub, font=("Segoe UI", 9),
                     bg=C["panel"], fg=C["text_lo"]).pack(anchor="w")
        for c in range(COLS):
            cards_f.columnconfigure(c, weight=1)

        # ── Barra de progreso ──────────────────────────────────────────────
        tk.Frame(f, bg=C["border"], height=1).pack(fill="x", padx=pad, pady=(0,14))
        prog_f = tk.Frame(f, bg=C["panel"], padx=18, pady=14)
        prog_f.pack(fill="x", padx=pad, pady=(0,14))

        header_row = tk.Frame(prog_f, bg=C["panel"])
        header_row.pack(fill="x")
        tk.Label(header_row, text="Progreso del lote",
                 font=("Segoe UI Semibold", 10), bg=C["panel"],
                 fg=C["text_hi"]).pack(side="left")
        tk.Label(header_row, text=f"{pct}%  ·  {impresas} de {total_ped}",
                 font=("Segoe UI Semibold", 10), bg=C["panel"],
                 fg=C["success"] if pct==100 else C["accent"]).pack(side="right")

        bar_bg = tk.Frame(prog_f, bg=C["bar_bg"], height=12)
        bar_bg.pack(fill="x", pady=(10,4))
        bar_bg.pack_propagate(False)
        def _draw(e=None):
            w = bar_bg.winfo_width() or 600
            for ww in bar_bg.winfo_children(): ww.destroy()
            fill_w = max(2, int(w * pct / 100))
            clr = C["bar_ok"] if pct==100 else C["bar_fg"]
            tk.Frame(bar_bg, bg=clr, width=fill_w, height=12).pack(side="left")
        bar_bg.bind("<Configure>", _draw)
        self.root.after(120, _draw)

        # ── Estado de conexiones ───────────────────────────────────────────
        tk.Frame(f, bg=C["border"], height=1).pack(fill="x", padx=pad, pady=(0,14))
        tk.Label(f, text="Conexiones activas", font=("Segoe UI Semibold", 10),
                 bg=C["bg_dark"], fg=C["text_hi"]).pack(anchor="w", padx=pad, pady=(0,8))

        conex_f = tk.Frame(f, bg=C["bg_dark"])
        conex_f.pack(fill="x", padx=pad, pady=(0,14))

        cuentas_list = self._ml_cuentas if hasattr(self,"_ml_cuentas") else {}
        COLS_C = ["#7C3AED","#0891B2","#D97706","#DC2626","#059669"]
        conexiones = [("🌐  Servidor Railway", RAILWAY_URL or "No configurado",
                       C["success"] if RAILWAY_URL else C["danger"])]
        if cuentas_list:
            for i,(cid,nick) in enumerate(cuentas_list.items()):
                conexiones.append((f"🛒  ML: {nick}", "Conectado", COLS_C[i%5]))
        else:
            conexiones.append(("🛒  MercadoLibre","Sin cuentas", C["warning"]))
        imp = self.config.get("impresora","")
        conexiones.append(("🖨   Impresora", imp or "No configurada",
                           C["success"] if imp else C["warning"]))

        for tit, detalle, color in conexiones:
            row_f = tk.Frame(conex_f, bg=C["panel"], pady=6)
            row_f.pack(fill="x", pady=2)
            tk.Frame(row_f, bg=color, width=6, height=6).pack(side="left", padx=(10,8))
            tk.Label(row_f, text=tit, font=("Segoe UI Semibold", 9),
                     bg=C["panel"], fg=C["text_hi"],
                     width=24, anchor="w").pack(side="left")
            tk.Label(row_f, text=detalle, font=("Segoe UI", 9),
                     bg=C["panel"], fg=C["text_mid"]).pack(side="left", padx=6)

        # ── Acciones rápidas ───────────────────────────────────────────────
        tk.Frame(f, bg=C["border"], height=1).pack(fill="x", padx=pad, pady=(0,14))
        tk.Label(f, text="Acciones rápidas", font=("Segoe UI Semibold", 10),
                 bg=C["bg_dark"], fg=C["text_hi"]).pack(anchor="w", padx=pad, pady=(0,10))

        acc_f = tk.Frame(f, bg=C["bg_dark"])
        acc_f.pack(fill="x", padx=pad, pady=(0,22))

        for txt, accion, color in [
            ("📦  Pedidos ML",   "ml",      C["accent"]),
            ("🔍  Picking",      "picking", C["success"]),
            ("📄  Cargar PDF",   "pdf",     "#7C3AED"),
            ("⚙   Configuración","cfg",     C["text_lo"]),
        ]:
            def _cmd(a=accion):
                if   a == "pdf": self._switch_tab("picking"); self.cargar_pdf()
                elif a == "cfg": self.abrir_configuracion()
                else:            self._switch_tab(a)
            tk.Button(acc_f, text=txt,
                      font=("Segoe UI Semibold", 10),
                      bg=color, fg="white",
                      activebackground=C["accent2"], activeforeground="white",
                      relief="flat", cursor="hand2", padx=18, pady=10, bd=0,
                      command=_cmd).pack(side="left", padx=(0,10))

        # ── Timestamp ──────────────────────────────────────────────────────
        tk.Label(f, text=f"Actualizado a las {_dt.now().strftime('%H:%M:%S')}",
                 font=("Segoe UI", 8), bg=C["bg_dark"],
                 fg=C["text_lo"]).pack(anchor="e", padx=pad, pady=(0,20))

    def _refresh_dashboard(self):
        try:
            self._build_dashboard_content()
        except Exception:
            pass

    # =========================================================================
    # COLUMNA 1: FASE 1 (COLECTA EN DEPÓSITO)
    # =========================================================================
    def _build_fase1(self):
        p = self.col_fase1

        # Header
        hdr = tk.Frame(p, bg=C["accent"], pady=10)
        hdr.pack(fill="x")
        self.lbl_col1_header = tk.Label(
            hdr, text="📦  FASE 1: COLECTA EN DEPÓSITO",
            font=("Segoe UI Black", 10), bg=C["accent"], fg="white")
        self.lbl_col1_header.pack()

        # Contenedor con scroll
        canvas_frame = tk.Frame(p, bg=C["panel"])
        canvas_frame.pack(fill="both", expand=True, padx=0, pady=0)

        self.canvas_fase1 = tk.Canvas(canvas_frame, bg=C["panel"], highlightthickness=0)
        canvas = self.canvas_fase1
        scrollbar = tk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        self.frame_fase1 = tk.Frame(canvas, bg=C["panel"])

        win_id = canvas.create_window((0, 0), window=self.frame_fase1, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Actualizar scrollregion Y ajustar ancho del frame al canvas
        def _on_frame_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        def _on_canvas_configure(e):
            canvas.itemconfig(win_id, width=e.width)

        self.frame_fase1.bind("<Configure>", _on_frame_configure)
        canvas.bind("<Configure>", _on_canvas_configure)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Scroll con rueda del mouse — bind al canvas Y al frame raíz
        def _scroll(event):
            self.canvas_fase1.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind("<MouseWheel>", _scroll)
        canvas_frame.bind("<MouseWheel>", _scroll)
        p.bind("<MouseWheel>", _scroll)
        # Guardar referencia para bind recursivo posterior
        self._scroll_fase1_fn = _scroll

    # =========================================================================
    # COLUMNA 2: CONTROLES DE ESCANEO
    # =========================================================================
    def _build_control_panel(self):
        p = self.col_control

        # Sección: cargar PDF
        sec1 = tk.Frame(p, bg=C["panel"], padx=18, pady=15)
        sec1.pack(fill="x")

        tk.Label(sec1, text="ARCHIVO DE TRABAJO",
                 font=("Segoe UI", 8, "bold"), bg=C["panel"],
                 fg=C["text_lo"]).pack(anchor="w", pady=(0, 8))

        self.btn_cargar = tk.Button(
            sec1, text="📄  Cargar PDF",
            command=self.cargar_pdf, font=("Segoe UI Semibold", 10),
            bg=C["accent"], fg="white", activebackground=C["accent2"],
            activeforeground="white", relief="flat", cursor="hand2",
            pady=9, bd=0)
        self.btn_cargar.pack(fill="x")

        self.lbl_estado_pdf = tk.Label(
            sec1, text="Esperando archivo PDF...",
            fg=C["text_lo"], bg=C["panel"], font=FONT_SMALL,
            wraplength=360, justify="left")
        self.lbl_estado_pdf.pack(anchor="w", pady=(6, 0))

        tk.Frame(sec1, bg=C["border"], height=1).pack(fill="x", pady=(12, 8))

        # Estado de la base Excel (carga automática)
        excel_hdr = tk.Frame(sec1, bg=C["panel"])
        excel_hdr.pack(fill="x")
        tk.Label(excel_hdr, text="BASE SKU → NOMBRE",
                 font=("Segoe UI", 8, "bold"), bg=C["panel"],
                 fg=C["text_lo"]).pack(side="left", anchor="w")
        tk.Button(excel_hdr, text="Editar lista",
                  font=("Segoe UI", 7), bg=C["success"], fg="white",
                  activebackground="#059669", activeforeground="white",
                  relief="flat", cursor="hand2", padx=6, pady=2, bd=0,
                  command=self._form_agregar_sku_rapido).pack(side="right")
        tk.Button(excel_hdr, text="Ver todos",
                  font=("Segoe UI", 7), bg=C["accent2"], fg="white",
                  activebackground=C["accent"], activeforeground="white",
                  relief="flat", cursor="hand2", padx=6, pady=2, bd=0,
                  command=self.abrir_base_skus).pack(side="right")

        self.lbl_estado_excel = tk.Label(
            sec1, text="⏳  Cargando base de datos...",
            fg=C["text_lo"], bg=C["panel"], font=FONT_SMALL,
            wraplength=360, justify="left")
        self.lbl_estado_excel.pack(anchor="w", pady=(4, 0))

        tk.Frame(p, bg=C["border"], height=1).pack(fill="x", padx=18)

        # Sección: escaneo
        sec2 = tk.Frame(p, bg=C["panel"], padx=18, pady=15)
        sec2.pack(fill="x")

        tk.Label(sec2, text="ESCANEAR CÓDIGO DE BARRAS",
                 font=("Segoe UI", 8, "bold"), bg=C["panel"],
                 fg=C["text_lo"]).pack(anchor="w", pady=(0, 8))

        entry_frame = tk.Frame(sec2, bg=C["accent"], padx=2, pady=2)
        entry_frame.pack(fill="x")
        inner = tk.Frame(entry_frame, bg=C["card"])
        inner.pack(fill="x")

        self.entrada_sku = tk.Entry(
            inner, font=("Consolas", 12, "bold"), justify="center",
            bg=C["card"], fg=C["text_hi"], insertbackground=C["accent"],
            relief="flat", bd=0)
        self.entrada_sku.pack(fill="x", ipady=11, padx=10)
        self.entrada_sku.bind('<Return>', self.procesar_escaneo)
        self.entrada_sku.config(state="disabled")

        self.lbl_placeholder = tk.Label(
            sec2, text="Apuntá la pistola y escaneá",
            font=FONT_SMALL, bg=C["panel"], fg=C["text_lo"])
        self.lbl_placeholder.pack(anchor="w", pady=(6, 0))

        # Botones
        btn_frame = tk.Frame(sec2, bg=C["panel"])
        btn_frame.pack(fill="x", pady=(10, 0))

        self.btn_cancelar = tk.Button(
            btn_frame, text="✕  CANCELAR",
            state="disabled", font=FONT_BTN,
            bg=C["card"], fg=C["text_lo"],
            activebackground=C["danger"], activeforeground="white",
            relief="flat", cursor="hand2", pady=6, bd=0,
            command=self.cancelar_paquete)
        self.btn_cancelar.pack(side="left", fill="x", expand=True, padx=(0, 5))

        self.btn_saltar = tk.Button(
            btn_frame, text="⏭  SALTAR",
            state="disabled", font=FONT_BTN,
            bg=C["card"], fg=C["text_lo"],
            activebackground=C["warning"], activeforeground="white",
            relief="flat", cursor="hand2", pady=6, bd=0,
            command=self.saltar_sku_actual)
        self.btn_saltar.pack(side="right", fill="x", expand=True, padx=(5, 0))

        tk.Frame(p, bg=C["border"], height=1).pack(fill="x", padx=18)

        # Indicador de fase y botón transición
        sec_fase = tk.Frame(p, bg=C["panel"], padx=18, pady=12)
        sec_fase.pack(fill="x")

        self.lbl_fase_actual = tk.Label(
            sec_fase, text="● FASE 1: COLECTA EN DEPÓSITO",
            font=("Segoe UI Semibold", 10), bg=C["panel"], fg=C["accent"])
        self.lbl_fase_actual.pack(anchor="w")

        self.btn_fase2 = tk.Button(
            sec_fase, text="▶  PASAR A FASE 2",
            state="disabled", font=FONT_BTN,
            bg=C["card"], fg=C["text_lo"],
            activebackground=C["success"], activeforeground="white",
            relief="flat", cursor="hand2", pady=8, bd=0,
            command=self._switch_to_fase2)
        self.btn_fase2.pack(fill="x", pady=(8, 0))

        tk.Frame(p, bg=C["border"], height=1).pack(fill="x", padx=18)

        # Estado del pedido
        sec3 = tk.Frame(p, bg=C["panel"], padx=18, pady=12)
        sec3.pack(fill="x")

        self.lbl_num_caja = tk.Label(
            sec3, text="", font=("Segoe UI Black", 40, "bold"),
            bg=C["panel"], fg=C["accent"])
        self.lbl_num_caja.pack()

        self.lbl_resultado = tk.Label(
            sec3, text="", font=("Segoe UI Semibold", 10),
            bg=C["panel"], fg=C["text_hi"],
            justify="center", wraplength=360)
        self.lbl_resultado.pack(pady=(2, 0))

        self.lbl_imprimiendo = tk.Label(
            sec3, text="", font=FONT_SMALL,
            bg=C["panel"], fg=C["success"])
        self.lbl_imprimiendo.pack(pady=(4, 0))

    # =========================================================================
    # COLUMNA 3: VISTA PREVIA DE ETIQUETA
    # =========================================================================
    def _build_preview_panel(self):
        p = self.col_preview

        visor_hdr = tk.Frame(p, bg=C["panel"], height=38)
        visor_hdr.pack(fill="x")
        visor_hdr.pack_propagate(False)

        tk.Label(visor_hdr, text="▪  FASE 2: VISTA PREVIA DE ETIQUETA",
                 font=("Segoe UI", 8, "bold"), bg=C["panel"],
                 fg=C["text_lo"]).pack(side="left", padx=16, pady=10)

        self.lbl_visor_info = tk.Label(
            visor_hdr, text="",
            font=FONT_SMALL, bg=C["panel"], fg=C["accent"])
        self.lbl_visor_info.pack(side="right", padx=16, pady=10)

        canvas_frame = tk.Frame(p, bg=C["preview_bg"])
        canvas_frame.pack(fill="both", expand=True)

        self.canvas_preview = tk.Canvas(
            canvas_frame, bg=C["preview_bg"],
            highlightthickness=0, cursor="crosshair")
        self.canvas_preview.pack(fill="both", expand=True)
        self.canvas_preview.bind("<Configure>", self._on_canvas_resize)
        self._mostrar_placeholder_visor()

    def _on_canvas_resize(self, event):
        if self.imagen_actual is None:
            self._mostrar_placeholder_visor()

    def _mostrar_placeholder_visor(self):
        self.canvas_preview.delete("all")
        w = self.canvas_preview.winfo_width() or 600
        h = self.canvas_preview.winfo_height() or 500
        self.canvas_preview.create_line(w//2-30, h//2, w//2+30, h//2,
                                        fill=C["border"], width=1)
        self.canvas_preview.create_line(w//2, h//2-30, w//2, h//2+30,
                                        fill=C["border"], width=1)
        self.canvas_preview.create_text(
            w//2, h//2 + 50,
            text="La etiqueta aparecerá aquí",
            fill=C["text_lo"], font=("Segoe UI", 11), justify="center")

    # =========================================================================
    # MÉTODOS DE LÓGICA
    # =========================================================================

    def _texto_impresora(self):
        imp = self.config.get("impresora", "")
        return f"🖨  {imp}" if imp else "⚠  Sin impresora"

    def _autocargar_excel(self):
        """Carga automáticamente el Excel guardado en config al iniciar."""
        ruta = self.config.get("excel_path", "").strip()
        if not ruta:
            self.lbl_estado_excel.config(
                text="⚠  Sin base Excel. Configurá la ruta en ⚙ Configurar.",
                fg=C["warning"])
            return
        if not os.path.exists(ruta):
            self.lbl_estado_excel.config(
                text=f"❌  Archivo no encontrado: {os.path.basename(ruta)}",
                fg=C["danger"])
            return
        self.lbl_estado_excel.config(
            text=f"⏳  Cargando {os.path.basename(ruta)}...", fg=C["accent"])
        self.root.update_idletasks()
        self._cargar_excel_desde_ruta(ruta)

    def _cargar_excel_desde_ruta(self, ruta):
        """Lee el Excel en un thread para no bloquear la UI."""
        import threading
        def _worker():
            db = leer_excel_skus(ruta)
            self.root.after(0, lambda: self._excel_cargado(db, ruta))
        threading.Thread(target=_worker, daemon=True).start()

    def _excel_cargado(self, db, ruta):
        """Callback en el hilo principal cuando termina la carga del Excel."""
        self.db_nombres = db
        if db:
            self.lbl_estado_excel.config(
                text=f"✅  {len(db)} productos cargados — {os.path.basename(ruta)}",
                fg=C["success"])
        else:
            self.lbl_estado_excel.config(
                text=f"⚠  No se encontraron SKUs en {os.path.basename(ruta)}",
                fg=C["warning"])
        if self.pedidos:
            if self.fase_actual == 1:
                self._dibujar_fase1()
            else:
                self._dibujar_fase2()

    def cargar_base_excel(self):
        """Carga manual — mantenido por compatibilidad, ya no expuesto en UI."""
        if not _OPENPYXL_OK:
            messagebox.showerror(
                "Módulo faltante",
                "Instalá openpyxl para usar esta función:\n\npip install openpyxl",
                parent=self.root)
            return
        ruta = filedialog.askopenfilename(
            title="Seleccionar base de datos Excel",
            filetypes=[("Archivos Excel", "*.xlsx *.xls *.xlsm")])
        if not ruta:
            return
        self._cargar_excel_desde_ruta(ruta)

    def _nombre_sku(self, sku):
        """Devuelve el nombre del producto para un SKU: Excel primero, luego PDF."""
        if sku in self.db_nombres:
            entrada = self.db_nombres[sku]
            # Soporta tanto el formato nuevo (dict) como el viejo (str)
            return entrada["nombre"] if isinstance(entrada, dict) else entrada
        return self.sku_descripciones.get(sku, "")

    def _ubicacion_sku(self, sku):
        """Devuelve (pasillo, estanteria) para un SKU desde el Excel."""
        if sku in self.db_nombres:
            entrada = self.db_nombres[sku]
            if isinstance(entrada, dict):
                return entrada.get("pasillo", ""), entrada.get("estanteria", "")
        return "", ""

    def _ubicacion_texto(self, sku):
        """Devuelve un string compacto de ubicación: 'Pasillo 1 · Estantería A4'"""
        pasillo, estanteria = self._ubicacion_sku(sku)
        if pasillo and estanteria:
            return f"📍 {pasillo}  ·  {estanteria}"
        elif pasillo:
            return f"📍 {pasillo}"
        return ""

    def abrir_configuracion(self):
        VentanaConfiguracion(self.root, self.config, self._on_config_guardada)

    def _verificar_primera_vez(self):
        """Solo muestra bienvenida si NO hay impresora configurada."""
        if not self.config.get("impresora", "").strip():
            self._mostrar_bienvenida()

    def _mostrar_bienvenida(self):
        win = tk.Toplevel(self.root)
        win.title("Bienvenido")
        win.geometry("500x380")
        win.resizable(False, False)
        win.config(bg=C["bg_dark"])
        win.grab_set()
        win.lift()

        hdr = tk.Frame(win, bg=C["accent"], pady=16)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Bienvenido al Sistema de Picking",
                 font=("Segoe UI Black", 14), bg=C["accent"], fg="white").pack()
        tk.Label(hdr, text="Everest Group  |  MercadoLibre Uruguay",
                 font=("Segoe UI", 9), bg=C["accent"], fg="#BFDBFE").pack(pady=(2,0))

        body = tk.Frame(win, bg=C["bg_dark"], padx=30, pady=20)
        body.pack(fill="both", expand=True)

        tk.Label(body,
                 text="Es la primera vez que abris la app.\n"
                      "Necesitamos configurar algunas cosas antes de empezar.",
                 font=("Segoe UI", 11), bg=C["bg_dark"], fg=C["text_hi"],
                 justify="center", wraplength=420).pack(pady=(0, 18))

        for icon, titulo_item, desc in [
            ("impresora de etiquetas", "Obligatorio para imprimir las etiquetas ML"),
            ("codigo de supervisor",   "Para autorizar el paso a Fase 2 (Packing)"),
            ("base de SKUs",           "Importar Excel o agregar uno por uno (se puede despues)"),
        ]:
            row = tk.Frame(body, bg=C["card"], pady=8)
            row.pack(fill="x", pady=3)
            col = tk.Frame(row, bg=C["card"])
            col.pack(side="left", fill="x", expand=True, padx=12)
            tk.Label(col, text=icon.upper(), font=("Segoe UI Semibold", 10),
                     bg=C["card"], fg=C["text_hi"], anchor="w").pack(anchor="w")
            tk.Label(col, text=titulo_item, font=("Segoe UI", 8),
                     bg=C["card"], fg=C["text_mid"], anchor="w").pack(anchor="w")

        sep = tk.Frame(body, bg=C["border"], height=1)
        sep.pack(fill="x", pady=(16, 10))
        btn_row = tk.Frame(body, bg=C["bg_dark"])
        btn_row.pack(fill="x")

        tk.Button(btn_row, text="Saltar por ahora",
                  font=FONT_BODY, bg=C["panel"], fg=C["text_mid"],
                  relief="flat", cursor="hand2", padx=14, pady=7, bd=0,
                  command=win.destroy).pack(side="left")
        tk.Button(btn_row, text="Ir a Configuracion  →",
                  font=("Segoe UI Semibold", 11), bg=C["accent"], fg="white",
                  activebackground=C["accent2"], activeforeground="white",
                  relief="flat", cursor="hand2", padx=20, pady=7, bd=0,
                  command=lambda: [win.destroy(), self.abrir_configuracion()]).pack(side="right")

    def _form_agregar_sku_rapido(self):
        """Formulario rapido para agregar un SKU sin abrir la ventana completa."""
        win = tk.Toplevel(self.root)
        win.title("Agregar SKU")
        win.geometry("420x340")
        win.resizable(False, False)
        win.config(bg=C["bg_dark"])
        win.grab_set()

        hdr = tk.Frame(win, bg=C["success"], pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="+ Agregar SKU a la base de datos",
                 font=("Segoe UI Semibold", 11), bg=C["success"], fg="white").pack()

        footer = tk.Frame(win, bg=C["bg_dark"], padx=20, pady=10)
        footer.pack(side="bottom", fill="x")
        lbl_err = tk.Label(footer, text="", font=FONT_SMALL,
                           bg=C["bg_dark"], fg=C["danger"])
        lbl_err.pack(anchor="w", pady=(0, 6))
        tk.Frame(footer, bg=C["border"], height=1).pack(fill="x", pady=(0, 8))
        btn_row = tk.Frame(footer, bg=C["bg_dark"])
        btn_row.pack(fill="x")

        body = tk.Frame(win, bg=C["bg_dark"], padx=20, pady=14)
        body.pack(fill="both", expand=True)

        campos = [
            ("SKU *",       "sku",        "Ej: 7478"),
            ("Nombre *",    "nombre",     "Ej: Lima Para Pies"),
            ("Pasillo",     "pasillo",    "Ej: Pasillo 1 - Belleza"),
            ("Estanteria",  "estanteria", "Ej: A6"),
        ]
        vars_   = {}
        entries = {}
        for label, key, placeholder in campos:
            tk.Label(body, text=label, font=("Segoe UI", 8, "bold"),
                     bg=C["bg_dark"], fg=C["text_mid"]).pack(anchor="w", pady=(6, 2))
            wrap = tk.Frame(body, bg=C["border"], padx=1, pady=1)
            wrap.pack(fill="x")
            inner = tk.Frame(wrap, bg=C["card"]); inner.pack(fill="x")
            var = tk.StringVar()
            vars_[key] = var
            e = tk.Entry(inner, textvariable=var, font=FONT_BODY,
                         bg=C["card"], fg=C["text_hi"],
                         insertbackground=C["accent"], relief="flat", bd=0)
            e.pack(fill="x", ipady=5, padx=6)
            entries[key] = e

        def _guardar():
            sku    = vars_["sku"].get().strip().upper()
            nombre = vars_["nombre"].get().strip()
            if not sku:
                lbl_err.config(text="El SKU es obligatorio."); return
            if not nombre:
                lbl_err.config(text="El nombre es obligatorio."); return
            import sys as _sys
            base = os.path.dirname(_sys.executable) if getattr(_sys, "frozen", False) \
                   else os.path.dirname(os.path.abspath(__file__))
            db_path = os.path.join(base, "skus_db.json")
            try:
                db = json.load(open(db_path, encoding="utf-8")) if os.path.exists(db_path) else {}
            except Exception:
                db = {}
            db[sku] = {
                "nombre":     nombre,
                "pasillo":    vars_["pasillo"].get().strip(),
                "estanteria": vars_["estanteria"].get().strip(),
            }
            with open(db_path, "w", encoding="utf-8") as f:
                json.dump(db, f, ensure_ascii=False, indent=2)
            self.db_nombres[sku] = db[sku]
            self.lbl_estado_excel.config(
                text=f"BD: {len(self.db_nombres)} SKUs cargados", fg=C["success"])
            lbl_err.config(text="")
            for v in vars_.values():
                v.set("")
            entries["sku"].focus()
            messagebox.showinfo("Guardado", f"SKU '{sku}' agregado.", parent=win)

        def _guardar_cerrar():
            sku    = vars_["sku"].get().strip().upper()
            nombre = vars_["nombre"].get().strip()
            if not sku:
                lbl_err.config(text="El SKU es obligatorio."); return
            if not nombre:
                lbl_err.config(text="El nombre es obligatorio."); return
            _guardar()
            win.destroy()

        tk.Button(btn_row, text="Guardar y agregar otro",
                  font=FONT_SMALL, bg=C["panel"], fg=C["text_mid"],
                  relief="flat", cursor="hand2", padx=12, pady=7, bd=0,
                  command=_guardar).pack(side="left")
        tk.Button(btn_row, text="Guardar y cerrar",
                  font=("Segoe UI Semibold", 10), bg=C["success"], fg="white",
                  activebackground="#059669", activeforeground="white",
                  relief="flat", cursor="hand2", padx=16, pady=7, bd=0,
                  command=_guardar_cerrar).pack(side="right")

        for e in entries.values():
            e.bind("<Return>", lambda ev: _guardar())
        entries["sku"].focus()

    def abrir_base_skus(self):
        VentanaBaseSKUs(self.root, self._on_db_interna_actualizada)

    def _on_db_interna_actualizada(self, nueva_db):
        """Callback cuando se guarda la BD interna — merge con db_nombres."""
        for sku, datos in nueva_db.items():
            # La BD interna tiene prioridad
            self.db_nombres[sku] = datos
        n = len(nueva_db)
        self.lbl_estado_excel.config(
            text=f"✅  BD interna: {n} SKUs  +  Excel cargado",
            fg=C["success"])
        if self.pedidos:
            if self.fase_actual == 1:
                self._dibujar_fase1()
            else:
                self._dibujar_fase2()

    def _cargar_db_interna(self):
        """Carga la BD interna al iniciar y la fusiona con db_nombres."""
        import sys as _sys
        _base = os.path.dirname(_sys.executable) if getattr(_sys,'frozen',False) else os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(_base, "skus_db.json")
        if os.path.exists(db_path):
            try:
                with open(db_path, "r", encoding="utf-8") as f:
                    db = json.load(f)
                for sku, datos in db.items():
                    self.db_nombres[sku] = datos
            except Exception:
                pass
        VentanaConfiguracion(self.root, self.config, self._on_config_guardada)

    def _on_config_guardada(self, nueva):
        ruta_anterior = self.config.get("excel_path", "")
        nueva["servidor_nube"] = RAILWAY_URL
        self.config = nueva
        guardar_config(nueva)
        self.lbl_printer.config(text=self._texto_impresora())
        # Si cambió la ruta del Excel, recargarlo inmediatamente
        ruta_nueva = nueva.get("excel_path", "")
        if ruta_nueva and ruta_nueva != ruta_anterior:
            self._cargar_excel_desde_ruta(ruta_nueva)

    def cargar_pdf(self):
        ruta = filedialog.askopenfilename(filetypes=[("Archivos PDF", "*.pdf")])
        if ruta:
            self.ruta_pdf = ruta
            nombre = os.path.basename(ruta)
            self.lbl_estado_pdf.config(
                text=f"⏳  Procesando {nombre}...", fg=C["accent"])
            self.root.update_idletasks()
            # Deshabilitar botón mientras procesa
            self.btn_cargar.config(state="disabled")
            import threading
            threading.Thread(target=self._mapear_pdf_thread, daemon=True).start()

    def _mapear_pdf_thread(self):
        """Parsea el PDF en un thread y notifica al hilo principal al terminar."""
        try:
            resultado = self._parsear_pdf(self.ruta_pdf)
            self.root.after(0, lambda: self._pdf_cargado(resultado))
        except Exception as e:
            self.root.after(0, lambda: self._pdf_error(str(e)))

    def mapear_skus_desde_pdf(self):
        """Punto de entrada legacy — ahora delega al thread."""
        self._limpiar_estado_previo()
        import threading
        threading.Thread(target=self._mapear_pdf_thread, daemon=True).start()

    def _limpiar_estado_previo(self):
        self.pedidos.clear()
        self.pedido_en_proceso = None
        self.fase_actual       = 1
        self.colecta_global    = {}
        self.sku_descripciones = {}
        self.lbl_resultado.config(text="")
        self.lbl_num_caja.config(text="")
        self.lbl_imprimiendo.config(text="")
        self._limpiar_fase1()
        self._mostrar_placeholder_visor()

    def _parsear_pdf(self, ruta_pdf):
        """
        Parsea el PDF y devuelve (pedidos, sku_descripciones).
        Corre en un thread secundario — NO toca ningún widget de tkinter.
        """
        pedidos         = {}
        sku_descripciones = {}

        documento      = fitz.open(ruta_pdf)
        inicio_resumen = len(documento)

        for n in range(len(documento)):
            txt = documento.load_page(n).get_text("text")
            if "Despacha tus" in txt or "The following table" in txt:
                inicio_resumen = n
                break

        id_to_page = {}

        for n in range(inicio_resumen):
            txt  = documento.load_page(n).get_text("text")
            ids  = re.findall(r'20000\d{11}', txt.replace(" ", "").replace("\n", ""))
            desc = self._extraer_descripcion_pagina(txt)
            pedidos[n + 1] = {
                "pagina": n + 1, "skus_requeridos": {},
                "skus_escaneados": {}, "impreso": False,
                "descripcion": desc
            }
            for id_val in ids:
                id_to_page[id_val] = n + 1
            for linea in txt.split('\n'):
                if "SKU:" in linea:
                    p = linea.split("SKU:")
                    if len(p) > 1 and p[1].strip():
                        sku = p[1].strip().split()[0].upper()
                        pedidos[n + 1]["skus_requeridos"][sku] = 1

        for n in range(inicio_resumen, len(documento)):
            page      = documento.load_page(n)
            page_w    = page.rect.width
            split_x   = page_w * 0.42

            raw_blocks = [b for b in page.get_text("blocks")
                          if b[6] == 0 and b[4].strip()]

            left_blocks = sorted(
                [(b[1], b[3], b[4].strip()) for b in raw_blocks if b[2] <= split_x],
                key=lambda b: b[0])
            right_blocks = sorted(
                [(b[1], b[3], b[4].strip()) for b in raw_blocks if b[0] >= split_x * 0.85],
                key=lambda b: b[0])

            order_intervals = []
            for idx, (ly0, ly1, ltext) in enumerate(left_blocks):
                ids   = re.findall(r'20000\d{11}', ltext.replace(" ", "").replace("\n", ""))
                pages = {id_to_page[i] for i in ids if i in id_to_page}
                if pages:
                    y_next = left_blocks[idx + 1][0] if idx + 1 < len(left_blocks) else 1e9
                    order_intervals.append((ly0 - 5, y_next, pages))

            for ry0, ry1, rtext in right_blocks:
                active_pages = set()
                for oy0, oy_end, pages in order_intervals:
                    if oy0 <= ry0 < oy_end:
                        active_pages = pages
                        break
                if not active_pages:
                    if order_intervals:
                        active_pages = order_intervals[-1][2]
                    else:
                        continue

                lines    = [l.strip() for l in rtext.split('\n') if l.strip()]
                cur_desc = ""
                cur_sku  = ""

                for line in lines:
                    sku_m  = re.match(r'SKU:\s*(\S+)', line)
                    qty_m  = re.search(r'(?:Cantidad|Unidades):\s*(\d+)', line, re.IGNORECASE)
                    attr_m = re.match(r'^[A-Za-záéíóúÁÉÍÓÚñÑ]{3,}\s*:', line)

                    if sku_m:
                        cur_sku = sku_m.group(1).strip().upper()
                        if cur_desc and cur_sku not in sku_descripciones:
                            sku_descripciones[cur_sku] = cur_desc
                        for pg in active_pages:
                            if cur_sku not in pedidos[pg]["skus_requeridos"]:
                                pedidos[pg]["skus_requeridos"][cur_sku] = 1
                            if cur_desc and not pedidos[pg].get("descripcion"):
                                pedidos[pg]["descripcion"] = cur_desc
                        cur_desc = ""
                    elif qty_m:
                        qty = int(qty_m.group(1))
                        if cur_sku:
                            for pg in active_pages:
                                pedidos[pg]["skus_requeridos"][cur_sku] = qty
                    elif not attr_m and len(line) > 5 and not re.match(r'^\d', line):
                        cleaned = re.sub(r'^[□☐✓✗\s]+', '', line).strip()
                        if cleaned and not re.search(r'20000\d{9}', cleaned.replace(" ", "")):
                            cur_desc = cleaned

        documento.close()
        pedidos = {k: v for k, v in pedidos.items() if v["skus_requeridos"]}

        for d in pedidos.values():
            desc = d.get("descripcion", "").strip()
            if desc:
                for sku in d["skus_requeridos"]:
                    if sku not in sku_descripciones:
                        sku_descripciones[sku] = desc

        return pedidos, sku_descripciones

    def _pdf_cargado(self, resultado):
        """Callback en hilo principal — recibe los datos parseados y actualiza UI."""
        pedidos, sku_descripciones = resultado
        self.pedidos           = pedidos
        self.sku_descripciones = sku_descripciones

        cant = len(self.pedidos)
        self.lbl_estado_pdf.config(
            text=f"✅  {cant} paquetes cargados. Subiendo a Railway…",
            fg=C["accent"])
        self.btn_cargar.config(state="normal")
        self.actualizar_contador_global()
        self._dibujar_fase1()
        self.entrada_sku.config(state="normal")
        self.entrada_sku.focus()
        self.btn_fase2.config(state="normal", bg=C["success"], fg="white")
        self.lbl_fase_actual.config(
            text="● FASE 1: COLECTA EN DEPOSITO", fg=C["accent"])
        self.lbl_col1_header.config(
            text="FASE 1: COLECTA EN DEPOSITO")
        self._exportar_estado_movil()
        # Siempre subir a Railway y arrancar polling
        self._subir_a_nube_async()
        self.root.after(5000, self._sincronizar_desde_nube)

    def _pdf_error(self, msg):
        self.btn_cargar.config(state="normal")
        self.lbl_estado_pdf.config(text="❌  Error al cargar el PDF.", fg=C["danger"])

    def _limpiar_fase1(self):
        """Limpia el contenedor de FASE 1"""
        for w in self.frame_fase1.winfo_children():
            w.destroy()

    def _dibujar_fase1(self):
        """Lista consolidada de SKUs agrupados por pasillo."""
        self._limpiar_fase1()
        self.fase1_items = {}

        # Sumar cantidades de cada SKU entre todos los pedidos
        total_req = {}
        for d in self.pedidos.values():
            for sku, qty in d["skus_requeridos"].items():
                total_req[sku] = total_req.get(sku, 0) + qty

        # Encabezado resumen
        hdr_info = tk.Frame(self.frame_fase1, bg=C["panel"], pady=6)
        hdr_info.pack(fill="x", padx=6)
        tk.Label(hdr_info,
                 text=f"{len(total_req)} SKUs distintos  ·  {sum(total_req.values())} unidades en total",
                 font=FONT_SMALL, bg=C["panel"], fg=C["text_mid"]).pack(anchor="w")

        # Agrupar SKUs por pasillo
        grupos = {}   # { pasillo: [ (sku, qty, desc, estanteria) ] }
        SIN_PASILLO = "Sin ubicación en Excel"
        for sku in sorted(total_req.keys()):
            pasillo, estanteria = self._ubicacion_sku(sku)
            key = pasillo if pasillo else SIN_PASILLO
            grupos.setdefault(key, []).append((sku, total_req[sku], self._nombre_sku(sku), estanteria))

        # Ordenar grupos: pasillos numerados primero, luego el resto
        def _orden_pasillo(nombre):
            m = re.search(r'(\d+)', nombre)
            return (0, int(m.group(1)), nombre) if m else (1, 0, nombre)

        for pasillo in sorted(grupos.keys(), key=_orden_pasillo):
            items_pasillo = grupos[pasillo]

            # ── Encabezado de pasillo ─────────────────────────────────────────
            sep = tk.Frame(self.frame_fase1, bg=C["border"], height=1)
            sep.pack(fill="x", padx=6, pady=(8, 0))

            hdr_p = tk.Frame(self.frame_fase1, bg=C["bar_bg"], pady=5)
            hdr_p.pack(fill="x", padx=6)

            n_skus  = len(items_pasillo)
            n_units = sum(q for _, q, _, _ in items_pasillo)
            tk.Label(hdr_p,
                     text=f"📦  {pasillo.upper()}",
                     font=("Segoe UI Black", 9),
                     bg=C["bar_bg"], fg=C["accent"]).pack(side="left", padx=(8, 12))
            tk.Label(hdr_p,
                     text=f"{n_skus} SKU{'s' if n_skus>1 else ''}  ·  {n_units} ud.",
                     font=("Segoe UI", 8),
                     bg=C["bar_bg"], fg=C["text_mid"]).pack(side="left")

            # ── Filas de SKU dentro del pasillo ──────────────────────────────
            for sku, qty, desc, estanteria in items_pasillo:
                fila = tk.Frame(self.frame_fase1, bg=C["card"], pady=6)
                fila.pack(fill="x", padx=6, pady=1)

                # Checkbox — izquierda
                chk = tk.Label(fila, text="○", font=("Segoe UI", 15),
                               bg=C["card"], fg=C["text_lo"], width=3)
                chk.pack(side="left", padx=(6, 4))

                # Bloque central
                bloque = tk.Frame(fila, bg=C["card"])
                bloque.pack(side="left", fill="x", expand=True, padx=(2, 4))

                if desc:
                    tk.Label(bloque, text=desc,
                             font=("Segoe UI Semibold", 10),
                             bg=C["card"], fg=C["text_hi"],
                             anchor="w", wraplength=250, justify="left").pack(anchor="w")

                # SKU + contador en la misma línea
                sku_row = tk.Frame(bloque, bg=C["card"])
                sku_row.pack(anchor="w")
                tk.Label(sku_row, text=sku,
                         font=("Consolas", 9),
                         bg=C["card"], fg=C["text_mid"],
                         anchor="w").pack(side="left")
                lbl_cnt = tk.Label(sku_row, text=f"  0 / {qty}",
                                   font=("Segoe UI Semibold", 10),
                                   bg=C["card"], fg=C["accent"],
                                   anchor="w")
                lbl_cnt.pack(side="left", padx=(6, 0))

                # Estantería (solo si la tiene, el pasillo ya está en el header)
                if estanteria:
                    tk.Label(bloque, text=f"🗂  {estanteria}",
                             font=("Segoe UI", 8),
                             bg=C["card"], fg=C["accent2"],
                             anchor="w").pack(anchor="w")

                self.fase1_items[sku] = {"checkbox": chk, "lbl_cnt": lbl_cnt, "req": qty}

        self._bind_scroll_recursivo(self.frame_fase1)

    def _dibujar_fase2(self):
        """Lista por caja con checkmarks por SKU para el armado de pedidos."""
        self._limpiar_fase1()

        for pn in sorted(self.pedidos.keys()):
            d = self.pedidos[pn]
            d["skus_items"] = {}

            pedido_frame = tk.Frame(self.frame_fase1, bg=C["card"], relief="flat", bd=1)
            pedido_frame.pack(fill="x", pady=6, padx=0)

            header = tk.Frame(pedido_frame, bg=C["accent"], height=28)
            header.pack(fill="x")
            header.pack_propagate(False)

            nombre = self._nombre_pedido(pn)
            tk.Label(header, text=nombre, font=("Segoe UI Black", 11),
                     bg=C["accent"], fg="white",
                     wraplength=280, justify="left").pack(side="left", padx=10, pady=4)

            lbl_estado = tk.Label(header, text="○ Pendiente",
                                  font=("Segoe UI", 8), bg=C["accent"],
                                  fg=C["text_mid"])
            lbl_estado.pack(side="right", padx=10, pady=4)
            d["_lbl_estado"] = lbl_estado

            skus_frame = tk.Frame(pedido_frame, bg=C["card"], padx=10, pady=8)
            skus_frame.pack(fill="x")

            for sku, cant in d["skus_requeridos"].items():
                fila = tk.Frame(skus_frame, bg=C["card"])
                fila.pack(fill="x", pady=3)

                chk = tk.Label(fila, text="○", font=("Segoe UI", 12),
                               bg=C["card"], fg=C["text_lo"], width=3)
                chk.pack(side="left")
                d["skus_items"][sku] = {"checkbox": chk}

                # Cantidad a la derecha
                tk.Label(fila, text=f"×{cant}", font=("Segoe UI Semibold", 10),
                         bg=C["card"], fg=C["accent"]).pack(side="right", padx=(0, 4))

                # Bloque central: nombre + SKU
                bloque = tk.Frame(fila, bg=C["card"])
                bloque.pack(side="left", padx=(5, 0), fill="x", expand=True)

                nombre_prod = self._nombre_sku(sku)
                if nombre_prod:
                    tk.Label(bloque, text=nombre_prod,
                             font=("Segoe UI Semibold", 9),
                             bg=C["card"], fg=C["text_hi"],
                             anchor="w", wraplength=220, justify="left").pack(anchor="w")
                    tk.Label(bloque, text=sku,
                             font=("Consolas", 8),
                             bg=C["card"], fg=C["text_mid"], anchor="w").pack(anchor="w")
                else:
                    tk.Label(bloque, text=sku,
                             font=("Consolas", 10, "bold"),
                             bg=C["card"], fg=C["text_hi"],
                             width=16, anchor="w").pack(anchor="w")

        self._bind_scroll_recursivo(self.frame_fase1)

    def _animar_exito(self):
        """Flash verde en el panel de control al escanear correctamente."""
        self.col_control.config(bg=C["success"])
        self.root.after(120, lambda: self.col_control.config(bg=C["success"]))
        self.root.after(240, lambda: self.col_control.config(bg=C["panel"]))

    def _animar_error(self):
        """Flash rojo en el panel de control al escanear un SKU incorrecto."""
        self.col_control.config(bg=C["danger"])
        self.root.after(120, lambda: self.col_control.config(bg=C["danger"]))
        self.root.after(240, lambda: self.col_control.config(bg=C["panel"]))

    def _animar_checkmark(self, sku):
        """Animacion del checkmark en la lista Fase 1 al colectar un SKU."""
        if not hasattr(self,"fase1_items") or sku not in self.fase1_items:
            return
        item = self.fase1_items[sku]
        chk  = item.get("checkbox")
        cnt  = item.get("lbl_cnt")
        if not chk or not chk.winfo_exists():
            return

        def _pulse(step=0):
            if not chk.winfo_exists():
                return
            colores = [C["success"], "#FFFFFF", C["success"], C["success"]]
            if step < len(colores):
                chk.config(fg=colores[step])
                if cnt and cnt.winfo_exists():
                    cnt.config(fg=colores[step])
                self.root.after(90, lambda: _pulse(step+1))

        _pulse()

    def _animar_lote_completo(self):
        """Animacion al completar todo el lote — parpadeo del contador."""
        def _toggle(n=0, activo=True):
            if n >= 8:
                self.lbl_contador.config(fg=C["success"])
                return
            color = C["success"] if activo else C["warning"]
            self.lbl_contador.config(fg=color)
            self.root.after(180, lambda: _toggle(n+1, not activo))
        _toggle()

    def _animar_pedido_completo(self):
        """Flash verde completo al terminar un pedido en Fase 2."""
        self.col_control.config(bg=C["success"])
        self.lbl_num_caja.config(fg="#FFFFFF")
        def _restaurar():
            self.col_control.config(bg=C["panel"])
            self.lbl_num_caja.config(fg=C["success"])
        self.root.after(400, _restaurar)
        self.canvas_fase1.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _bind_scroll_recursivo(self, widget):
        """Enlaza scroll a todos los widgets hijos recursivamente."""
        fn = getattr(self, '_scroll_fase1_fn', self._on_scroll_fase1)
        try:
            widget.bind("<MouseWheel>", fn)
        except Exception:
            pass
        for child in widget.winfo_children():
            self._bind_scroll_recursivo(child)

    def actualizar_contador_global(self):
        pend = sum(1 for d in self.pedidos.values() if not d["impreso"])
        total = len(self.pedidos)
        done  = total - pend
        self.lbl_contador.config(
            text=f"{pend} pendientes   ·   {done} impresas   ·   {total} total")
        return pend

    def cancelar_paquete(self):
        if self.pedido_en_proceso:
            p = self.pedido_en_proceso
            self.pedidos[p]["skus_escaneados"].clear()
            self._actualizar_estado_fase1(p)
            self.pedido_en_proceso = None
            self.lbl_resultado.config(
                text="Paquete cancelado.",
                fg=C["warning"])
            self.lbl_num_caja.config(text="")
            self.btn_cancelar.config(state="disabled", bg=C["card"], fg=C["text_lo"])
            self.btn_saltar.config(state="disabled", bg=C["card"], fg=C["text_lo"])
            self._mostrar_placeholder_visor()
            self.lbl_visor_info.config(text="")
            self.entrada_sku.focus()

    def saltar_sku_actual(self):
        """Marca el siguiente SKU incompleto como 'saltado'"""
        if not self.pedido_en_proceso:
            return

        pn = self.pedido_en_proceso
        d = self.pedidos[pn]

        sku_para_saltar = None
        for sku, req in d["skus_requeridos"].items():
            esc = d["skus_escaneados"].get(sku, 0)
            if esc < req:
                sku_para_saltar = sku
                break

        if sku_para_saltar:
            d["skus_escaneados"][sku_para_saltar] = d["skus_requeridos"][sku_para_saltar]
            self._actualizar_estado_fase1(pn)
            self.lbl_resultado.config(
                text=f"⏭  Se saltó '{sku_para_saltar}'.",
                fg=C["warning"])
            self._evaluar_pac(pn)
        else:
            self.lbl_resultado.config(
                text="Todos los artículos completados.",
                fg=C["success"])

        self.entrada_sku.focus()

    def procesar_escaneo(self, event):
        sku = self.entrada_sku.get().strip().upper()
        self.entrada_sku.delete(0, tk.END)
        if not sku or not self.pedidos:
            return

        if self.fase_actual == 1:
            self._procesar_escaneo_fase1(sku)
            return

        # ── FASE 2 ──────────────────────────────────────────────
        # PAC activo
        if self.pedido_en_proceso:
            self._agregar_sku_pac(sku)
            return

        # Buscar pedido
        destino = None
        for pn in sorted(self.pedidos.keys()):
            d = self.pedidos[pn]
            if d["impreso"]:
                continue
            if sku in d["skus_requeridos"]:
                req = d["skus_requeridos"][sku]
                esc = d["skus_escaneados"].get(sku, 0)
                if esc < req:
                    destino = pn
                    break

        if destino is None:
            self.lbl_resultado.config(
                text=f"⚠  SKU '{sku}' no encontrado.",
                fg=C["danger"])
            self.lbl_num_caja.config(text="⚠", fg=C["danger"])
            self._mostrar_placeholder_visor()
            self.lbl_visor_info.config(text="")
            # Parpadeo
            self.col_control.config(bg=C["danger"])
            self.root.after(200, lambda: self.col_control.config(bg=C["panel"]))
            return

        d = self.pedidos[destino]
        d["skus_escaneados"][sku] = d["skus_escaneados"].get(sku, 0) + 1
        total_req = sum(d["skus_requeridos"].values())
        self._actualizar_estado_fase1(destino)

        if total_req == 1:
            self._mostrar_caja_completa(destino)
            self.root.after(80, lambda: self._imprimir_automatico(destino))
        else:
            self.pedido_en_proceso = destino
            self.btn_cancelar.config(state="normal", bg=C["danger"], fg="white")
            self.btn_saltar.config(state="normal", bg=C["warning"], fg="white")
            self._evaluar_pac(destino)

    # =========================================================================
    # FASE 1: COLECTA
    # =========================================================================

    def _procesar_escaneo_fase1(self, sku):
        total_req = {}
        for d in self.pedidos.values():
            for s, q in d["skus_requeridos"].items():
                total_req[s] = total_req.get(s, 0) + q

        if sku not in total_req:
            self.lbl_resultado.config(
                text=f"⚠  '{sku}' no está en ningún pedido.", fg=C["danger"])
            self.lbl_num_caja.config(text="⚠", fg=C["danger"])
            self._animar_error()
            return

        required = total_req[sku]
        collected_prev = self.colecta_global.get(sku, 0)

        if collected_prev >= required:
            self.lbl_resultado.config(
                text=f"✅  '{sku}' ya colectado ({required}/{required}).",
                fg=C["warning"])
            self.lbl_num_caja.config(text="✔", fg=C["warning"])
            return

        self.colecta_global[sku] = collected_prev + 1
        collected = self.colecta_global[sku]

        self._actualizar_checkmarks_fase1(sku, total_req)
        # Exportar con debounce (máximo una vez cada 600ms)
        self._exportar_debounced()

        if collected >= required:
            self.lbl_resultado.config(
                text=f"✅  '{sku}' completado ({collected}/{required})",
                fg=C["success"])
            self.lbl_num_caja.config(text="✔", fg=C["success"])
            self._animar_exito()
            self._animar_checkmark(sku)
        else:
            self.lbl_resultado.config(
                text=f"SKU '{sku}': {collected}/{required} unidades",
                fg=C["text_hi"])
            self.lbl_num_caja.config(
                text=f"{collected}/{required}", fg=C["accent"])
            self._animar_exito()

        # Auto-transición si toda la colecta está completa
        todo_completo = all(
            self.colecta_global.get(s, 0) >= q for s, q in total_req.items()
        )
        if todo_completo:
            self.lbl_resultado.config(
                text="✅  ¡Colecta completa! Pasando a Fase 2...", fg=C["success"])
            self.lbl_num_caja.config(text="✔✔", fg=C["success"])
            self._animar_lote_completo()
            self.root.after(1200, self._ejecutar_transicion_fase2)

    def _actualizar_checkmarks_fase1(self, sku, total_req=None):
        if not hasattr(self, "fase1_items") or sku not in self.fase1_items:
            return

        item      = self.fase1_items[sku]
        required  = item["req"]
        collected = self.colecta_global.get(sku, 0)

        item["lbl_cnt"].config(text=f"  {collected} / {required}")

        if collected >= required:
            item["checkbox"].config(text="✔", fg=C["success"])
            item["lbl_cnt"].config(fg=C["success"])
        else:
            item["checkbox"].config(text="○", fg=C["text_lo"])
            item["lbl_cnt"].config(fg=C["accent"])

    def _switch_to_fase2(self):
        total_req = {}
        for d in self.pedidos.values():
            for s, q in d["skus_requeridos"].items():
                total_req[s] = total_req.get(s, 0) + q

        incompletos = [
            (s, self.colecta_global.get(s, 0), q)
            for s, q in total_req.items()
            if self.colecta_global.get(s, 0) < q
        ]

        if incompletos:
            # Requiere código de supervisor para saltear la Fase 1
            codigo = self.config.get("codigo_supervisor", "1234")
            VentanaCodigoSupervisor(
                self.root,
                codigo_correcto=codigo,
                incompletos=incompletos,
                callback_ok=self._ejecutar_transicion_fase2)
        else:
            self._ejecutar_transicion_fase2()

    def _ejecutar_transicion_fase2(self):
        self.fase_actual = 2

        for d in self.pedidos.values():
            d["skus_escaneados"].clear()
            d["impreso"] = False

        self.pedido_en_proceso = None
        self._dibujar_fase2()
        self.actualizar_contador_global()

        self.lbl_fase_actual.config(
            text="● FASE 2: ARMADO DE PEDIDOS", fg=C["success"])
        self.btn_fase2.config(
            state="disabled", bg=C["card"], fg=C["text_lo"])
        self.lbl_col1_header.config(
            text="📦  FASE 2: ARMADO DE PEDIDOS")
        self.lbl_resultado.config(
            text="Escaneá el SKU para asignar a su CAJA.", fg=C["text_mid"])
        self.lbl_num_caja.config(text="")
        self.lbl_imprimiendo.config(text="")
        self.btn_cancelar.config(state="disabled", bg=C["card"], fg=C["text_lo"])
        self.btn_saltar.config(state="disabled", bg=C["card"], fg=C["text_lo"])
        self._mostrar_placeholder_visor()
        self.lbl_visor_info.config(text="")
        self.entrada_sku.focus()

    # =========================================================================
    # FASE 2: ARMADO (checkmarks por pedido)
    # =========================================================================

    def _actualizar_estado_fase1(self, pn):
        """Actualiza los checkboxes verdes en FASE 1"""
        d = self.pedidos[pn]
        if "skus_items" not in d:
            return

        for sku, item in d["skus_items"].items():
            chk = item["checkbox"]
            req = d["skus_requeridos"][sku]
            esc = d["skus_escaneados"].get(sku, 0)

            if esc >= req:
                chk.config(text="✔", fg=C["success"])
            else:
                chk.config(text="○", fg=C["text_lo"])

    def _agregar_sku_pac(self, sku):
        """Procesa escaneo para PAC en curso"""
        pn = self.pedido_en_proceso
        d  = self.pedidos[pn]

        if sku in d["skus_requeridos"]:
            req = d["skus_requeridos"][sku]
            esc = d["skus_escaneados"].get(sku, 0)
            if esc < req:
                d["skus_escaneados"][sku] = esc + 1
                self._actualizar_estado_fase1(pn)
                self._evaluar_pac(pn)
            else:
                self.lbl_resultado.config(
                    text=f"✅  '{sku}' ya completo.",
                    fg=C["success"])
        else:
            self.lbl_resultado.config(
                text=f"⚠  '{sku}' no pertenece a: {self._nombre_pedido(pn)}",
                fg=C["danger"])
            self.col_control.config(bg=C["danger"])
            self.root.after(200, lambda: self.col_control.config(bg=C["panel"]))

    def _evaluar_pac(self, pn):
        """Evalúa estado del PAC"""
        d         = self.pedidos[pn]
        total_req = sum(d["skus_requeridos"].values())
        total_esc = sum(d["skus_escaneados"].values())

        self.mostrar_vista_previa(pn)

        if total_req - total_esc > 0:
            nombre = self._nombre_pedido(pn)
            font_sz = 22 if len(nombre) > 20 else 32
            self.lbl_num_caja.config(text=nombre, fg=C["accent"],
                                     font=("Segoe UI Black", font_sz, "bold"),
                                     wraplength=360)
            self.lbl_resultado.config(
                text=f"{total_esc}/{total_req} artículos",
                fg=C["text_mid"])
        else:
            self._mostrar_caja_completa(pn)
            self.root.after(80, lambda: self._imprimir_automatico(pn))

    def _mostrar_caja_completa(self, pn):
        d = self.pedidos[pn]
        nombre = self._nombre_pedido(pn)
        font_sz = 22 if len(nombre) > 20 else 32
        self.lbl_num_caja.config(text=nombre, fg=C["success"],
                                 font=("Segoe UI Black", font_sz, "bold"),
                                 wraplength=360)
        self.lbl_resultado.config(
            text="✅  ¡Completa! Imprimiendo...",
            fg=C["success"])
        self._animar_pedido_completo()
        self.mostrar_vista_previa(pn)
        self.root.update_idletasks()

    def _extraer_descripcion_pagina(self, txt):
        lineas = [l.strip() for l in txt.split('\n')]

        _EXCLUIR = (
            "Recorta", "Pack ID", "Destinatario", "Direccion",
            "Envio", "FLEX", "ZONA", "Barrio", "SKU:", "@",
            "http", "www.", "Aramburu", "GROUP", "SRL", "S.A.",
        )

        def _es_candidata(linea):
            if len(linea) < 8:
                return False
            if re.match(r'^\d', linea):
                return False
            if re.search(r'20000\d{11}', linea.replace(" ", "")):
                return False
            for exc in _EXCLUIR:
                if exc.lower() in linea.lower():
                    return False
            return True

        # Estrategia 1: buscar la línea con " | " (formato ML: "Título | N u.")
        for linea in lineas:
            if " | " in linea and _es_candidata(linea):
                return linea.split(" | ")[0].strip()

        # Estrategia 2: retroceder desde "SKU:" saltando líneas de atributo (Color:, Talle:, etc.)
        for i, linea in enumerate(lineas):
            if "SKU:" in linea:
                for j in range(i - 1, max(i - 6, -1), -1):
                    cand = lineas[j]
                    if not cand:
                        continue
                    # Saltar líneas de atributo como "Color: Negro" o "Talle: M"
                    if re.match(r'^[A-Za-záéíóúÁÉÍÓÚñÑ]+\s*:', cand):
                        continue
                    if _es_candidata(cand):
                        return cand
                break

        return ""

    def _nombre_pedido(self, pn):
        skus   = self.pedidos[pn]["skus_requeridos"]
        n_skus = len(skus)
        # Intentar nombre desde Excel usando el primer SKU del pedido
        primer_sku = next(iter(skus), None)
        desc_excel = self._nombre_sku(primer_sku) if primer_sku else ""
        desc = desc_excel or self.pedidos[pn].get("descripcion", "").strip()
        if desc and n_skus > 1:
            return f"{desc}  (+{n_skus - 1} más)"
        return desc if desc else f"CAJA {pn}"

    def mostrar_vista_previa(self, num_pagina):
        try:
            documento = fitz.open(self.ruta_pdf)
            pagina    = documento.load_page(num_pagina - 1)
            pix = pagina.get_pixmap(matrix=fitz.Matrix(3.0, 3.0))
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            documento.close()

            self.canvas_preview.update_idletasks()
            cw = self.canvas_preview.winfo_width()  or 600
            ch = self.canvas_preview.winfo_height() or 500

            img_w, img_h = img.size
            ratio = min((cw - 40) / img_w, (ch - 40) / img_h)
            new_w = int(img_w * ratio)
            new_h = int(img_h * ratio)
            img = img.resize((new_w, new_h), Image.LANCZOS)

            self.imagen_actual = ImageTk.PhotoImage(img)
            self.canvas_preview.delete("all")

            ox, oy = cw // 2, ch // 2
            self.canvas_preview.create_rectangle(
                ox - new_w//2 + 3, oy - new_h//2 + 3,
                ox + new_w//2 + 3, oy + new_h//2 + 3,
                fill="#000000", width=0)

            self.canvas_preview.create_image(ox, oy, anchor="center",
                                              image=self.imagen_actual)

            self.canvas_preview.create_rectangle(
                ox - new_w//2 - 1, oy - new_h//2 - 1,
                ox + new_w//2 + 1, oy + new_h//2 + 1,
                outline=C["border"], width=1, fill="")

            self.lbl_visor_info.config(text=f"Etiqueta {num_pagina}")

        except Exception as e:
            pass

    def _imprimir_automatico(self, p_num):
        impresora = self.config.get("impresora", "").strip()
        if not impresora:
            resp = messagebox.askokcancel(
                "Sin impresora",
                "⚠  No hay impresora configurada.\nPresioná OK para configurarla.",
                parent=self.root)
            if resp:
                self.abrir_configuracion()
            return

        try:
            self.lbl_imprimiendo.config(text="⏳  Imprimiendo...", fg=C["accent"])
            self.btn_cancelar.config(state="disabled", bg=C["card"], fg=C["text_lo"])
            self.btn_saltar.config(state="disabled", bg=C["card"], fg=C["text_lo"])
            self.root.update_idletasks()

            doc_o = fitz.open(self.ruta_pdf)
            doc_p = fitz.open()
            doc_p.insert_pdf(doc_o, from_page=p_num - 1, to_page=p_num - 1)
            tmp = os.path.join(tempfile.gettempdir(), f"etiqueta_ml_{p_num}.pdf")
            doc_p.save(tmp)
            doc_p.close()
            doc_o.close()

            ok = self._enviar_a_impresora(tmp, impresora)

            if ok:
                self.lbl_imprimiendo.config(text="✅  Impresa", fg=C["success"])
            else:
                self.lbl_imprimiendo.config(text="❌  Error al imprimir", fg=C["danger"])
                messagebox.showerror(
                    "Error de impresión",
                    f"No se pudo enviar la etiqueta a:\n{impresora}\n\n"
                    "Verificá que:\n"
                    "• La impresora esté encendida y conectada\n"
                    "• El nombre de impresora sea correcto (Configurar impresora)\n"
                    "• SumatraPDF esté instalado para mayor compatibilidad",
                    parent=self.root)
                return

            self.pedidos[p_num]["impreso"] = True
            self.pedido_en_proceso = None
            pendientes = self.actualizar_contador_global()

            # Actualizar FASE 1
            d = self.pedidos[p_num]
            if "_lbl_estado" in d:
                d["_lbl_estado"].config(text="✔ Hecha", fg=C["success"])

            self.limpiar_vista_previa()

            if pendientes == 0:
                self.lbl_num_caja.config(text="")
                self.lbl_imprimiendo.config(text="")
                messagebox.showinfo(
                    "¡Lote terminado!",
                    "🎉  Todos los pedidos procesados.",
                    parent=self.root)
                self.lbl_resultado.config(text="")
                self.btn_cargar.focus()
            else:
                self.lbl_num_caja.config(text="")
                self.lbl_resultado.config(text="✅  Siguiente pedido.", fg=C["text_mid"])
                self.entrada_sku.focus()

        except Exception as e:
            self.lbl_imprimiendo.config(text="")
            if self.pedido_en_proceso:
                self.btn_cancelar.config(state="normal", bg=C["danger"], fg="white")
                self.btn_saltar.config(state="normal", bg=C["warning"], fg="white")

    def limpiar_vista_previa(self):
        self.imagen_actual = None
        self._mostrar_placeholder_visor()
        self.lbl_visor_info.config(text="")

    def _ps_encoded(self, script, timeout=45):
        """Ejecuta PowerShell via EncodedCommand (base64) — evita todo problema de escaping."""
        import base64
        enc = base64.b64encode(script.encode("utf-16-le")).decode("ascii")
        r = subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive",
             "-ExecutionPolicy", "Bypass", "-EncodedCommand", enc],
            capture_output=True, text=True, timeout=timeout)
        return r.returncode == 0

    def _renderizar_pagina_imagen(self, ruta_pdf, dpi_scale=4.17):
        """Renderiza la primera página del PDF como imagen PIL (300 DPI aprox.)."""
        doc = fitz.open(ruta_pdf)
        pix = doc.load_page(0).get_pixmap(matrix=fitz.Matrix(dpi_scale, dpi_scale))
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        doc.close()
        return img

    def _enviar_a_impresora(self, ruta_pdf, impresora):
        imp_ps      = impresora.replace("'", "''")
        pdf_ps      = ruta_pdf.replace("'", "''")
        folder_ps   = os.path.dirname(ruta_pdf).replace("'", "''")
        filename_ps = os.path.basename(ruta_pdf).replace("'", "''")

        # ── Método 0: win32print + win32ui directo (más confiable) ──────────
        # Renderiza el PDF a imagen y lo envía directo al spooler de Windows.
        # No depende de SumatraPDF ni del visor PDF del sistema.
        try:
            import win32print
            import win32ui
            from PIL import ImageWin

            img = self._renderizar_pagina_imagen(ruta_pdf)

            hdc = win32ui.CreateDC()
            hdc.CreatePrinterDC(impresora)
            pw = hdc.GetDeviceCaps(110)   # HORZRES: ancho imprimible en píxeles
            ph = hdc.GetDeviceCaps(111)   # VERTRES: alto imprimible en píxeles

            scale = min(pw / img.width, ph / img.height)
            nw    = int(img.width  * scale)
            nh    = int(img.height * scale)
            img   = img.resize((nw, nh), Image.LANCZOS)

            hdc.StartDoc(os.path.basename(ruta_pdf))
            hdc.StartPage()
            x0 = (pw - nw) // 2
            y0 = (ph - nh) // 2
            dib = ImageWin.Dib(img)
            dib.draw(hdc.GetHandleOutput(), (x0, y0, x0 + nw, y0 + nh))
            hdc.EndPage()
            hdc.EndDoc()
            hdc.DeleteDC()
            return True
        except ImportError:
            pass
        except Exception:
            pass

        # ── Método 1: SumatraPDF ────────────────────────────────────────────
        for ruta_s in [
            r"C:\Program Files\SumatraPDF\SumatraPDF.exe",
            r"C:\Program Files (x86)\SumatraPDF\SumatraPDF.exe",
            os.path.join(os.environ.get("LOCALAPPDATA", ""), "SumatraPDF", "SumatraPDF.exe"),
        ]:
            if os.path.exists(ruta_s):
                try:
                    r = subprocess.run(
                        [ruta_s, "-print-to", impresora,
                         "-print-settings", "fit", "-silent", ruta_pdf],
                        capture_output=True, timeout=30)
                    return r.returncode == 0
                except Exception:
                    pass

        # ── Método 2: PowerShell System.Drawing.Printing ────────────────────
        # Convierte el PDF a PNG y lo imprime vía .NET sin depender del visor PDF.
        tmp_png = ""
        try:
            img     = self._renderizar_pagina_imagen(ruta_pdf)
            tmp_png = ruta_pdf.replace(".pdf", "_print.png")
            img.save(tmp_png)
            png_ps  = tmp_png.replace("'", "''")

            script2 = f"""
Add-Type -AssemblyName System.Drawing
$script:imgPath  = '{png_ps}'
$script:printer  = '{imp_ps}'
$script:pd = New-Object System.Drawing.Printing.PrintDocument
$script:pd.PrinterSettings.PrinterName = $script:printer
$script:pd.add_PrintPage({{
    param($s, $e)
    $b = New-Object System.Drawing.Bitmap $script:imgPath
    try {{
        $r = $e.MarginBounds
        $iw = $b.Width; $ih = $b.Height
        $scale = [Math]::Min($r.Width / $iw, $r.Height / $ih)
        $nw = [int]($iw * $scale); $nh = [int]($ih * $scale)
        $x = $r.X + ($r.Width  - $nw) / 2
        $y = $r.Y + ($r.Height - $nh) / 2
        $e.Graphics.DrawImage($b, [int]$x, [int]$y, $nw, $nh)
    }} finally {{ $b.Dispose() }}
    $e.HasMorePages = $false
}})
try {{
    $script:pd.Print()
    Start-Sleep -Milliseconds 800
    exit 0
}} catch {{
    exit 1
}} finally {{
    $script:pd.Dispose()
}}
"""
            if self._ps_encoded(script2, timeout=30):
                return True
        except Exception:
            pass
        finally:
            try:
                if tmp_png and os.path.exists(tmp_png):
                    os.remove(tmp_png)
            except Exception:
                pass

        # ── Método 3: win32api.ShellExecute ─────────────────────────────────
        try:
            import win32api
            win32api.ShellExecute(0, "printto", ruta_pdf, f'"{impresora}"', ".", 0)
            return True
        except ImportError:
            pass
        except Exception:
            pass

        # ── Método 4: Start-Process -Verb PrintTo ───────────────────────────
        script4 = f"""
try {{
    $proc = Start-Process -FilePath '{pdf_ps}' -Verb PrintTo `
            -ArgumentList ('"' + '{imp_ps}' + '"') -PassThru -ErrorAction Stop
    if ($proc) {{ $proc.WaitForExit(15000) | Out-Null }}
    exit 0
}} catch {{
    exit 1
}}
"""
        if self._ps_encoded(script4):
            return True

        # ── Método 5: Shell.Application InvokeVerbEx ────────────────────────
        script5 = f"""
try {{
    $sh  = New-Object -ComObject Shell.Application
    $dir = $sh.NameSpace('{folder_ps}')
    $fil = $dir.ParseName('{filename_ps}')
    $fil.InvokeVerbEx('printto', ('"' + '{imp_ps}' + '"'))
    Start-Sleep -Seconds 2
    exit 0
}} catch {{
    exit 1
}}
"""
        if self._ps_encoded(script5):
            return True

        return False


    # =========================================================================
    # SERVIDOR MÓVIL
    # =========================================================================

    def _exportar_debounced(self):
        """Debounce: exporta al JSON como máximo una vez cada 600ms."""
        if not self._export_pending:
            self._export_pending = True
            self.root.after(600, self._exportar_flush)

    def _exportar_flush(self):
        self._export_pending = False
        if self._servidor_activo:
            self._exportar_estado_movil()
        # Subir también al servidor en la nube si está configurado
        url_nube = self.config.get("servidor_nube", "").strip()
        if url_nube:
            self._subir_a_nube_async()

    def _exportar_estado_movil(self):
        """Escribe estado_picking.json. Solo se llama si el servidor está activo."""
        if not self._servidor_activo:
            return
        try:
            base_dir    = os.path.dirname(os.path.abspath(__file__))
            estado_path = os.path.join(base_dir, "estado_picking.json")

            total_req = {}
            for d in self.pedidos.values():
                for sku, qty in d["skus_requeridos"].items():
                    total_req[sku] = total_req.get(sku, 0) + qty

            grupos_dict = {}
            SIN_PASILLO = "Sin ubicación"
            for sku in sorted(total_req.keys()):
                pasillo, estanteria = self._ubicacion_sku(sku)
                key = pasillo if pasillo else SIN_PASILLO
                grupos_dict.setdefault(key, []).append({
                    "sku":        sku,
                    "nombre":     self._nombre_sku(sku),
                    "req":        total_req[sku],
                    "pasillo":    pasillo,
                    "estanteria": estanteria,
                })

            def _orden(nombre):
                m = re.search(r'(\d+)', nombre)
                return (0, int(m.group(1)), nombre) if m else (1, 0, nombre)

            grupos = [
                {"pasillo": k, "items": v}
                for k, v in sorted(grupos_dict.items(), key=lambda x: _orden(x[0]))
            ]
            estado = {
                "fase":             self.fase_actual,
                "grupos":           grupos,
                "colecta":          dict(self.colecta_global),
                "colecta_completa": False,
                "total_skus":       len(total_req),
                "total_uds":        sum(total_req.values()),
            }
            # Escritura atómica: escribir en temp y renombrar
            tmp = estado_path + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(estado, f, ensure_ascii=False)
            os.replace(tmp, estado_path)
        except Exception:
            pass

    def toggle_servidor_movil(self):
        if self._servidor_activo:
            self._detener_servidor()
        else:
            self._iniciar_servidor()

    def _iniciar_servidor(self):
        import socket, threading

        if not self.pedidos:
            messagebox.showwarning(
                "Sin datos", "Cargá un PDF primero antes de iniciar el servidor.",
                parent=self.root)
            return

        self._exportar_estado_movil()

        # Detectar IP local
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
        except Exception:
            ip = "127.0.0.1"

        # Generar certificado SSL
        base_dir  = os.path.dirname(os.path.abspath(__file__))
        cert_path = os.path.join(base_dir, "ssl_cert.pem")
        key_path  = os.path.join(base_dir, "ssl_key.pem")
        ssl_ok    = self._generar_certificado_ssl(ip, cert_path, key_path)

        # Construir el HTML de la app móvil embebido
        html_movil = self._html_app_movil()

        # Crear app Flask inline
        try:
            from flask import Flask as _Flask, jsonify as _jsonify, request as _req
        except ImportError:
            messagebox.showerror(
                "Módulo faltante",
                "Instalá Flask para usar el servidor móvil:\n\npip install flask",
                parent=self.root)
            return

        import json as _json

        flask_app   = _Flask(__name__)
        estado_path = os.path.join(base_dir, "estado_picking.json")
        _lock2      = threading.Lock()

        def _leer():
            if not os.path.exists(estado_path):
                return {"grupos": [], "colecta": {}}
            with _lock2:
                with open(estado_path, "r", encoding="utf-8") as f:
                    return _json.load(f)

        def _guardar(e):
            with _lock2:
                with open(estado_path, "w", encoding="utf-8") as f:
                    _json.dump(e, f, ensure_ascii=False)

        @flask_app.route("/")
        def _index():
            return html_movil, 200, {"Content-Type": "text/html; charset=utf-8"}

        @flask_app.route("/manifest.json")
        def _manifest():
            m = _json.dumps({"name":"Sistema de Picking","short_name":"Picking",
                             "start_url":"/","display":"standalone",
                             "background_color":"#0F172A","theme_color":"#1E293B"})
            return m, 200, {"Content-Type": "application/json"}

        @flask_app.route("/api/estado")
        def _estado():
            return _jsonify(_leer())

        @flask_app.route("/api/escanear", methods=["POST"])
        def _escanear():
            data    = _req.get_json(force=True)
            sku     = str(data.get("sku","")).strip().upper()
            if not sku:
                return _jsonify({"ok": False, "msg": "SKU vacío"})
            estado  = _leer()
            colecta = estado.get("colecta", {})
            sku_info = None
            for g in estado.get("grupos", []):
                for it in g.get("items", []):
                    if it["sku"] == sku:
                        sku_info = it; break
                if sku_info: break
            if not sku_info:
                return _jsonify({"ok": False, "tipo": "no_encontrado",
                                 "msg": f"'{sku}' no está en ningún pedido"})
            req = sku_info["req"]
            col = colecta.get(sku, 0)
            if col >= req:
                return _jsonify({"ok": True, "tipo": "ya_completo",
                                 "msg": f"'{sku}' ya completo ({req}/{req})",
                                 "collected": col, "req": req})
            colecta[sku] = col + 1
            estado["colecta"] = colecta
            todo = all(colecta.get(it["sku"],0) >= it["req"]
                       for g in estado["grupos"] for it in g["items"])
            estado["colecta_completa"] = todo
            _guardar(estado)
            # Sincronizar colecta de vuelta a la app de escritorio
            self.root.after(0, lambda: self._sincronizar_colecta_movil(colecta))
            nuevo = colecta[sku]
            return _jsonify({"ok": True,
                             "tipo": "completo" if nuevo >= req else "parcial",
                             "sku": sku, "nombre": sku_info.get("nombre",""),
                             "pasillo": sku_info.get("pasillo",""),
                             "estanteria": sku_info.get("estanteria",""),
                             "collected": nuevo, "req": req,
                             "todo_completo": todo,
                             "msg": f"✔ {nuevo}/{req}" if nuevo >= req else f"{nuevo}/{req}"})

        @flask_app.route("/api/reset_sku", methods=["POST"])
        def _reset():
            data    = _req.get_json(force=True)
            sku     = str(data.get("sku","")).strip().upper()
            estado  = _leer()
            colecta = estado.get("colecta", {})
            if sku in colecta and colecta[sku] > 0:
                colecta[sku] -= 1
                if colecta[sku] == 0: del colecta[sku]
                estado["colecta"] = colecta
                estado["colecta_completa"] = False
                _guardar(estado)
                self.root.after(0, lambda: self._sincronizar_colecta_movil(colecta))
                return _jsonify({"ok": True, "msg": f"Deshecho: {sku}"})
            return _jsonify({"ok": False, "msg": "Nada que deshacer"})

        import logging
        log = logging.getLogger("werkzeug")
        log.setLevel(logging.ERROR)

        def _run():
            try:
                if ssl_ok:
                    flask_app.run(host="0.0.0.0", port=5050, debug=False,
                                  use_reloader=False,
                                  ssl_context=(cert_path, key_path))
                else:
                    flask_app.run(host="0.0.0.0", port=5050, debug=False,
                                  use_reloader=False)
            except Exception:
                pass

        self._flask_app    = flask_app
        self._servidor_thread = threading.Thread(target=_run, daemon=True)
        self._servidor_thread.start()
        self._servidor_activo = True

        self.btn_servidor.config(text="📱  Detener servidor", bg=C["danger"], fg="white")
        proto = "https" if ssl_ok else "http"
        self.lbl_servidor_ip.config(text=f"{proto}://{ip}:5050", fg=C["success"])

        aviso_ssl = (
            "\n⚠  Chrome mostrará 'Conexión no segura' → normal.\n"
            "   Tocá 'Configuración avanzada' → 'Continuar'\n"
            "   (solo la primera vez)\n") if ssl_ok else (
            "\n⚠  Sin HTTPS: la cámara no funcionará.\n"
            "   Instalá pyopenssl:  pip install pyopenssl\n")

        messagebox.showinfo(
            "Servidor iniciado",
            f"✅  Servidor activo en:\n\n"
            f"   {proto}://{ip}:5050\n"
            f"{aviso_ssl}\n"
            f"En el celular (Chrome):\n"
            f"1. Conectate al mismo WiFi que esta PC\n"
            f"2. Entrá a la URL de arriba\n"
            f"3. Tocá ⋮ → 'Agregar a pantalla de inicio'\n\n"
            f"Sin WiFi: activá Hotspot en esta PC\ny conectá el celular a ese hotspot.",
            parent=self.root)

    def _generar_certificado_ssl(self, ip, cert_path, key_path):
        if os.path.exists(cert_path) and os.path.exists(key_path):
            return True
        try:
            from OpenSSL import crypto
            k = crypto.PKey()
            k.generate_key(crypto.TYPE_RSA, 2048)
            cert = crypto.X509()
            cert.get_subject().C  = "AR"
            cert.get_subject().O  = "Picking"
            cert.get_subject().CN = ip
            cert.set_serial_number(1)
            cert.gmtime_adj_notBefore(0)
            cert.gmtime_adj_notAfter(365*24*60*60)
            cert.set_issuer(cert.get_subject())
            cert.set_pubkey(k)
            san = f"IP:{ip},IP:127.0.0.1,DNS:localhost"
            cert.add_extensions([
                crypto.X509Extension(b"subjectAltName", False, san.encode()),
                crypto.X509Extension(b"basicConstraints", True, b"CA:TRUE"),
            ])
            cert.sign(k, "sha256")
            with open(cert_path, "wb") as f:
                f.write(crypto.dump_certificate(crypto.FILETYPE_PEM, cert))
            with open(key_path, "wb") as f:
                f.write(crypto.dump_privatekey(crypto.FILETYPE_PEM, k))
            return True
        except Exception:
            return False

    def _sincronizar_colecta_movil(self, colecta_movil):
        """Recibe escaneos del celular y actualiza checkmarks en tiempo real."""
        cambios = False
        nuevos  = []   # SKUs que cambiaron en esta sincronización

        for sku, qty in colecta_movil.items():
            qty = int(qty)
            ant = self.colecta_global.get(sku, 0)
            if ant != qty:
                self.colecta_global[sku] = qty
                self._actualizar_checkmarks_fase1(sku)
                if qty > ant:   # Nuevo escaneo (no retroceso)
                    nuevos.append(sku)
                cambios = True

        if cambios:
            # Actualizar totales globales
            self.actualizar_contador_global()
            self.root.update_idletasks()

            # Parpadeo verde en el panel izquierdo por cada SKU nuevo escaneado
            for sku in nuevos[:3]:   # máximo 3 simultáneos
                self._flash_sku_escaneado(sku)

            # Si toda la colecta está completa, notificar
            total_req = {}
            for d in self.pedidos.values():
                for s, q in d["skus_requeridos"].items():
                    total_req[s] = total_req.get(s, 0) + q
            if all(self.colecta_global.get(s,0) >= q for s,q in total_req.items()):
                self.lbl_resultado.config(
                    text="✅ ¡Colecta completa desde celular! Pasá a Fase 2.",
                    fg=C["success"])
                self.col_control.config(bg=C["success"])
                self.root.after(1000, lambda: self.col_control.config(bg=C["panel"]))

    def _flash_sku_escaneado(self, sku):
        """Parpadeo verde en el item de la lista cuando llega un escaneo del celular."""
        if not hasattr(self, "fase1_items") or sku not in self.fase1_items:
            return
        item = self.fase1_items[sku]
        chk  = item.get("checkbox")
        cnt  = item.get("lbl_cnt")
        if not chk or not chk.winfo_exists():
            return
        # Flash: cambiar color brevemente
        orig_bg_chk = chk.cget("fg")
        if chk.winfo_exists():
            chk.config(fg="#FFFFFF")
        def _restaurar():
            if chk.winfo_exists():
                chk.config(fg=orig_bg_chk)
        self.root.after(400, _restaurar)

    def _construir_payload_nube(self):
        """Construye el dict para subir al servidor en la nube."""
        total_req = {}
        for d in self.pedidos.values():
            for sku, qty in d["skus_requeridos"].items():
                total_req[sku] = total_req.get(sku, 0) + qty

        grupos_dict = {}
        for sku in sorted(total_req.keys()):
            pasillo, estanteria = self._ubicacion_sku(sku)
            key = pasillo if pasillo else "Sin ubicación"
            grupos_dict.setdefault(key, []).append({
                "sku": sku, "nombre": self._nombre_sku(sku),
                "req": total_req[sku], "pasillo": pasillo, "estanteria": estanteria,
            })

        def _orden(n):
            m = re.search(r'(\d+)', n)
            return (0, int(m.group(1)), n) if m else (1, 0, n)

        grupos = [{"pasillo": k, "items": v}
                  for k, v in sorted(grupos_dict.items(), key=lambda x: _orden(x[0]))]

        return {
            "fase": self.fase_actual, "grupos": grupos,
            "colecta": dict(self.colecta_global),
            "colecta_completa": False,
            "total_skus": len(total_req),
            "total_uds": sum(total_req.values()),
        }

    def _subir_a_nube_async(self):
        """Sube el estado al servidor en la nube en un thread para no bloquear la UI."""
        import threading
        threading.Thread(target=self._subir_a_nube_worker, daemon=True).start()

    def _subir_a_nube_worker(self):
        """Sube el estado al servidor Railway. Siempre usa RAILWAY_URL y clave fija."""
        try:
            import urllib.request, json as _json
            url     = RAILWAY_URL.rstrip("/")
            key     = "everest2024"   # siempre fija — igual a PICKING_API_KEY en Railway
            payload = _json.dumps(self._construir_payload_nube()).encode("utf-8")
            req = urllib.request.Request(
                url + "/api/subir_estado",
                data=payload,
                headers={"Content-Type": "application/json", "X-API-Key": key},
                method="POST")
            with urllib.request.urlopen(req, timeout=10) as resp:
                resultado = _json.loads(resp.read())
                if resultado.get("ok"):
                    self.root.after(0, lambda: self._on_subida_ok(resultado))
                else:
                    self.root.after(0, lambda: self._on_subida_error(str(resultado)))
        except Exception as e:
            self.root.after(0, lambda: self._on_subida_error(str(e)))

    def _mostrar_info_movil(self):
        """Muestra la ventana con QR y URL para que los celulares entren a Railway."""
        win = tk.Toplevel(self.root)
        win.title("App Móvil — Picking")
        win.geometry("480x420")
        win.resizable(False, False)
        win.config(bg=C["bg_dark"])
        win.grab_set()

        # Header
        hdr = tk.Frame(win, bg=C["success"], pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text="📱  App Móvil de Picking",
                 font=("Segoe UI Black", 13), bg=C["success"], fg="white").pack()
        tk.Label(hdr, text="Los operarios abren esta URL en sus celulares",
                 font=("Segoe UI", 9), bg=C["success"], fg="#D1FAE5").pack(pady=(2,0))

        body = tk.Frame(win, bg=C["bg_dark"], padx=28, pady=20)
        body.pack(fill="both", expand=True)

        url_movil = RAILWAY_URL.rstrip("/") + "/movil"

        # URL grande y copiable
        tk.Label(body, text="URL para los celulares:",
                 font=("Segoe UI", 9, "bold"), bg=C["bg_dark"],
                 fg=C["text_mid"]).pack(anchor="w")

        url_frame = tk.Frame(body, bg=C["accent"], padx=2, pady=2)
        url_frame.pack(fill="x", pady=(6,0))
        url_in = tk.Frame(url_frame, bg=C["card"])
        url_in.pack(fill="x")
        lbl_url = tk.Label(url_in, text=url_movil,
                           font=("Consolas", 11, "bold"),
                           bg=C["card"], fg=C["success"],
                           anchor="w", padx=10, pady=8)
        lbl_url.pack(fill="x")

        # Botón copiar
        def _copiar():
            win.clipboard_clear()
            win.clipboard_append(url_movil)
            btn_copiar.config(text="✅ Copiado!", bg=C["success"])
            win.after(2000, lambda: btn_copiar.config(text="📋 Copiar URL", bg=C["accent"]))

        btn_copiar = tk.Button(body, text="📋 Copiar URL",
                               font=("Segoe UI Semibold", 10),
                               bg=C["accent"], fg="white",
                               activebackground=C["accent2"], activeforeground="white",
                               relief="flat", cursor="hand2", padx=14, pady=7, bd=0,
                               command=_copiar)
        btn_copiar.pack(fill="x", pady=(8,0))

        # Estado del lote
        tk.Frame(body, bg=C["border"], height=1).pack(fill="x", pady=(16,12))

        self.lbl_railway_status = tk.Label(
            body, text="🔄  Verificando estado en Railway...",
            font=("Segoe UI", 9), bg=C["bg_dark"], fg=C["text_mid"],
            wraplength=400, justify="left")
        self.lbl_railway_status.pack(anchor="w")

        # Botón subir/actualizar lote
        def _subir_ahora():
            if not self.pedidos:
                messagebox.showwarning(
                    "Sin lote", "Cargá un PDF o generá un lote ML primero.",
                    parent=win)
                return
            btn_subir.config(text="⏳ Subiendo...", state="disabled")
            self._subir_a_nube_async_con_callback(
                lambda ok, msg: _on_resultado(ok, msg))

        def _on_resultado(ok, msg):
            if ok:
                btn_subir.config(
                    text="✅ Lote enviado a Railway",
                    bg=C["success"], state="normal")
                self.lbl_railway_status.config(
                    text=f"✅ Lote activo en Railway — {len(self.pedidos)} pedidos cargados.\n"
                         f"Los celulares verán la lista en segundos.",
                    fg=C["success"])
            else:
                btn_subir.config(text="❌ Error — Reintentar", bg=C["danger"], state="normal")
                self.lbl_railway_status.config(
                    text=f"❌ Error al conectar con Railway:\n{msg}",
                    fg=C["danger"])

        btn_subir = tk.Button(body, text="🚀  Enviar lote a Railway (celulares)",
                              font=("Segoe UI Semibold", 10),
                              bg="#7C3AED", fg="white",
                              activebackground=C["accent2"], activeforeground="white",
                              relief="flat", cursor="hand2", padx=14, pady=7, bd=0,
                              command=_subir_ahora)
        btn_subir.pack(fill="x", pady=(8,0))

        # Instrucciones
        tk.Label(body,
                 text="Instrucciones para el operario:\n"
                      "1. Abrir Chrome en el celular\n"
                      "2. Escribir la URL de arriba\n"
                      "3. Escanear con el lector integrado del teléfono",
                 font=("Segoe UI", 8), bg=C["bg_dark"], fg=C["text_lo"],
                 justify="left").pack(anchor="w", pady=(12,0))

        # Verificar estado Railway al abrir
        self._verificar_estado_railway_en_ventana()

    def _verificar_estado_railway_en_ventana(self):
        """Consulta Railway y actualiza lbl_railway_status si existe."""
        import threading, urllib.request, json as _j
        def _w():
            try:
                url = RAILWAY_URL.rstrip("/") + "/api/estado"
                with urllib.request.urlopen(url, timeout=5) as r:
                    d = _j.loads(r.read())
                cargado   = d.get("cargado", False)
                total_skus = d.get("total_skus", 0)
                total_uds  = d.get("total_uds", 0)
                ts         = d.get("ultima_actualizacion", "")
                if cargado:
                    txt = (f"✅ Railway tiene un lote activo:\n"
                           f"   {total_skus} SKUs · {total_uds} unidades · actualizado {ts}\n"
                           f"   Los celulares ya pueden ver la lista.")
                    col = C["success"]
                else:
                    txt = ("⚠  Railway NO tiene lote activo todavía.\n"
                           "   Presioná '🚀 Enviar lote a Railway' para que los celulares vean la lista.")
                    col = C["warning"]
                def _update():
                    if hasattr(self, 'lbl_railway_status') and \
                       self.lbl_railway_status.winfo_exists():
                        self.lbl_railway_status.config(text=txt, fg=col)
                self.root.after(0, _update)
            except Exception as e:
                def _err():
                    if hasattr(self, 'lbl_railway_status') and \
                       self.lbl_railway_status.winfo_exists():
                        self.lbl_railway_status.config(
                            text=f"❌ No se pudo conectar a Railway:\n{e}",
                            fg=C["danger"])
                self.root.after(0, _err)
        threading.Thread(target=_w, daemon=True).start()

    def _subir_a_nube_async_con_callback(self, callback):
        """Sube a Railway y llama callback(ok, msg) al terminar."""
        import threading
        def _w():
            try:
                import urllib.request, json as _json
                url     = RAILWAY_URL.rstrip("/")
                key     = "everest2024"   # clave fija
                payload = _json.dumps(self._construir_payload_nube()).encode("utf-8")
                req = urllib.request.Request(
                    url + "/api/subir_estado",
                    data=payload,
                    headers={"Content-Type": "application/json", "X-API-Key": key},
                    method="POST")
                with urllib.request.urlopen(req, timeout=12) as resp:
                    resultado = _json.loads(resp.read())
                ok  = resultado.get("ok", False)
                msg = resultado.get("msg", "")
                self.root.after(0, lambda: callback(ok, msg))
                if ok:
                    self.root.after(0, lambda: self._on_subida_ok(resultado))
            except Exception as e:
                self.root.after(0, lambda: callback(False, str(e)))
        threading.Thread(target=_w, daemon=True).start()

    def _on_subida_ok(self, resultado):
        msg = resultado.get("msg", "")
        total = len(self.pedidos)
        self.lbl_estado_pdf.config(
            text=f"✅ {total} pedidos en Railway — celulares listos",
            fg=C["success"])
        self.root.after(5000, self._sincronizar_desde_nube)

    def _on_subida_error(self, err):
        # Silencioso — el error se muestra solo si el usuario abre la ventana
        self.lbl_estado_pdf.config(
            text=f"⚠  Lote cargado localmente (Railway sin conexión)",
            fg=C["warning"])

    def _sincronizar_desde_nube(self):
        """Polling: lee la colecta del servidor Railway y sincroniza checkmarks."""
        import threading
        def _worker():
            try:
                import urllib.request, json as _json
                url = RAILWAY_URL.rstrip("/")
                with urllib.request.urlopen(url + "/api/estado", timeout=5) as resp:
                    data = _json.loads(resp.read())
                colecta = data.get("colecta", {})
                if colecta:
                    self.root.after(0, lambda: self._sincronizar_colecta_movil(colecta))
            except Exception:
                pass
        threading.Thread(target=_worker, daemon=True).start()
        # Re-programar cada 5 segundos
        self.root.after(5000, self._sincronizar_desde_nube)

    def _html_app_movil(self):
        """HTML de la app móvil embebido. Intenta leer el archivo externo primero
        (desarrollo), si no existe usa el string embebido (producción/exe)."""
        # Intentar archivo externo (cuando se corre como .py con los archivos al lado)
        try:
            ruta = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                "movil_static", "index.html")
            if os.path.exists(ruta):
                with open(ruta, "r", encoding="utf-8") as f:
                    return f.read()
        except Exception:
            pass
        # Fallback: HTML embebido directamente
        return self._HTML_MOVIL_EMBEBIDO

    # ── HTML de la app móvil embebido (no depende de archivos externos) ───────
    _HTML_MOVIL_EMBEBIDO = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
<meta name="theme-color" content="#1E293B">
<meta name="mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-capable" content="yes">
<title>Sistema de Picking</title>
<style>
:root{--bg:#0F172A;--panel:#1E293B;--card:#162032;--border:#334155;--accent:#3B82F6;--accent2:#6366F1;--success:#10B981;--warning:#F59E0B;--danger:#EF4444;--hi:#F1F5F9;--mid:#94A3B8;--lo:#475569;--bar:#1E3A5F}
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
body{background:var(--bg);color:var(--hi);font-family:'Segoe UI',system-ui,sans-serif;min-height:100vh}
.topbar{background:var(--panel);padding:12px 16px 10px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100;border-bottom:1px solid var(--border)}
.topbar-left{display:flex;align-items:center;gap:10px}
.topbar-title{font-size:14px;font-weight:700;color:var(--hi)}
.topbar-sub{font-size:11px;color:var(--mid)}
.badge{background:var(--accent);color:white;font-size:12px;font-weight:700;padding:4px 10px;border-radius:20px;white-space:nowrap}
.badge.done{background:var(--success)}
.scanner-section{padding:12px 14px;background:var(--panel);border-bottom:1px solid var(--border)}
.scanner-label{font-size:10px;font-weight:700;color:var(--lo);letter-spacing:.08em;margin-bottom:6px}
.input-wrap{display:flex;gap:8px;align-items:center}
#sku-input{flex:1;background:var(--card);border:2px solid var(--accent);color:var(--hi);font-size:16px;font-family:'Consolas',monospace;font-weight:700;padding:11px 14px;border-radius:10px;outline:none;text-align:center;text-transform:uppercase}
.btn-cam{background:var(--accent);border:none;border-radius:10px;color:white;font-size:22px;width:48px;height:48px;display:flex;align-items:center;justify-content:center;cursor:pointer;flex-shrink:0}
.btn-cam.active{background:var(--danger)}
#feedback{margin-top:8px;min-height:36px;border-radius:8px;padding:8px 12px;font-size:13px;font-weight:600;display:flex;align-items:center;gap:8px;transition:background .2s,color .2s}
#feedback.ok{background:rgba(16,185,129,.15);color:var(--success)}
#feedback.warn{background:rgba(245,158,11,.15);color:var(--warning)}
#feedback.error{background:rgba(239,68,68,.15);color:var(--danger)}
#feedback.neutral{background:transparent;color:var(--mid)}
#cam-container{display:none;background:#000;position:relative}
#cam-container.open{display:block}
#cam-video{width:100%;max-height:240px;object-fit:cover;display:block}
.cam-overlay{position:absolute;inset:0;display:flex;align-items:center;justify-content:center;pointer-events:none}
.cam-frame{width:65%;height:60px;border:2px solid var(--accent);border-radius:6px;box-shadow:0 0 0 2000px rgba(0,0,0,.45)}
.cam-hint{position:absolute;bottom:8px;width:100%;text-align:center;font-size:11px;color:rgba(255,255,255,.7)}
.content{padding:10px 10px 80px}
.grupo{margin-bottom:10px;border-radius:10px;overflow:hidden;border:1px solid var(--border)}
.grupo-header{background:var(--bar);padding:10px 14px;display:flex;align-items:center;justify-content:space-between;cursor:pointer;user-select:none}
.grupo-header-left{display:flex;align-items:center;gap:8px}
.grupo-nombre{font-size:13px;font-weight:800;color:var(--accent);text-transform:uppercase;letter-spacing:.04em}
.grupo-stats{font-size:11px;color:var(--mid)}
.grupo-prog{font-size:13px;font-weight:700}
.grupo-prog.done{color:var(--success)}
.grupo-prog.pend{color:var(--accent)}
.grupo-items{background:var(--card)}
.grupo-items.collapsed{display:none}
.item{display:flex;align-items:center;gap:10px;padding:10px 14px;border-bottom:1px solid var(--border)}
.item:last-child{border-bottom:none}
.item.done{opacity:.55}
.item-check{font-size:20px;flex-shrink:0;width:26px;text-align:center}
.item-check.ok{color:var(--success)}
.item-check.pend{color:var(--lo)}
.item-body{flex:1;min-width:0}
.item-nombre{font-size:13px;font-weight:600;color:var(--hi);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.item-sku-row{display:flex;align-items:center;gap:8px;margin-top:2px}
.item-sku{font-family:'Consolas',monospace;font-size:11px;color:var(--mid)}
.item-cnt{font-size:12px;font-weight:700}
.item-cnt.ok{color:var(--success)}
.item-cnt.pend{color:var(--accent)}
.item-estante{font-size:10px;color:var(--accent2);margin-top:2px}
#toast{position:fixed;bottom:24px;left:50%;transform:translateX(-50%) translateY(80px);background:var(--panel);border:1px solid var(--border);color:var(--hi);padding:12px 22px;border-radius:30px;font-size:14px;font-weight:600;z-index:999;transition:transform .3s ease,opacity .3s;opacity:0;white-space:nowrap}
#toast.show{transform:translateX(-50%) translateY(0);opacity:1}
#toast.success{border-color:var(--success);color:var(--success)}
#toast.error{border-color:var(--danger);color:var(--danger)}
#flash{position:fixed;inset:0;pointer-events:none;opacity:0;transition:opacity .15s;z-index:200}
#flash.ok-flash{background:rgba(16,185,129,.22)}
#flash.error-flash{background:rgba(239,68,68,.22)}
.empty{text-align:center;padding:60px 20px;color:var(--lo)}
.empty-icon{font-size:52px;margin-bottom:12px}
.fab{position:fixed;bottom:20px;right:20px;background:var(--accent);color:white;border:none;border-radius:50%;width:52px;height:52px;font-size:22px;cursor:pointer;box-shadow:0 4px 16px rgba(0,0,0,.4);display:flex;align-items:center;justify-content:center;z-index:50}
.fab:active{transform:scale(.92)}
.spinning{animation:spin .7s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}
#completado-banner{display:none;background:var(--success);color:white;text-align:center;padding:14px;font-size:15px;font-weight:800;letter-spacing:.05em}
#completado-banner.show{display:block}
</style>
</head>
<body>
<div id="flash"></div>
<div class="topbar">
  <div class="topbar-left">
    <span style="font-size:22px">⬡</span>
    <div><div class="topbar-title">PICKING · FASE 1</div><div class="topbar-sub">Colecta en depósito</div></div>
  </div>
  <div id="badge-global" class="badge">— / —</div>
</div>
<div id="completado-banner">🎉 ¡COLECTA COMPLETA! Volvé a la PC para pasar a Fase 2.</div>
<div class="scanner-section">
  <div class="scanner-label">ESCANEAR CÓDIGO</div>
  <div class="input-wrap">
    <input id="sku-input" type="text" inputmode="text" autocomplete="off"
           autocorrect="off" autocapitalize="characters" spellcheck="false"
           placeholder="Escaneá o escribí el SKU">
    <button class="btn-cam" id="btn-cam">📷</button>
  </div>
  <div id="feedback" class="neutral">Listo para escanear</div>
</div>
<div id="cam-container">
  <video id="cam-video" autoplay playsinline muted></video>
  <div class="cam-overlay"><div class="cam-frame"></div></div>
  <div class="cam-hint">Apuntá el código de barras al recuadro</div>
</div>
<div class="content" id="content">
  <div class="empty"><div class="empty-icon">📦</div><div class="empty-text">Cargando lista…</div></div>
</div>
<button class="fab" id="btn-refresh">🔄</button>
<div id="toast"></div>
<script>
let estado=null,scannerActivo=false,streamCam=null,barcodeDetector=null,scanLoop=null,ultimoEscaneado=null;
document.addEventListener('DOMContentLoaded',()=>{
  cargarEstado();
  const input=document.getElementById('sku-input');
  input.addEventListener('keydown',e=>{
    if(e.key==='Enter'){e.preventDefault();const v=input.value.trim();if(v){procesarSKU(v);input.value='';}}
  });
  input.addEventListener('input',()=>{
    clearTimeout(window._scanTimeout);
    window._scanTimeout=setTimeout(()=>{const v2=input.value.trim();if(v2.length>=4){procesarSKU(v2);input.value='';}},400);
  });
  document.getElementById('btn-refresh').addEventListener('click',()=>{
    document.getElementById('btn-refresh').classList.add('spinning');
    cargarEstado().finally(()=>setTimeout(()=>document.getElementById('btn-refresh').classList.remove('spinning'),500));
  });
  document.getElementById('btn-cam').addEventListener('click',toggleCamara);
  document.addEventListener('click',e=>{if(!e.target.closest('#cam-container')&&!e.target.closest('#btn-cam'))input.focus();});
  input.focus();
  setInterval(cargarEstadoSilencioso,8000);
});
async function cargarEstado(){try{const r=await fetch('/api/estado');estado=await r.json();renderLista();}catch(e){mostrarFeedback('error','❌ Sin conexión con la PC. Verificá el WiFi.');}}
async function cargarEstadoSilencioso(){try{const r=await fetch('/api/estado');estado=await r.json();renderLista(true);}catch(e){}}
function renderLista(silent=false){
  if(!estado)return;
  const content=document.getElementById('content');
  const grupos=estado.grupos||[];
  const colecta=estado.colecta||{};
  if(!grupos.length){content.innerHTML='<div class="empty"><div class="empty-icon">📋</div><div class="empty-text">Cargá un PDF en la PC primero.</div></div>';document.getElementById('badge-global').textContent='— / —';return;}
  let totalSKUs=0,donesGlobal=0;
  grupos.forEach(g=>g.items.forEach(it=>{totalSKUs++;if((colecta[it.sku]||0)>=it.req)donesGlobal++;}));
  const badge=document.getElementById('badge-global');
  badge.textContent=`${donesGlobal} / ${totalSKUs}`;
  badge.className=donesGlobal===totalSKUs?'badge done':'badge';
  document.getElementById('completado-banner').className=estado.colecta_completa?'show':'';
  const colapsados=new Set();
  document.querySelectorAll('.grupo').forEach(el=>{if(el.querySelector('.grupo-items.collapsed'))colapsados.add(el.dataset.pasillo);});
  content.innerHTML=grupos.map(grupo=>{
    const pasillo=grupo.pasillo,items=grupo.items;
    const dones=items.filter(it=>(colecta[it.sku]||0)>=it.req).length;
    const totalUds=items.reduce((s,it)=>s+it.req,0);
    const grupoDone=dones===items.length;
    const collapsed=colapsados.has(pasillo)?'collapsed':'';
    const itemsHTML=items.map(it=>{
      const col=colecta[it.sku]||0,done=col>=it.req;
      return `<div class="item ${done?'done':''}" data-sku="${it.sku}">
        <div class="item-check ${done?'ok':'pend'}">${done?'✔':'○'}</div>
        <div class="item-body">
          <div class="item-nombre">${it.nombre||it.sku}</div>
          <div class="item-sku-row"><span class="item-sku">${it.sku}</span><span class="item-cnt ${done?'ok':'pend'}">${col} / ${it.req}</span></div>
          ${it.estanteria?`<div class="item-estante">🗂 ${it.estanteria}</div>`:''}
        </div></div>`;
    }).join('');
    return `<div class="grupo" data-pasillo="${pasillo}">
      <div class="grupo-header" onclick="toggleGrupo(this)">
        <div class="grupo-header-left"><span style="font-size:16px">📦</span>
          <div><div class="grupo-nombre">${pasillo}</div><div class="grupo-stats">${items.length} SKU${items.length>1?'s':''} · ${totalUds} ud.</div></div>
        </div>
        <div class="grupo-prog ${grupoDone?'done':'pend'}">${dones}/${items.length}</div>
      </div>
      <div class="grupo-items ${collapsed}">${itemsHTML}</div></div>`;
  }).join('');
}
function toggleGrupo(h){h.nextElementSibling.classList.toggle('collapsed');}
async function procesarSKU(raw){
  const sku=raw.trim().toUpperCase();
  if(!sku||sku.length<2)return;
  if(ultimoEscaneado===sku)return;
  ultimoEscaneado=sku;setTimeout(()=>ultimoEscaneado=null,100);
  try{
    const r=await fetch('/api/escanear',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({sku})});
    const data=await r.json();
    if(!data.ok){flashOverlay('error-flash');vibrar([200,80,200]);mostrarFeedback('error',`❌ ${data.msg}`);toast(data.msg,'error');}
    else if(data.tipo==='ya_completo'){mostrarFeedback('warn',`⚠ ${data.nombre||sku} ya completo`);vibrar([80]);}
    else{
      flashOverlay('ok-flash');vibrar([60]);
      const txt=data.tipo==='completo'?`✔ ${data.nombre||sku} — ¡Completo!`:`${data.nombre||sku}  ${data.collected}/${data.req}`;
      mostrarFeedback('ok',txt);
      if(!estado.colecta)estado.colecta={};
      estado.colecta[sku]=data.collected;
      if(data.todo_completo){estado.colecta_completa=true;toast('🎉 ¡Colecta completa!','success');}
      renderLista(true);
      setTimeout(()=>{const el=document.querySelector(`[data-sku="${sku}"]`);if(el)el.scrollIntoView({behavior:'smooth',block:'center'});},100);
    }
  }catch(e){mostrarFeedback('error','❌ Error de conexión');}
}
async function toggleCamara(){if(scannerActivo)detenerCamara();else iniciarCamara();}
async function iniciarCamara(){
  const container=document.getElementById('cam-container'),video=document.getElementById('cam-video'),btn=document.getElementById('btn-cam');
  try{
    streamCam=await navigator.mediaDevices.getUserMedia({video:{facingMode:'environment',width:{ideal:1280},height:{ideal:720}}});
    video.srcObject=streamCam;container.classList.add('open');btn.classList.add('active');btn.textContent='⏹';scannerActivo=true;
    if('BarcodeDetector' in window){
      barcodeDetector=new BarcodeDetector({formats:['code_128','code_39','ean_13','ean_8','qr_code','upc_a','upc_e','itf','codabar']});
      scanLoop=requestAnimationFrame(detectarBarcode);
    }else{mostrarFeedback('warn','⚠ Usá el lector físico o escribí el SKU manualmente.');}
  }catch(e){mostrarFeedback('error','❌ No se pudo acceder a la cámara: '+e.message);}
}
async function detectarBarcode(){
  if(!scannerActivo||!barcodeDetector)return;
  const video=document.getElementById('cam-video');
  if(video.readyState===video.HAVE_ENOUGH_DATA){
    try{const bs=await barcodeDetector.detect(video);if(bs.length>0){procesarSKU(bs[0].rawValue);await new Promise(r=>setTimeout(r,1500));}}catch(e){}
  }
  if(scannerActivo)scanLoop=requestAnimationFrame(detectarBarcode);
}
function detenerCamara(){
  if(streamCam){streamCam.getTracks().forEach(t=>t.stop());streamCam=null;}
  if(scanLoop){cancelAnimationFrame(scanLoop);scanLoop=null;}
  document.getElementById('cam-container').classList.remove('open');
  document.getElementById('btn-cam').classList.remove('active');
  document.getElementById('btn-cam').textContent='📷';
  scannerActivo=false;document.getElementById('sku-input').focus();
}
function mostrarFeedback(tipo,msg){const el=document.getElementById('feedback');el.className=tipo;el.textContent=msg;}
function toast(msg,tipo=''){const el=document.getElementById('toast');el.textContent=msg;el.className='show '+tipo;setTimeout(()=>el.className='',2800);}
function flashOverlay(cls){const el=document.getElementById('flash');el.className=cls;el.style.opacity='1';setTimeout(()=>{el.style.opacity='0';setTimeout(()=>el.className='',300);},150);}
function vibrar(p){if(navigator.vibrate)navigator.vibrate(p);}
</script>
</body>
</html>"""


    def _detener_servidor(self):
        try:
            # Apagar Flask haciendo una request interna
            import urllib.request
            urllib.request.urlopen("http://127.0.0.1:5050/shutdown", timeout=2)
        except Exception:
            pass
        self._servidor_activo = False
        self._servidor_thread = None
        self.btn_servidor.config(text="📱  App Móvil", bg=C["success"], fg="white")
        self.lbl_servidor_ip.config(text="")


if __name__ == "__main__":
    ventana_principal = tk.Tk()
    app = AsistenteDepositoApp(ventana_principal)
    ventana_principal.mainloop()